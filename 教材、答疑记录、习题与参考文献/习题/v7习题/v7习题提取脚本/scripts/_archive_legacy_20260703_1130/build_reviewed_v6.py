from __future__ import annotations

import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(r"D:\守正公司工作区\cams考试\教材、答疑记录、习题与参考文献\习题\v7习题")
INPUT = BASE / "output_2s" / "semantic_aligned_cn_en_reviewed_v5_merged.xlsx"
OUTPUT = BASE / "output_2s" / "semantic_aligned_cn_en_reviewed_v6_fixed.xlsx"

CN_SHEET = "中文题库"
EN_SHEET = "英文题库"
ALIGN_SHEET = "语义对齐"
UNMATCHED_SHEET = "英文未匹配"
RISK_SHEET = "中英对齐风险清单"
CONFLICT_SHEET = "答案解析冲突记录"
MANUAL_SHEET = "需人工审核清单"
SUMMARY_SHEET = "质量复核总览"
ROLLING_FIX_SHEET = "英文滚屏修复记录_v6"
ALIGN_REVIEW_SHEET = "语义对齐复核记录"


def compact(value: Any, limit: int | None = None) -> str:
    text = " ".join(str(value or "").split())
    return text if limit is None else text[:limit]


def norm_answer(value: Any) -> str:
    return "".join(re.findall(r"[A-F]", str(value or "").upper()))


def headers(ws) -> dict[str, int]:
    return {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}


def row_by_qno(ws, qno: int) -> int | None:
    h = headers(ws)
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, h["题号"]).value == qno:
            return row
    return None


def copy_row_style(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def set_hyperlink(cell, value: Any) -> None:
    cell.value = value
    if value and isinstance(value, str) and (value.endswith(".jpg") or value.endswith(".png")):
        cell.hyperlink = value
        cell.style = "Hyperlink"


def update_bank_row(ws, data: dict[str, Any]) -> None:
    h = headers(ws)
    qno = int(data["题号"])
    row = row_by_qno(ws, qno)
    if row is None:
        # Insert before the next larger question number so the sheet remains readable.
        insert_at = ws.max_row + 1
        for probe in range(2, ws.max_row + 1):
            value = ws.cell(probe, h["题号"]).value
            if isinstance(value, int) and value > qno:
                insert_at = probe
                break
        ws.insert_rows(insert_at)
        copy_row_style(ws, max(2, insert_at - 1), insert_at)
        row = insert_at
    for key, value in data.items():
        col = h[key]
        if key in ("题目截图", "解析截图"):
            set_hyperlink(ws.cell(row, col), value)
        else:
            ws.cell(row, col).value = value


def get_bank_row(ws, qno: int) -> dict[str, Any]:
    h = headers(ws)
    row = row_by_qno(ws, qno)
    if row is None:
        raise KeyError(f"{ws.title} Q{qno} not found")
    return {key: ws.cell(row, col).value for key, col in h.items()}


def cn_analysis(wb, qno: int) -> str:
    return compact(get_bank_row(wb[CN_SHEET], qno)["解析文本"])


def write_sheet(wb, title: str, sheet_headers: list[str], rows: list[list[Any]], widths: list[int] | None = None) -> None:
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(sheet_headers)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = widths or []
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1] if col <= len(widths) else 24
    for row_idx in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row_idx, col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def alignment_target(cn_qno: int) -> tuple[int, str]:
    if 1 <= cn_qno <= 81:
        return cn_qno, "同号顺序段"
    if cn_qno == 82:
        return 288, "显式换位：中文82在英文288"
    if 83 <= cn_qno <= 205:
        return cn_qno - 1, "顺序段：英文提前1题"
    if cn_qno == 206:
        return 383, "显式换位：中文206在英文383"
    if 207 <= cn_qno <= 289:
        return cn_qno - 2, "顺序段：英文提前2题"
    if 290 <= cn_qno <= 383:
        return cn_qno - 1, "顺序段：英文提前1题"
    return cn_qno, "同号顺序段"


def answer_status(cn_answer: Any, en_answer: Any) -> tuple[str, str]:
    cn = norm_answer(cn_answer)
    en = norm_answer(en_answer)
    if cn and en and cn == en:
        return cn, "一致"
    if cn and en:
        return "", f"不一致：中文={cn}; 英文={en}"
    if cn or en:
        return cn or en, "单侧识别"
    return "", "未识别"


def count_expected_answers(text: str) -> int | None:
    text = compact(text).lower()
    patterns = [
        (r"选择一项|请选择一项|select one|choose one|单选", 1),
        (r"选择两项|请选择两项|select two|choose two", 2),
        (r"选择三项|请选择三项|select three|choose three", 3),
        (r"选择四项|请选择四项|select four|choose four", 4),
        (r"选择五项|请选择五项|select five|choose five", 5),
    ]
    for pattern, count in patterns:
        if re.search(pattern, text, re.IGNORECASE):
            return count
    return None


def main() -> None:
    wb = load_workbook(INPUT)
    en_ws = wb[EN_SHEET]

    frame = BASE / "output_2s" / "cache" / "frames_2_0s"
    fine = BASE / "output_2s" / "fine_review_frames"

    # The v5 workbook treated EN121/EN272 as deleted cross-screen continuations.
    # v6 restores them after dense frame review, so the old record would be misleading.
    if "英文跨屏合并记录" in wb.sheetnames:
        del wb["英文跨屏合并记录"]

    corrected_rows: dict[int, dict[str, Any]] = {
        75: {
            "题号": 75,
            "分卷": "英文版1",
            "答案": "B",
            "时间": "494-500",
            "帧数": 3,
            "题目文本": (
                "单选 -[AML/CFT Compliance Programs] The Wolfsberg Group's AML Principles on Private Banking: "
                "A Assist financial institutions conducting business in jurisdictions with high data privacy standards in working with industries susceptible to money laundering. "
                "B Advise banks to accept only those clients whose source of funds and beneficial ownership is understood. "
                "C Establish rules for private bankers on how to deal with politically exposed persons (PEPs) and persons residing in high-risk countries. "
                "D Require banks to better manage reputational risk and protect the privacy of wealthy clients."
            ),
            "解析文本": cn_analysis(wb, 75),
            "来源视频": "英文版1.mp4",
            "题目截图": str(frame / "英文版1" / "英文版1_000249.jpg"),
            "解析截图": str(frame / "英文版1" / "英文版1_000250.jpg"),
        },
        121: {
            "题号": 121,
            "分卷": "英文版2",
            "答案": "A",
            "时间": "127.25-129.75",
            "帧数": "0.25s抽帧复核",
            "题目文本": (
                "单选 -[Sanctions Compliance and Screening] Which of the following best describes the use of fuzzy logic in customer screening systems? "
                "A It produces outputs that include a range of intermediate possibilities between \"Yes\" and \"No\". "
                "B It is an advanced analytics tool widely used to implement AFC controls. "
                "C It allows for a greater number of exact matches, reducing the need for manual review. "
                "D It is a new technique for enhancing the quality of alerts for review."
            ),
            "解析文本": cn_analysis(wb, 122),
            "来源视频": "英文版2.mp4",
            "题目截图": str(fine / "en121_dense" / "en121_025.jpg"),
            "解析截图": str(fine / "en121_dense" / "en121_029.jpg"),
        },
        238: {
            "题号": 238,
            "分卷": "英文版3",
            "答案": "AC",
            "时间": "226-228",
            "帧数": 2,
            "题目文本": (
                "多选 -[AML/CFT Compliance Programs] Perpetual KYC includes: (Choose two.) "
                "A triggers identifying static data changes and data based on client's behaviors in real time. "
                "B periodic refreshing at one-, three- and five-year cycles. "
                "C ability to prioritize better based on client data rather than driven by schedules alone. "
                "D classification of clients into categories of risk."
            ),
            "解析文本": cn_analysis(wb, 240),
            "来源视频": "英文版3.mp4",
            "题目截图": str(frame / "英文版3" / "英文版3_000114.jpg"),
            "解析截图": str(frame / "英文版3" / "英文版3_000115.jpg"),
        },
        247: {
            "题号": 247,
            "分卷": "英文版3",
            "答案": "ABD",
            "时间": "294-296",
            "帧数": 2,
            "题目文本": (
                "多选 -[Risks and Methods of Money Laundering and Terrorist Financing] Which attributes are typically used to assess the vulnerability to money laundering risk of products offered by an insurer? (Choose three.) "
                "A Purpose and intended use. "
                "B Liquidity. "
                "C Age of the beneficiary. "
                "D Customer anonymity or third-party transactions. "
                "E Length of waiting period."
            ),
            "解析文本": cn_analysis(wb, 249),
            "来源视频": "英文版3.mp4",
            "题目截图": str(frame / "英文版3" / "英文版3_000148.jpg"),
            "解析截图": str(frame / "英文版3" / "英文版3_000149.jpg"),
        },
        253: {
            "题号": 253,
            "分卷": "英文版3",
            "答案": "BD",
            "时间": "326-332",
            "帧数": 4,
            "题目文本": (
                "多选 -[Risks and Methods of Money Laundering and Terrorist Financing] In a large US bank, an individual leads a team in charge of overseeing the governance and effectiveness of the bank's transaction monitoring approach. Which strategies should the team implement? (Select Two.) "
                "A Periodic and ad hoc cooperation with the legal team to appropriately investigate and monitor the transactions of subjects of subpoenas or government inquiries. "
                "B Periodic review of client profiles to ensure that the most up-to-date information is on file for high-risk clients in line with the bank's internal policies and procedures. "
                "C Periodic review of suspicious activity reports (SARs) filed with FinCEN to determine whether any should be withdrawn. "
                "D Periodic review of the transaction monitoring scenarios and their productivity to ensure that appropriate AML typologies are reflected."
            ),
            "解析文本": cn_analysis(wb, 255),
            "来源视频": "英文版3.mp4",
            "题目截图": str(frame / "英文版3" / "英文版3_000164.jpg"),
            "解析截图": str(frame / "英文版3" / "英文版3_000166.jpg"),
        },
        272: {
            "题号": 272,
            "分卷": "英文版3",
            "答案": "BD",
            "时间": "449.25-452.75",
            "帧数": "0.25s抽帧复核",
            "题目文本": (
                "多选 -[Risks and Methods of Money Laundering and Terrorist Financing] According to the Financial Action Task Force (FATF) Recommendation 22, when involved in customer transactions that include the buying and selling of real estate, real estate professionals are required to apply which customer due diligence (CDD) measures? (Select Two.) "
                "A Obtaining prior senior manager approval for all related transactions. "
                "B Understanding the purpose of the business relationship. "
                "C Performing adverse media searches on the customer. "
                "D Identifying and verifying the customer's identity."
            ),
            "解析文本": cn_analysis(wb, 274),
            "来源视频": "英文版3.mp4",
            "题目截图": str(fine / "en272_dense" / "en272_022.jpg"),
            "解析截图": str(fine / "en272_dense" / "en272_034.jpg"),
        },
        341: {
            "题号": 341,
            "分卷": "英文版4",
            "答案": "ADEF",
            "时间": "215.25-216.75",
            "帧数": "0.25s抽帧复核",
            "题目文本": (
                "多选 -[AML/CFT Compliance Programs] Challenges in the implementation of new technologies for AML/CFT include: (Select Four.) "
                "A data privacy. "
                "B enhanced due diligence (EDD) policies. "
                "C the Travel Rule. "
                "D data quality. "
                "E complexity. "
                "F regulatory."
            ),
            "解析文本": cn_analysis(wb, 342),
            "来源视频": "英文版4.mp4",
            "题目截图": str(fine / "en341_dense" / "en341_dense_022.jpg"),
            "解析截图": str(fine / "en341_dense" / "en341_dense_028.jpg"),
        },
        346: {
            "题号": 346,
            "分卷": "英文版4",
            "答案": "DE",
            "时间": "243.00-244.75",
            "帧数": "0.25s抽帧复核",
            "题目文本": (
                "多选 -[AML/CFT Compliance Programs] Which factors should be prioritized when choosing an anti-financial crime (AFC) tool for an organization? (Select Two.) "
                "A Real-time data analysis capabilities as part of a broader strategy. "
                "B Preference for the lowest-cost solution. "
                "C Complete elimination of manual processes. "
                "D Compatibility with existing IT infrastructure. "
                "E Scalability to handle increasing transaction volumes."
            ),
            "解析文本": cn_analysis(wb, 347),
            "来源视频": "英文版4.mp4",
            "题目截图": str(fine / "en346_dense" / "en346_dense_021.jpg"),
            "解析截图": str(fine / "en346_dense" / "en346_dense_028.jpg"),
        },
        359: {
            "题号": 359,
            "分卷": "英文版4",
            "答案": "BD",
            "时间": "318.00-322.75",
            "帧数": "0.25s抽帧复核",
            "题目文本": (
                "多选 -[Compliance Standards for AML and CFT] Public-private partnerships (PPPs) that involve the sharing of information between law enforcement authorities, Financial Intelligence Units (FIUs), and the private sector are established to: (Choose two.) "
                "A create a common database of key information and share analysis of suspicious activities with FATF. "
                "B exchange strategic information between FIUs and obliged entities. "
                "C exchange strategic information between financial institutions. "
                "D exchange operational information between public authorities and obliged entities."
            ),
            "解析文本": cn_analysis(wb, 360),
            "来源视频": "英文版4.mp4",
            "题目截图": str(fine / "en359_dense" / "en359_dense_001.jpg"),
            "解析截图": str(fine / "en359_dense" / "en359_dense_014.jpg"),
        },
    }

    fix_rows: list[list[Any]] = []
    for qno, data in corrected_rows.items():
        old_row = row_by_qno(en_ws, qno)
        old_summary = ""
        old_answer = ""
        if old_row:
            h = headers(en_ws)
            old_answer = en_ws.cell(old_row, h["答案"]).value
            old_summary = compact(en_ws.cell(old_row, h["题目文本"]).value, 180)
        update_bank_row(en_ws, data)
        fix_rows.append(
            [
                qno,
                "恢复缺失题" if old_row is None else "修复滚屏污染",
                old_answer,
                data["答案"],
                old_summary,
                compact(data["题目文本"], 220),
                data["题目截图"],
                data["解析截图"],
                "相邻滚动帧被误识别为下一题；按视频帧复核恢复真实题面/答案/解析。",
            ]
        )

    # Ensure English sheet is sorted by question number after inserted rows.
    h = headers(en_ws)
    data_rows = [[en_ws.cell(r, c).value for c in range(1, en_ws.max_column + 1)] for r in range(2, en_ws.max_row + 1)]
    data_rows.sort(key=lambda row: int(row[h["题号"] - 1]))
    for idx, row_values in enumerate(data_rows, start=2):
        for col, value in enumerate(row_values, start=1):
            en_ws.cell(idx, col).value = value
            if en_ws.cell(1, col).value in ("题目截图", "解析截图"):
                set_hyperlink(en_ws.cell(idx, col), value)

    cn_ws = wb[CN_SHEET]
    en_ws = wb[EN_SHEET]
    cn_rows = {qno: get_bank_row(cn_ws, qno) for qno in range(1, 396)}
    en_rows = {qno: get_bank_row(en_ws, qno) for qno in range(1, 396)}

    align_rows: list[list[Any]] = []
    risk_rows: list[list[Any]] = []
    used_en: set[int] = set()
    for cn_qno in range(1, 396):
        en_qno, method = alignment_target(cn_qno)
        used_en.add(en_qno)
        cn_row = cn_rows[cn_qno]
        en_row = en_rows[en_qno]
        suggested, status = answer_status(cn_row["答案"], en_row["答案"])
        diff = abs(cn_qno - en_qno)
        confidence = "高"
        reasons: list[str] = []
        if "显式换位" in method:
            confidence = "中"
            reasons.append("显式换位")
        if status != "一致":
            confidence = "中" if confidence == "高" else confidence
            reasons.append("答案状态不一致")
        expected_cn = count_expected_answers(str(cn_row["题目文本"]))
        expected_en = count_expected_answers(str(en_row["题目文本"]))
        if expected_cn and len(norm_answer(cn_row["答案"])) not in (0, expected_cn):
            reasons.append(f"中文答案数量与题干不符：题干{expected_cn}项/答案{cn_row['答案']}")
        if expected_en and len(norm_answer(en_row["答案"])) not in (0, expected_en):
            reasons.append(f"英文答案数量与题干不符：题干{expected_en}项/答案{en_row['答案']}")
        if (cn_qno, en_qno) in {(206, 383), (342, 341)}:
            reasons.append("源材料答案/解析或中英答案需人工判定")
            confidence = "低"
        if diff > 20:
            reasons.append("题号差>20但已由换位规则定位")
        align_rows.append(
            [
                confidence,
                method,
                suggested,
                status,
                cn_qno,
                en_qno,
                cn_row["答案"],
                en_row["答案"],
                compact(cn_row["题目文本"], 900),
                compact(en_row["题目文本"], 900),
                compact(cn_row["解析文本"], 900),
                compact(en_row["解析文本"], 900),
                cn_row["来源视频"],
                en_row["来源视频"],
                cn_row["时间"],
                en_row["时间"],
                cn_row["题目截图"],
                en_row["题目截图"],
            ]
        )
        if reasons:
            risk_rows.append(
                [
                    cn_qno,
                    en_qno,
                    confidence,
                    status,
                    "；".join(dict.fromkeys(reasons)),
                    compact(cn_row["题目文本"], 220),
                    compact(en_row["题目文本"], 220),
                    cn_row["答案"],
                    en_row["答案"],
                    cn_row["题目截图"],
                    en_row["题目截图"],
                ]
            )

    unmatched = []
    for en_qno in range(1, 396):
        if en_qno not in used_en:
            row = en_rows[en_qno]
            unmatched.append([en_qno, row["答案"], compact(row["题目文本"], 260), row["来源视频"], row["时间"], row["题目截图"]])

    conflict_rows = [
        ["中文题库", 4, "D", "解析指向C", "答案/解析不一致", "保留视频答案字段；需人工判定", cn_rows[4]["题目截图"], "答案帧显示D，但解析文字指向C。"],
        ["英文题库", 322, "ABCDE", "解析指向AD", "答案/解析不一致", "保留视频答案字段；需人工判定", en_rows[322]["题目截图"], "题干为Select Five，答案帧为ABCDE，解析文字指向AD。"],
        ["中文题库", 206, "A", "题干写请选择三项", "题干/答案数量不一致", "保留视频答案字段；需人工判定", cn_rows[206]["题目截图"], "中文视频截图显示“请选择三项”，但正确答案只显示A；英文同题为ACD。"],
        ["英文题库", 341, "ADEF", "解析更接近ACDE", "答案/解析不一致", "保留英文视频答案字段；需人工判定", en_rows[341]["题目截图"], "英文视频勾选ADEF；解析文字讨论Travel Rule(C)、数据质量(D)、复杂性(E)，与答案字段存在冲突。"],
    ]

    manual_rows = [
        ["中文题库", 4, "答案/解析冲突", "答案字段D，解析指向C。", cn_rows[4]["题目截图"]],
        ["英文题库", 322, "答案/解析冲突", "题干Select Five且答案ABCDE，解析指向AD。", en_rows[322]["题目截图"]],
        ["中文题库", 206, "题干/答案数量冲突", "题干要求选择三项，但视频答案只显示A；英文同题为ACD。", cn_rows[206]["题目截图"]],
        ["英文题库", 341, "答案/解析冲突", "视频答案ADEF，解析更接近ACDE。", en_rows[341]["题目截图"]],
    ]

    write_sheet(
        wb,
        ALIGN_SHEET,
        [
            "匹配置信",
            "匹配方法",
            "推荐答案",
            "答案状态",
            "中文题号",
            "英文题号",
            "中文答案",
            "英文答案",
            "中文题目",
            "英文题目",
            "中文解析",
            "英文解析",
            "中文来源",
            "英文来源",
            "中文时间",
            "英文时间",
            "中文题目截图",
            "英文题目截图",
        ],
        align_rows,
        [10, 24, 12, 22, 10, 10, 10, 10, 55, 55, 65, 65, 18, 18, 16, 16, 38, 38],
    )
    write_sheet(wb, UNMATCHED_SHEET, ["英文题号", "答案", "题目文本", "来源视频", "时间", "题目截图"], unmatched, [12, 10, 90, 18, 16, 38])
    write_sheet(
        wb,
        RISK_SHEET,
        ["中文题号", "英文题号", "置信", "答案状态", "原因", "中文题干摘要", "英文题干摘要", "中文答案", "英文答案", "中文截图", "英文截图"],
        risk_rows,
        [10, 10, 8, 24, 42, 55, 55, 10, 10, 38, 38],
    )
    write_sheet(wb, CONFLICT_SHEET, ["工作表", "题号", "答案字段", "解析中指向", "冲突类型", "处理方式", "证据截图", "说明"], conflict_rows, [12, 10, 14, 18, 20, 28, 42, 70])
    write_sheet(wb, MANUAL_SHEET, ["工作表", "题号", "问题类型", "说明", "证据帧/截图"], manual_rows, [12, 10, 24, 80, 42])
    write_sheet(
        wb,
        ROLLING_FIX_SHEET,
        ["英文题号", "处理类型", "旧答案", "新答案", "旧题干摘要", "新题干摘要", "题目证据截图", "解析证据截图", "处理说明"],
        fix_rows,
        [10, 16, 10, 10, 55, 65, 42, 42, 60],
    )
    write_sheet(
        wb,
        ALIGN_REVIEW_SHEET,
        ["项目", "结论"],
        [
            ["英文滚屏污染修复", "修复/恢复 9 个英文题号：75、121、238、247、253、272、341、346、359。"],
            ["英文题库完整性", "修复后英文题库为 395 行，题号 1-395 连续，无缺号。"],
            ["中文题库完整性", "中文题库为 395 行，题号 1-395 连续。"],
            ["对齐规则", "1-81同号；82->288；83-205减1；206->383；207-289减2；290-383减1；384-395同号。"],
            ["未匹配英文", f"{len(unmatched)} 条。"],
            ["中英答案风险", f"{len(risk_rows)} 条进入“中英对齐风险清单”，主要是中英文答案字母不一致或答案数量异常；题干语义仍按规则对齐。"],
            ["强制人工判定", "4 条源材料内部或强冲突：中文4、英文322、中文206、英文341。"],
        ],
        [26, 100],
    )
    write_sheet(
        wb,
        SUMMARY_SHEET,
        ["类别", "范围", "结论"],
        [
            ["中文题库", "1-395", "395行；题号连续；保留源视频答案；中文4、中文206需人工判定。"],
            ["英文题库", "1-395", "395行；题号连续；已修复滚屏污染/缺失题；英文322、英文341需人工判定。"],
            ["中英对齐", "1-395", "按顺序段和2个显式换位重建；英文未匹配0条。"],
            ["中英答案风险", "32条", "集中在中英文答案字母不一致或题干要求项数与答案数量不一致；见“中英对齐风险清单”。"],
            ["滚屏修复", "英文", "修复9条：75、121、238、247、253、272、341、346、359。"],
            ["人工审核", "源冲突", "4条：中文4、英文322、中文206、英文341。"],
        ],
        [18, 18, 100],
    )

    wb.save(OUTPUT)
    print(OUTPUT)
    print(f"fix_rows={len(fix_rows)} align_rows={len(align_rows)} risk_rows={len(risk_rows)} unmatched={len(unmatched)} conflicts={len(conflict_rows)}")


if __name__ == "__main__":
    main()
