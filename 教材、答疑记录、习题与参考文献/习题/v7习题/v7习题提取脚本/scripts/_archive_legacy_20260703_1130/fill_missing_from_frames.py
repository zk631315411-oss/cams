from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ANSWER_RE = re.compile(r"^[A-E]{1,5}$")


def compact(value) -> str:
    return " ".join(str(value or "").split())


def parse_time(value) -> tuple[float | None, float | None]:
    text = str(value or "")
    nums = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", text)]
    if not nums:
        return None, None
    if len(nums) == 1:
        return nums[0], nums[0]
    return nums[0], nums[1]


def video_to_json_name(video: str) -> str:
    return Path(str(video)).stem + "_frames.json"


def load_frames(json_dir: Path) -> dict[str, list[dict]]:
    frames_by_video = {}
    for path in json_dir.glob("*_frames.json"):
        frames_by_video[path.name] = json.loads(path.read_text(encoding="utf-8"))
    return frames_by_video


def nearby_frames(frames: list[dict], start: float | None, end: float | None, before=2, after=18) -> list[dict]:
    if start is None or end is None:
        return frames
    lo = start - before
    hi = end + after
    return [f for f in frames if lo <= float(f.get("time_sec") or 0) <= hi]


def candidate_answer(frames: list[dict], qno: int, start: float | None, end: float | None) -> tuple[str | None, str]:
    q = str(qno)
    windows = [
        [f for f in nearby_frames(frames, start, end, 2, 14) if str(f.get("page_no") or "") == q],
        nearby_frames(frames, start, end, 2, 10),
        [f for f in frames if str(f.get("page_no") or "") == q],
    ]
    for window in windows:
        answers = [str(f.get("answer") or "").strip().upper() for f in window]
        answers = [a for a in answers if ANSWER_RE.match(a)]
        if not answers:
            continue
        counts = Counter(answers)
        answer, count = counts.most_common(1)[0]
        # Accept a unique winner, or a repeated winner over one-off transition noise.
        if len(counts) == 1 or count >= 2:
            return answer, f"answer_candidates={dict(counts)}"
    return None, "no_unique_answer"


def clean_analysis(text: str) -> str:
    if not text:
        return ""
    idx = text.find("试题详解")
    if idx < 0:
        idx = text.find("原解析")
    if idx >= 0:
        text = text[idx:]
    # Drop common UI/source tail only when present near the end.
    for marker in ["考友笔记", "来 源", "来源 2026", "2026新真题"]:
        pos = text.find(marker)
        if pos > 80:
            text = text[:pos]
            break
    return compact(text)


def candidate_analysis(frames: list[dict], qno: int, start: float | None, end: float | None) -> tuple[str | None, str]:
    q = str(qno)
    windows = [
        nearby_frames(frames, start, end, 0, 24),
        [f for f in frames if str(f.get("page_no") or "") == q],
    ]
    best = ""
    best_meta = ""
    for window in windows:
        for f in window:
            text = str(f.get("clean_text") or "")
            if "试题详解" not in text and "原解析" not in text:
                continue
            page_no = str(f.get("page_no") or "")
            if page_no and page_no != q:
                # Keep near-time transition frames only if they have no conflicting page number.
                continue
            cleaned = clean_analysis(text)
            if len(cleaned) > len(best):
                best = cleaned
                best_meta = f"frame={f.get('index')} time={f.get('time_sec')} page={page_no}"
    if len(best) >= 40:
        return best, best_meta
    return None, "no_analysis"


def sync_alignment(ws, lang: str, qno: int, answer, analysis, evidence) -> int:
    is_cn = lang == "CN"
    q_col = 9 if is_cn else 10
    answer_col = 11 if is_cn else 12
    analysis_col = 15 if is_cn else 16
    shot_col = 21 if is_cn else 22
    count = 0
    for row in range(2, ws.max_row + 1):
        try:
            if int(ws.cell(row, q_col).value) != qno:
                continue
        except (TypeError, ValueError):
            continue
        if answer:
            ws.cell(row, answer_col).value = answer
        if analysis is not None:
            ws.cell(row, analysis_col).value = analysis
        if evidence:
            ws.cell(row, shot_col).value = evidence
            ws.cell(row, shot_col).hyperlink = evidence
            ws.cell(row, shot_col).style = "Hyperlink"
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
        count += 1
    return count


def write_log(wb, rows: list[list]) -> None:
    title = "空答案解析回填记录"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    headers = ["语言", "题号", "行号", "字段", "旧值", "新值摘要", "证据", "方法", "状态", "同步行数"]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [10, 10, 10, 12, 16, 56, 72, 46, 12, 12]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
        if ws.cell(row, 7).value:
            ws.cell(row, 7).hyperlink = ws.cell(row, 7).value
            ws.cell(row, 7).style = "Hyperlink"
    ws.freeze_panes = "A2"


def main() -> None:
    out_dir = Path(sys.argv[1])
    source = out_dir / "semantic_aligned_cn_en_single_language_fixed.xlsx"
    target = out_dir / "semantic_aligned_cn_en_filled_missing.xlsx"
    frames_by_video = load_frames(out_dir / "json")

    wb = load_workbook(source)
    align_ws = wb.worksheets[2]
    sheets = [("CN", wb.worksheets[0]), ("EN", wb.worksheets[1])]
    log_rows: list[list] = []
    unresolved = defaultdict(int)

    for lang, ws in sheets:
        for row in range(2, ws.max_row + 1):
            qno = ws.cell(row, 1).value
            try:
                qno = int(qno)
            except (TypeError, ValueError):
                continue
            video = str(ws.cell(row, 8).value or "")
            frames = frames_by_video.get(video_to_json_name(video), [])
            start, end = parse_time(ws.cell(row, 4).value)
            answer = None
            analysis = None
            evidence = ""
            methods = []

            if not ws.cell(row, 3).value:
                answer, method = candidate_answer(frames, qno, start, end)
                methods.append(method)
                if answer:
                    ws.cell(row, 3).value = answer
                else:
                    unresolved[(lang, "answer")] += 1

            if not ws.cell(row, 7).value:
                analysis, method = candidate_analysis(frames, qno, start, end)
                methods.append(method)
                if analysis:
                    ws.cell(row, 7).value = analysis
                else:
                    unresolved[(lang, "analysis")] += 1

            if answer or analysis:
                near = nearby_frames(frames, start, end, 0, 24)
                evidence = next((str(f.get("image") or "") for f in near if str(f.get("page_no") or "") == str(qno)), "")
                if evidence:
                    for col in (9, 10):
                        ws.cell(row, col).value = evidence
                        ws.cell(row, col).hyperlink = evidence
                        ws.cell(row, col).style = "Hyperlink"
                for col in range(1, ws.max_column + 1):
                    ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
                synced = sync_alignment(align_ws, lang, qno, answer, analysis, evidence)
                if answer:
                    log_rows.append([lang, qno, row, "答案", "", answer, evidence, "; ".join(methods), "已回填", synced])
                if analysis:
                    log_rows.append([lang, qno, row, "解析", "", analysis[:180], evidence, "; ".join(methods), "已回填", synced])

    write_log(wb, log_rows)
    wb.save(target)
    print(target)
    print("filled_records", len(log_rows))
    print("unresolved", dict(unresolved))


if __name__ == "__main__":
    main()
