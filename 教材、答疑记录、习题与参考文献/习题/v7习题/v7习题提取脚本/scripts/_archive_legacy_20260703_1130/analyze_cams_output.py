from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


PART_RANGES = {
    1: (1, 100),
    2: (101, 200),
    3: (201, 300),
    4: (301, 395),
}


def page_set(clusters: list[dict]) -> set[int]:
    pages: set[int] = set()
    for cluster in clusters:
        page = str(cluster.get("page_no", "")).strip()
        if page.isdigit():
            pages.add(int(page))
    return pages


def compact_ranges(numbers: list[int]) -> str:
    if not numbers:
        return ""
    ranges: list[str] = []
    start = prev = numbers[0]
    for value in numbers[1:]:
        if value == prev + 1:
            prev = value
            continue
        ranges.append(str(start) if start == prev else f"{start}-{prev}")
        start = prev = value
    ranges.append(str(start) if start == prev else f"{start}-{prev}")
    return ", ".join(ranges)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("output_dir", type=Path)
    args = parser.parse_args()

    json_dir = args.output_dir / "json"
    cn_pages: set[int] = set()
    en_pages: set[int] = set()

    print("Per-video coverage")
    for path in sorted(json_dir.glob("*_clusters.json")):
        data = json.loads(path.read_text(encoding="utf-8"))
        lang = data["lang"]
        part = int(data["part"])
        pages = page_set(data["clusters"])
        if lang == "cn":
            cn_pages.update(pages)
        else:
            en_pages.update(pages)
        start, end = PART_RANGES[part]
        expected = set(range(start, end + 1))
        missing = sorted(expected - pages)
        page_range = f"{min(pages)}-{max(pages)}" if pages else "-"
        print(
            f"{data['video']}: lang={lang} part={part} "
            f"clusters={len(data['clusters'])} pages={len(pages)} "
            f"range={page_range} missing={len(missing)}"
        )
        if missing:
            print(f"  missing: {compact_ranges(missing)}")

    both = cn_pages & en_pages
    union = cn_pages | en_pages
    missing_all = [page for page in range(1, 396) if page not in union]
    cn_only = sorted(cn_pages - en_pages)
    en_only = sorted(en_pages - cn_pages)

    print()
    print("Overall coverage")
    print(f"Chinese pages: {len(cn_pages)}")
    print(f"English pages: {len(en_pages)}")
    print(f"Both languages: {len(both)}")
    print(f"At least one language: {len(union)}")
    print(f"Missing in both: {len(missing_all)}")
    print(f"Chinese only: {len(cn_only)}")
    print(f"English only: {len(en_only)}")
    if missing_all:
        print(f"Missing in both ranges: {compact_ranges(missing_all)}")
    if cn_only:
        print(f"Chinese-only ranges: {compact_ranges(cn_only)}")
    if en_only:
        print(f"English-only ranges: {compact_ranges(en_only)}")

    workbook = args.output_dir / "merged_cn_en.xlsx"
    if workbook.exists():
        wb = load_workbook(workbook, read_only=True, data_only=True)
        print()
        print("Workbook")
        for ws in wb.worksheets:
            print(f"{ws.title}: rows={ws.max_row - 1} columns={ws.max_column}")


if __name__ == "__main__":
    main()
