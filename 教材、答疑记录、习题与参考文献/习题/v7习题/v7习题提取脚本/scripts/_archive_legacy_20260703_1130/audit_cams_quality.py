from __future__ import annotations

import argparse
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


CN_CATEGORY_MAP = {
    "开展和回应调查": "conducting",
    "反洗钱/打击资助恐怖主义合规计划": "programs",
    "反洗钱和反恐怖融资合规标准": "standards",
    "反洗钱和反恐怖融资合规标准": "standards",
    "洗钱和恐怖融资的风险及方法": "risks",
    "洗钱和恐怖融资的风险及手段": "risks",
}

EN_CATEGORY_MAP = {
    "conductingandrespondingtoinvestigations": "conducting",
    "conducting and responding to investigations": "conducting",
    "aml/cftcomplianceprograms": "programs",
    "aml/cft compliance programs": "programs",
    "compliancestandardsforamlandcft": "standards",
    "compliance standards for aml and cft": "standards",
    "risksandmethodsofmoneylaunderingandterroristfinancing": "risks",
    "risks and methods of money laundering and terrorist financing": "risks",
}


def compact(text: object) -> str:
    return re.sub(r"\s+", "", str(text or "")).lower()


def text_len(value: object) -> int:
    return len(str(value or "").strip())


def extract_cn_category(text: str) -> str:
    match = re.search(r"-?\[([^\]]{2,40})\]", text or "")
    if not match:
        return ""
    raw = match.group(1)
    return CN_CATEGORY_MAP.get(raw, raw)


def extract_en_category(text: str) -> str:
    match = re.search(r"-?\[([^\]]{2,80})\]", text or "")
    if not match:
        return ""
    raw = compact(match.group(1))
    return EN_CATEGORY_MAP.get(raw, raw)


def category_from_text(text: object, lang: str) -> str:
    value = str(text or "")
    if lang == "cn":
        found = extract_cn_category(value)
        if found:
            return found
        c = compact(value)
        for key, mapped in CN_CATEGORY_MAP.items():
            if compact(key) in c:
                return mapped
    else:
        found = extract_en_category(value)
        if found:
            return found
        c = compact(value)
        for key, mapped in EN_CATEGORY_MAP.items():
            if compact(key) in c:
                return mapped
    return ""


def flag_row(row: dict[str, object]) -> list[str]:
    flags: list[str] = []
    cn_q = str(row.get("中文题目候选文本") or "")
    en_q = str(row.get("英文题目候选文本") or "")
    cn_e = str(row.get("中文解析候选文本") or "")
    en_e = str(row.get("英文解析候选文本") or "")
    cn_ans = str(row.get("中文答案") or "").strip()
    en_ans = str(row.get("英文答案") or "").strip()

    if not cn_q:
        flags.append("中文题干空")
    if not en_q:
        flags.append("英文题干空")
    if text_len(cn_q) < 80:
        flags.append("中文题干偏短")
    if text_len(en_q) < 120:
        flags.append("英文题干偏短")
    if not cn_e:
        flags.append("中文解析空")
    if not en_e:
        flags.append("英文解析空")
    if text_len(cn_e) < 80:
        flags.append("中文解析偏短")
    if text_len(en_e) < 80:
        flags.append("英文解析偏短")

    cn_q_compact = compact(cn_q)
    en_q_compact = compact(en_q)
    if cn_q_compact.startswith(("正确答案", "试题详解", "原解析")):
        flags.append("中文题干像解析页")
    if en_q_compact.startswith(("正确答案", "试题详解", "原解析")):
        flags.append("英文题干像解析页")

    cn_cat = category_from_text(cn_q, "cn") or category_from_text(cn_e, "cn")
    en_cat = category_from_text(en_q, "en") or category_from_text(en_e, "en")
    if cn_cat and en_cat and cn_cat != en_cat:
        flags.append(f"中英类别疑似不一致:{cn_cat}!={en_cat}")
    elif not cn_cat:
        flags.append("中文类别未识别")
    elif not en_cat:
        flags.append("英文类别未识别")

    if not cn_ans:
        flags.append("中文答案未识别")
    if not en_ans:
        flags.append("英文答案未识别")
    if cn_ans and en_ans and cn_ans != en_ans:
        flags.append("中英答案不一致")

    return flags


def status_bucket(flags: list[str]) -> str:
    hard = {
        "中文题干空",
        "英文题干空",
        "中文题干像解析页",
        "英文题干像解析页",
    }
    if any(flag in hard for flag in flags):
        return "高风险"
    if any("类别疑似不一致" in flag for flag in flags):
        return "高风险"
    if any(flag.endswith("答案未识别") or flag == "中英答案不一致" for flag in flags):
        return "需答案复核"
    if any("偏短" in flag or flag.endswith("空") or flag.endswith("未识别") for flag in flags):
        return "需文本复核"
    return "通过"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    wb_in = load_workbook(args.workbook, read_only=True, data_only=True)
    ws = wb_in["中英对照"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [
        dict(zip(headers, values))
        for values in ws.iter_rows(min_row=2, values_only=True)
    ]

    audited = []
    for row in rows:
        flags = flag_row(row)
        audited.append((row, flags, status_bucket(flags)))

    status_counts = Counter(status for _, _, status in audited)
    flag_counts = Counter(flag for _, flags, _ in audited for flag in flags)

    out = args.output or args.workbook.with_name("quality_audit.xlsx")
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "质量摘要"
    ws_flags = wb.create_sheet("可疑题目")
    ws_sample = wb.create_sheet("抽样题目")

    ws_summary.append(["项目", "数量"])
    ws_summary.append(["总题数", len(rows)])
    for key in ["通过", "需文本复核", "需答案复核", "高风险"]:
        ws_summary.append([key, status_counts.get(key, 0)])
    ws_summary.append(["", ""])
    ws_summary.append(["具体标记", "数量"])
    for flag, count in flag_counts.most_common():
        ws_summary.append([flag, count])

    out_headers = [
        "题号",
        "质量等级",
        "标记",
        "推荐答案",
        "答案核对状态",
        "中文答案",
        "英文答案",
        "中文类别",
        "英文类别",
        "中文时间",
        "英文时间",
        "中文题目候选文本",
        "英文题目候选文本",
        "中文解析候选文本",
        "英文解析候选文本",
        "中文题目截图",
        "英文题目截图",
    ]
    ws_flags.append(out_headers)
    for row, flags, status in audited:
        if status == "通过":
            continue
        cn_cat = category_from_text(row.get("中文题目候选文本"), "cn") or category_from_text(
            row.get("中文解析候选文本"),
            "cn",
        )
        en_cat = category_from_text(row.get("英文题目候选文本"), "en") or category_from_text(
            row.get("英文解析候选文本"),
            "en",
        )
        ws_flags.append(
            [
                row.get("题号"),
                status,
                "; ".join(flags),
                row.get("推荐答案"),
                row.get("答案核对状态"),
                row.get("中文答案"),
                row.get("英文答案"),
                cn_cat,
                en_cat,
                row.get("中文时间"),
                row.get("英文时间"),
                row.get("中文题目候选文本"),
                row.get("英文题目候选文本"),
                row.get("中文解析候选文本"),
                row.get("英文解析候选文本"),
                row.get("中文题目截图"),
                row.get("英文题目截图"),
            ]
        )

    sample_pages = [1, 2, 33, 54, 101, 132, 164, 194, 206, 284, 308, 346, 378, 395]
    ws_sample.append(out_headers)
    for row, flags, status in audited:
        if row.get("题号") not in sample_pages:
            continue
        cn_cat = category_from_text(row.get("中文题目候选文本"), "cn") or category_from_text(
            row.get("中文解析候选文本"),
            "cn",
        )
        en_cat = category_from_text(row.get("英文题目候选文本"), "en") or category_from_text(
            row.get("英文解析候选文本"),
            "en",
        )
        ws_sample.append(
            [
                row.get("题号"),
                status,
                "; ".join(flags),
                row.get("推荐答案"),
                row.get("答案核对状态"),
                row.get("中文答案"),
                row.get("英文答案"),
                cn_cat,
                en_cat,
                row.get("中文时间"),
                row.get("英文时间"),
                row.get("中文题目候选文本"),
                row.get("英文题目候选文本"),
                row.get("中文解析候选文本"),
                row.get("英文解析候选文本"),
                row.get("中文题目截图"),
                row.get("英文题目截图"),
            ]
        )

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        widths = {
            "A": 10,
            "B": 14,
            "C": 45,
            "D": 12,
            "E": 26,
            "F": 10,
            "G": 10,
            "H": 14,
            "I": 14,
            "J": 14,
            "K": 14,
            "L": 60,
            "M": 60,
            "N": 60,
            "O": 60,
            "P": 55,
            "Q": 55,
        }
        for col, width in widths.items():
            sheet.column_dimensions[col].width = width
        for idx in range(2, min(sheet.max_row, 2000) + 1):
            sheet.row_dimensions[idx].height = 70

    wb.save(out)

    print(f"rows={len(rows)}")
    for key in ["通过", "需文本复核", "需答案复核", "高风险"]:
        print(f"{key}={status_counts.get(key, 0)}")
    print(f"output={out}")


if __name__ == "__main__":
    main()
