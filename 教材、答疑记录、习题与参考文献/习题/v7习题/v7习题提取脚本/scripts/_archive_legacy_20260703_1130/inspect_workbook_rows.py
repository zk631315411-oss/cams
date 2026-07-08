from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def shorten(value, limit: int = 240):
    if value is None:
        return ""
    text = str(value).replace("\n", " ")
    text = " ".join(text.split())
    if len(text) <= limit:
        return text
    return text[:limit] + "..."


def rows_by_question(ws, question_numbers: set[int]) -> list[dict]:
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    out = []
    for row_index in range(2, ws.max_row + 1):
        value = ws.cell(row_index, 1).value
        try:
            question_no = int(value)
        except (TypeError, ValueError):
            continue
        if question_no not in question_numbers:
            continue
        row = {"excel_row": row_index}
        for col_index, header in enumerate(headers, start=1):
            row[str(header)] = shorten(ws.cell(row_index, col_index).value)
        out.append(row)
    return out


def main() -> None:
    if len(sys.argv) < 4:
        raise SystemExit("usage: inspect_workbook_rows.py WORKBOOK SHEET_INDEX QNO [QNO...]")
    workbook_path = Path(sys.argv[1])
    sheet_index = int(sys.argv[2])
    qnos = {int(arg) for arg in sys.argv[3:]}
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb.worksheets[sheet_index]
    result = {
        "sheet": ws.title,
        "headers": [ws.cell(1, c).value for c in range(1, ws.max_column + 1)],
        "rows": rows_by_question(ws, qnos),
    }
    print(json.dumps(result, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
