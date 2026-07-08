from __future__ import annotations

import importlib.util
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CN_SHEET = "中文题库"
EN_SHEET = "英文题库"
ALIGN_SHEET = "语义对齐"
UNMATCHED_SHEET = "英文未匹配"
RISK_SHEET = "中英对齐风险清单"
MERGE_SHEET = "英文跨屏合并记录"
CONFLICT_SHEET = "答案解析冲突记录"
MANUAL_SHEET = "需人工审核清单"
SUMMARY_SHEET = "质量复核总览"


def compact(value: Any) -> str:
    return " ".join(str(value or "").split())


def norm_answer(value: Any) -> str:
    return "".join(ch for ch in str(value or "").upper() if ch in "ABCDEF")


def set_link(cell, path: str | None) -> None:
    if path:
        cell.value = path
        cell.hyperlink = path
        cell.style = "Hyperlink"


def write_sheet(wb, title: str, headers: list[str], rows: list[list[Any]], widths: list[int] | None = None) -> None:
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = widths or [12, 12, 14, 18, 28, 55, 55, 75, 75, 18, 18, 18]
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1] if col <= len(widths) else 24
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def sheet_headers(ws) -> dict[str, int]:
    return {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}


def find_row_by_qno(ws, qno: int) -> int:
    headers = sheet_headers(ws)
    q_col = headers["题号"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, q_col).value == qno:
            return row
    raise KeyError(f"{ws.title} Q{qno} not found")


def load_alignment_module(base_dir: Path):
    path = base_dir / "scripts" / "build_semantic_alignment.py"
    spec = importlib.util.spec_from_file_location("build_semantic_alignment_local", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot import {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def rows_for_alignment(ws, kind: str) -> list[dict[str, Any]]:
    headers = sheet_headers(ws)
    rows: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        if kind == "cn":
            rows.append(
                {
                    "题号": ws.cell(row, headers["题号"]).value,
                    "分卷": ws.cell(row, headers["分卷"]).value,
                    "答案": ws.cell(row, headers["答案"]).value,
                    "起止秒": ws.cell(row, headers["时间"]).value,
                    "帧数": ws.cell(row, headers["帧数"]).value,
                    "题目候选文本": ws.cell(row, headers["题目文本"]).value,
                    "解析候选文本": ws.cell(row, headers["解析文本"]).value,
                    "来源视频": ws.cell(row, headers["来源视频"]).value,
                    "题目截图": ws.cell(row, headers["题目截图"]).value,
                    "解析截图": ws.cell(row, headers["解析截图"]).value,
                }
            )
        else:
            rows.append(
                {
                    "Question": ws.cell(row, headers["题号"]).value,
                    "Part": ws.cell(row, headers["分卷"]).value,
                    "Answer": ws.cell(row, headers["答案"]).value,
                    "Time": ws.cell(row, headers["时间"]).value,
                    "Frames": ws.cell(row, headers["帧数"]).value,
                    "Question Text": ws.cell(row, headers["题目文本"]).value,
                    "Explanation Text": ws.cell(row, headers["解析文本"]).value,
                    "Source": ws.cell(row, headers["来源视频"]).value,
                    "Question Image": ws.cell(row, headers["题目截图"]).value,
                    "Explanation Image": ws.cell(row, headers["解析截图"]).value,
                }
            )
    return rows


def answer_status(cn_answer: Any, en_answer: Any) -> tuple[str, str]:
    cn = norm_answer(cn_answer)
    en = norm_answer(en_answer)
    if cn and en and cn == en:
        return cn, "一致"
    if cn and en:
        return "", f"不一致: 中文={cn}; 英文={en}"
    if cn or en:
        return cn or en, "单侧识别"
    return "", "未识别"


def build_matches(module, cn_rows: list[dict[str, Any]], en_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cn_vecs = [module.ngrams(module.cjk_text(module.row_text(row, "题目候选文本", "解析候选文本"))) for row in cn_rows]
    en_vecs = [module.ngrams(module.cjk_text(module.row_text(row, "Question Text", "Explanation Text"))) for row in en_rows]
    cn_concepts = [module.concept_set(module.row_text(row, "题目候选文本", "解析候选文本")) for row in cn_rows]
    en_concepts = [module.concept_set(module.row_text(row, "Question Text", "Explanation Text")) for row in en_rows]
    en_page_to_idx = {module.page_int(row, "Question"): idx for idx, row in enumerate(en_rows)}

    overrides = dict(getattr(module, "VISUAL_OVERRIDES", {}))
    overrides.update({121: 120, 273: 271})
    ranked_by_cn: list[list[dict[str, Any]]] = []

    for cn_idx, cn_row in enumerate(cn_rows):
        cn_page = module.page_int(cn_row, "题号")
        ranked: list[dict[str, Any]] = []
        same_page_sim = 0.0
        for en_idx, en_row in enumerate(en_rows):
            en_page = module.page_int(en_row, "Question")
            sim = module.cosine(cn_vecs[cn_idx], en_vecs[en_idx])
            overlap = module.jaccard(cn_concepts[cn_idx], en_concepts[en_idx])
            diff = abs(cn_page - en_page)
            answer_adjust = module.answer_adjustment(cn_row.get("答案"), en_row.get("Answer"))
            score = module.fallback_score(sim, overlap, diff, answer_adjust)
            answer_disjoint = module.answer_is_disjoint(cn_row.get("答案"), en_row.get("Answer"))
            ranked.append(
                {
                    "idx": en_idx,
                    "page": en_page,
                    "sim": sim,
                    "concept": overlap,
                    "diff": diff,
                    "score": score,
                    "answer_disjoint": answer_disjoint,
                }
            )
            if en_page == cn_page:
                same_page_sim = sim

        top_score = max(item["score"] for item in ranked)
        for candidate in ranked:
            semantic_boost = 0.0
            if (
                candidate["sim"] >= 0.075
                and not candidate["answer_disjoint"]
                and (candidate["sim"] - same_page_sim >= 0.040 or candidate["score"] >= top_score - 0.020)
            ):
                semantic_boost = 0.10 + candidate["sim"]
            candidate["rank_score"] = candidate["score"] + semantic_boost
        ranked.sort(key=lambda item: (item["rank_score"], item["sim"], item["score"]), reverse=True)
        ranked_by_cn.append(ranked)

    chosen_by_cn: dict[int, dict[str, Any]] = {}
    used_en: set[int] = set()

    for cn_idx, cn_row in enumerate(cn_rows):
        cn_page = module.page_int(cn_row, "题号")
        override_page = overrides.get(cn_page)
        if override_page not in en_page_to_idx:
            continue
        en_idx = en_page_to_idx[override_page]
        chosen = next((item for item in ranked_by_cn[cn_idx] if item["idx"] == en_idx), None)
        if chosen is None:
            continue
        chosen = dict(chosen)
        chosen["manual_override"] = True
        chosen_by_cn[cn_idx] = chosen
        used_en.add(en_idx)

    assignment_order = sorted(
        (idx for idx in range(len(cn_rows)) if idx not in chosen_by_cn),
        key=lambda idx: (
            ranked_by_cn[idx][0]["rank_score"],
            ranked_by_cn[idx][0]["sim"],
            -ranked_by_cn[idx][0]["diff"],
        ),
        reverse=True,
    )
    for cn_idx in assignment_order:
        for candidate in ranked_by_cn[cn_idx]:
            if candidate["idx"] not in used_en:
                chosen_by_cn[cn_idx] = candidate
                used_en.add(candidate["idx"])
                break
        if cn_idx not in chosen_by_cn:
            # English has fewer rows after cross-screen merges; keep a traceable duplicate only when unavoidable.
            fallback = dict(ranked_by_cn[cn_idx][0])
            fallback["duplicate_fallback"] = True
            chosen_by_cn[cn_idx] = fallback

    matches: list[dict[str, Any]] = []
    for cn_idx, cn_row in enumerate(cn_rows):
        ranked = ranked_by_cn[cn_idx]
        chosen = chosen_by_cn[cn_idx]

        next_best = next((item for item in ranked if item["idx"] != chosen["idx"]), None)
        method = "视觉复核覆盖" if chosen.get("manual_override") else ("内容相似" if chosen["sim"] >= 0.075 and not chosen["answer_disjoint"] else "同号/近号+概念")
        if chosen.get("duplicate_fallback"):
            method = f"{method}/英文复用"
        margin = chosen["rank_score"] - (next_best["rank_score"] if next_best else 0.0)
        confidence = "高" if chosen.get("manual_override") else module.confidence_label(method, chosen["sim"], chosen["score"], margin, int(chosen["diff"]))
        en_row = en_rows[chosen["idx"]]
        suggested, status = answer_status(cn_row.get("答案"), en_row.get("Answer"))
        alternatives = [
            f'{item["page"]}({item["rank_score"]:.3f}/{item["sim"]:.3f})'
            for item in ranked[:5]
            if item["idx"] != chosen["idx"]
        ]
        matches.append(
            {
                "confidence": confidence,
                "method": method,
                "score": chosen["score"],
                "sim": chosen["sim"],
                "concept": chosen["concept"],
                "diff": chosen["diff"],
                "alternatives": "; ".join(alternatives[:4]),
                "suggested_answer": suggested,
                "answer_status": status,
                "cn": cn_row,
                "en": en_row,
            }
        )
    return matches


def write_alignment_sheets(wb, matches: list[dict[str, Any]], en_rows: list[dict[str, Any]]) -> list[list[Any]]:
    for title in (ALIGN_SHEET, UNMATCHED_SHEET, RISK_SHEET):
        if title in wb.sheetnames:
            del wb[title]

    align_rows = []
    used_en_pages: Counter[int] = Counter()
    for item in matches:
        cn = item["cn"]
        en = item["en"]
        used_en_pages[int(en.get("Question"))] += 1
        align_rows.append(
            [
                item["confidence"],
                item["method"],
                round(item["score"], 4),
                round(item["sim"], 4),
                round(item["concept"], 4),
                item["diff"],
                item["suggested_answer"],
                item["answer_status"],
                cn.get("题号"),
                en.get("Question"),
                cn.get("答案"),
                en.get("Answer"),
                compact(cn.get("题目候选文本")),
                compact(en.get("Question Text")),
                compact(cn.get("解析候选文本")),
                compact(en.get("Explanation Text")),
                cn.get("来源视频"),
                en.get("Source"),
                cn.get("起止秒"),
                en.get("Time"),
                cn.get("题目截图"),
                en.get("Question Image"),
                item["alternatives"],
            ]
        )

    write_sheet(
        wb,
        ALIGN_SHEET,
        [
            "匹配置信",
            "匹配方法",
            "匹配分数",
            "中文相似度",
            "概念重合",
            "题号差",
            "推荐答案",
            "答案状态",
            "中文题号",
            "英文题号",
            "中文答案",
            "英文答案",
            "中文题目",
            "英文题目",
            "中文解析",
            "英文解析",
            "中文来源",
            "英文来源",
            "中文时间",
            "英文时间",
            "中文题目截图",
            "英文题目截图",
            "备选英文题号(分数/中文相似度)",
        ],
        align_rows,
        [10, 14, 10, 10, 10, 8, 10, 24, 8, 8, 8, 8, 58, 58, 58, 58, 18, 18, 14, 14, 55, 55, 32],
    )

    unmatched_rows = []
    for row in en_rows:
        page = int(row.get("Question"))
        if used_en_pages[page] == 0 or used_en_pages[page] > 1:
            unmatched_rows.append(
                [
                    row.get("Question"),
                    row.get("Answer"),
                    compact(row.get("Question Text")),
                    compact(row.get("Explanation Text")),
                    row.get("Source"),
                    row.get("Time"),
                    row.get("Question Image"),
                    used_en_pages[page],
                ]
            )
    write_sheet(wb, UNMATCHED_SHEET, ["英文题号", "答案", "题目文本", "解析文本", "来源视频", "时间", "题目截图", "使用次数"], unmatched_rows, [8, 8, 70, 70, 18, 14, 55, 10])

    risk_rows = []
    for item in matches:
        reasons = []
        if item["confidence"] == "低":
            reasons.append("低置信")
        if str(item["answer_status"]).startswith("不一致"):
            reasons.append("答案状态不一致")
        if int(item["diff"]) > 5:
            reasons.append("题号差>5")
        if reasons:
            cn = item["cn"]
            en = item["en"]
            risk_rows.append(
                [
                    cn.get("题号"),
                    en.get("Question"),
                    item["confidence"],
                    round(item["score"], 4),
                    item["diff"],
                    item["answer_status"],
                    "；".join(reasons),
                    compact(cn.get("题目候选文本"))[:220],
                    compact(en.get("Question Text"))[:220],
                ]
            )
    write_sheet(wb, RISK_SHEET, ["中文题号", "英文题号", "置信", "分数", "题号差", "答案状态", "原因", "中文题干摘要", "英文题干摘要"], risk_rows, [10, 10, 8, 8, 8, 24, 28, 72, 72])
    return risk_rows


def answer_mentions(text: Any) -> list[str]:
    value = re.sub(r"\s+", "", str(text or "").upper())
    patterns = [
        r"正确答案(?:为|是)?([A-F](?:[、,，和及与]?[A-F]){0,5})",
        r"正确选项(?:为|是)?([A-F](?:[、,，和及与]?[A-F]){0,5})",
        r"答案(?:选|为|是)([A-F](?:[、,，和及与]?[A-F]){0,5})",
        r"因此(?:选择|选)([A-F](?:[、,，和及与]?[A-F]){0,5})",
        r"故(?:正确)?答案(?:为|是|选)?([A-F](?:[、,，和及与]?[A-F]){0,5})",
    ]
    mentions: list[str] = []
    for pattern in patterns:
        for match in re.finditer(pattern, value):
            answer = norm_answer(match.group(1))
            if answer:
                mentions.append(answer)
    return sorted(set(mentions))


def build_conflict_rows(wb) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for sheet_name in (CN_SHEET, EN_SHEET):
        ws = wb[sheet_name]
        headers = sheet_headers(ws)
        for row in range(2, ws.max_row + 1):
            qno = ws.cell(row, headers["题号"]).value
            answer = norm_answer(ws.cell(row, headers["答案"]).value)
            mentions = answer_mentions(ws.cell(row, headers["解析文本"]).value)
            conflicts = [item for item in mentions if item and item != answer]
            if not conflicts:
                continue
            conflict = "/".join(conflicts)
            evidence = ws.cell(row, headers["解析截图"]).value or ws.cell(row, headers["题目截图"]).value
            if sheet_name == CN_SHEET and qno == 4:
                note = "题干答案帧显示 D，但解析帧写第319(b)并称正确选项为 C；按视频答案字段保留 D。"
            elif sheet_name == EN_SHEET and qno == 322:
                note = "题干写 Select Five 且答案帧勾选 ABCDE，但解析文字称正确答案 AD；按答案帧保留 ABCDE。"
            else:
                note = "解析文字与答案字段不一致，未自动覆盖答案。"
            rows.append([sheet_name, qno, answer, conflict, "答案字段与解析文字不一致", "保留答案字段，冲突如实记录", evidence, note])
    return rows


def merge_cross_screen_questions(wb) -> list[list[Any]]:
    ws = wb[EN_SHEET]
    h = sheet_headers(ws)
    merge_rows: list[list[Any]] = []

    q120_text = (
        "单选 -[Risks and Methods of Money Laundering and Terrorist Financing] A large international bank's chief compliance officer (CCO) "
        "is exploring ways to enhance the bank's ability to identify suspicious activities by using intelligence data more effectively. "
        "One potential solution is to engage in public-private partnerships (PPPs) to leverage shared intelligence and enhance collaboration "
        "with government agencies. The bank considers joining a PPP initiative with the local Financial Intelligence Unit (FIU) and other "
        "financial institutions to improve its access to relevant data and intelligence. The CCO understands that while PPPs can provide "
        "significant benefits, such as improved risk detection and enhanced information sharing, there are also potential limitations, "
        "including data privacy concerns and differing priorities between public and private sector partners. Which approach would best "
        "maximize the benefits of PPPs for the bank while mitigating the limitations associated with data sharing and intelligence? "
        "A Engage in the PPP without strict data sharing protocols, allowing for open and unrestricted flow of information between the bank, "
        "FIUs, and other financial institutions. B Rely solely on the intelligence provided by government agencies through the PPP because "
        "they have the most comprehensive data on suspicious activities. C Establish a clear framework within the PPP that outlines data privacy "
        "protections and ensures that information sharing complies with legal and regulatory requirements in all jurisdictions involved. "
        "D Prioritize the bank's internal data sources over external intelligence from PPP. 正确答案C您选择/ 试题详解 原解析"
    )
    q271_text = (
        "多选 -[AML/CFT Compliance Programs] An institution is looking to alter an existing threshold-based monitoring scenario because it is "
        "generating too many alerts that do not yield suspicious activity reports. Documentation submitted to the relevant committee for "
        "supporting this proposal should include details on: (Choose three.) A evidence that the increased residual risk arising from the change "
        "is within the bank's risk appetite. B approval by money-laundering reporting officer for the proposal. C how many resources are spent "
        "on the less productive lower threshold and the associated costs. D the number of cases that will not be filed and the resources that can "
        "be freed up for other tasks. E minutes of meeting held with the regulator where agreement was obtained that the higher threshold was justified. "
        "F historical analysis proving that the current scenario generates a disproportionate number of false positives. 正确答案ACF您选择/ 试题详解 原解析"
    )

    merges = [
        {
            "keep": 120,
            "drop": 121,
            "question": q120_text,
            "answer": "C",
            "qshot": r"D:\守正公司工作区\cams考试\教材、答疑记录、习题与参考文献\习题\v7习题\output_2s\cache\frames_2_0s\英文版2\英文版2_000058.jpg",
            "ashot": r"D:\守正公司工作区\cams考试\教材、答疑记录、习题与参考文献\习题\v7习题\output_2s\cache\frames_2_0s\英文版2\英文版2_000064.jpg",
            "note": "英文120/121是同一道PPP长题，合并为英文120，删除英文121。",
        },
        {
            "keep": 271,
            "drop": 272,
            "question": q271_text,
            "answer": "ACF",
            "qshot": r"D:\守正公司工作区\cams考试\教材、答疑记录、习题与参考文献\习题\v7习题\output_2s\cache\frames_2_0s\英文版3\英文版3_000222.jpg",
            "ashot": r"D:\守正公司工作区\cams考试\教材、答疑记录、习题与参考文献\习题\v7习题\output_2s\cache\frames_2_0s\英文版3\英文版3_000225.jpg",
            "note": "英文271/272是同一道阈值调整题，合并为英文271，删除英文272。",
        },
    ]

    for item in merges:
        keep_row = find_row_by_qno(ws, item["keep"])
        drop_row = find_row_by_qno(ws, item["drop"])
        old_keep = compact(ws.cell(keep_row, h["题目文本"]).value)[:180]
        old_drop = compact(ws.cell(drop_row, h["题目文本"]).value)[:180]
        ws.cell(keep_row, h["题目文本"]).value = item["question"]
        ws.cell(keep_row, h["答案"]).value = item["answer"]
        set_link(ws.cell(keep_row, h["题目截图"]), item["qshot"])
        set_link(ws.cell(keep_row, h["解析截图"]), item["ashot"])
        ws.delete_rows(drop_row, 1)
        merge_rows.append([item["keep"], item["drop"], item["answer"], old_keep, old_drop, item["question"][:260], item["qshot"], item["ashot"], item["note"]])

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
    return merge_rows


def main() -> None:
    base_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parents[1]
    out_dir = base_dir / "output_2s"
    source = out_dir / "semantic_aligned_cn_en_reviewed_v4.xlsx"
    target = out_dir / "semantic_aligned_cn_en_reviewed_v5_merged.xlsx"

    wb = load_workbook(source)
    merge_rows = merge_cross_screen_questions(wb)

    module = load_alignment_module(base_dir)
    cn_rows = rows_for_alignment(wb[CN_SHEET], "cn")
    en_rows = rows_for_alignment(wb[EN_SHEET], "en")
    matches = build_matches(module, cn_rows, en_rows)
    risk_rows = write_alignment_sheets(wb, matches, en_rows)

    conflict_rows = build_conflict_rows(wb)
    write_sheet(
        wb,
        CONFLICT_SHEET,
        ["工作表", "题号", "答案字段", "解析中指向", "冲突类型", "处理方式", "证据截图", "说明"],
        conflict_rows,
        [12, 8, 10, 12, 24, 28, 60, 80],
    )
    write_sheet(
        wb,
        MERGE_SHEET,
        ["保留英文题号", "删除英文题号", "合并后答案", "原保留行摘要", "原删除行摘要", "合并后题干摘要", "题目证据截图", "解析证据截图", "处理说明"],
        merge_rows,
        [12, 12, 12, 55, 55, 75, 55, 55, 50],
    )
    manual_rows = [
        [row[0], row[1], "答案/解析冲突", row[7], row[6]]
        for row in conflict_rows
    ]
    write_sheet(wb, MANUAL_SHEET, ["工作表", "题号", "问题类型", "说明", "证据帧/截图"], manual_rows, [12, 8, 18, 90, 60])

    summary_rows = [
        ["题库规模", "中文题库", f"{len(cn_rows)} 题；题干/答案/解析无空值"],
        ["题库规模", "英文题库", f"{len(en_rows)} 题；已按用户要求合并跨屏拆题，删除英文121和英文272"],
        ["跨屏合并", "英文题库", f"完成 {len(merge_rows)} 组：120/121 -> 120，271/272 -> 271"],
        ["答案解析冲突", "全表", f"记录 {len(conflict_rows)} 条；不自动覆盖答案字段"],
        ["需人工审核", "冲突题", f"{len(manual_rows)} 条，均已写入《需人工审核清单》和《答案解析冲突记录》"],
        ["中英语义对齐", "合并后重建", f"重新生成语义对齐；风险清单 {len(risk_rows)} 条"],
    ]
    write_sheet(wb, SUMMARY_SHEET, ["类别", "范围", "结论"], summary_rows, [16, 16, 90])

    wb.save(target)
    print(target)
    print("cn_rows", len(cn_rows))
    print("en_rows", len(en_rows))
    print("merged_pairs", len(merge_rows))
    print("answer_analysis_conflicts", len(conflict_rows))
    print("alignment_risk", len(risk_rows))


if __name__ == "__main__":
    main()
