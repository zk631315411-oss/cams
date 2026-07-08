from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def compact(value) -> str:
    return " ".join(str(value or "").split())


def clean_analysis(text: str) -> str:
    text = compact(text)
    if not text:
        return ""
    markers = ["试题详解", "原解析"]
    starts = [text.find(marker) for marker in markers if text.find(marker) >= 0]
    if starts:
        text = text[min(starts) :]
    tail_markers = [
        "使用【深度解题】",
        "2026新真题",
        "2026新直",
        "来 源",
        "来源",
        "考友笔记",
        "写笔记",
        "0/395",
        "分享",
        "已做题",
        "收藏",
    ]
    for marker in tail_markers:
        pos = text.find(marker)
        if pos > 60:
            text = text[:pos]
    return compact(text)


def answer_status(cn_answer, en_answer) -> str:
    cn = compact(cn_answer).upper()
    en = compact(en_answer).upper()
    if cn and en:
        return "一致" if cn == en else f"不一致: 中文={cn}; 英文={en}"
    if cn or en:
        return "单侧识别"
    return "未识别"


def set_hyperlink(cell, value: str) -> None:
    if value:
        cell.value = value
        cell.hyperlink = value
        cell.style = "Hyperlink"


def load_frame(out_dir: Path, video_stem: str, index: int) -> dict:
    frames_path = out_dir / "json" / f"{video_stem}_frames.json"
    frames = json.loads(frames_path.read_text(encoding="utf-8"))
    for frame in frames:
        if int(frame.get("index") or -1) == index:
            return frame
    raise KeyError(f"Frame not found: {video_stem}#{index}")


EN344_ANALYSIS = (
    "试题详解 原解析 高风险业务部门和结构,如空壳公司,常涉及隐蔽资金来源和用途。"
    "选项A中,商品或服务与实际情况或金融活动性质不符,是空壳公司转移资金的典型手法;"
    "选项D中,无法识别资金转账发起人或受益人,符合空壳公司匿名交易的特征;"
    "选项E中,支付款项未注明用途,也未提及商品或服务,易被用于隐藏非法资金流动。"
    "选项B中,虽累计金额大但单笔支付低且符合商业惯例;选项C中,公司交易有完整文件和审计记录,"
    "均不符合高风险特征。因此,正确答案为ADE."
)


FIXES = {
    47: {"answer": "B", "answer_frame": ("英文版1", 155)},
    57: {"answer": "B", "answer_frame": ("英文版1", 189)},
    91: {"answer": "B", "answer_frame": ("英文版1", 304)},
    120: {"answer": "C", "answer_frame": ("英文版2", 63)},
    129: {"analysis_frame": ("英文版2", 91)},
    175: {"answer": "A", "answer_frame": ("英文版2", 230)},
    179: {"answer": "D", "answer_frame": ("英文版2", 245)},
    218: {"answer": "CD", "answer_frame": ("英文版3", 50)},
    252: {"answer": "BDE", "answer_frame": ("英文版3", 163)},
    255: {"answer": "ADE", "answer_frame": ("英文版3", 176)},
    264: {"answer": "BC", "answer_frame": ("英文版3", 204), "note": "答案来自解析帧中的“选择BC”"},
    277: {"answer": "ACE", "answer_frame": ("英文版3", 239)},
    283: {"analysis_frame": ("英文版3", 258)},
    292: {"answer": "BD", "answer_frame": ("英文版3", 289)},
    295: {"answer": "CF", "answer_frame": ("英文版3", 302)},
    320: {"answer": "BC", "answer_frame": ("英文版4", 57), "note": "答案来自解析帧中的“B和C”"},
    344: {
        "answer": "ADE",
        "answer_frame": ("英文版4", 119),
        "analysis": EN344_ANALYSIS,
        "analysis_frame": ("中文版4", 124),
        "note": "英文视频解析只露出开头, 采用语义对齐中文视频完整解析",
    },
    345: {"answer": "BC", "answer_frame": ("英文版4", 121), "analysis_frame": ("英文版4", 122)},
    363: {"answer": "BCD", "answer_frame": ("英文版4", 172)},
}


def write_log(wb, rows: list[list]) -> None:
    title = "空答案解析回填记录_v2"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    headers = ["语言", "题号", "字段", "旧值", "新值摘要", "证据截图", "来源帧", "备注", "同步行数"]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [10, 10, 12, 14, 64, 72, 18, 44, 12]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
        if ws.cell(row, 6).value:
            ws.cell(row, 6).hyperlink = ws.cell(row, 6).value
            ws.cell(row, 6).style = "Hyperlink"
    ws.freeze_panes = "A2"


def main() -> None:
    base_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parents[1]
    out_dir = base_dir / "output_2s"
    source = out_dir / "semantic_aligned_cn_en_filled_missing.xlsx"
    target = out_dir / "semantic_aligned_cn_en_filled_missing_v2.xlsx"

    wb = load_workbook(source)
    en_ws = wb["英文题库"]
    align_ws = wb["语义对齐"]
    log_rows: list[list] = []

    rows_by_qno = {}
    for row in range(2, en_ws.max_row + 1):
        try:
            rows_by_qno[int(en_ws.cell(row, 1).value)] = row
        except (TypeError, ValueError):
            continue

    for qno, fix in FIXES.items():
        row = rows_by_qno[qno]
        answer_evidence = ""
        analysis_evidence = ""
        note = fix.get("note", "")

        if "answer_frame" in fix:
            video, index = fix["answer_frame"]
            frame = load_frame(out_dir, video, index)
            answer_evidence = str(frame.get("image") or "")

        if "analysis_frame" in fix:
            video, index = fix["analysis_frame"]
            frame = load_frame(out_dir, video, index)
            analysis_evidence = str(frame.get("image") or "")

        new_answer = fix.get("answer")
        if new_answer and not en_ws.cell(row, 3).value:
            old = en_ws.cell(row, 3).value
            en_ws.cell(row, 3).value = new_answer
            log_rows.append(
                [
                    "EN",
                    qno,
                    "答案",
                    old or "",
                    new_answer,
                    answer_evidence,
                    f"{fix['answer_frame'][0]}#{fix['answer_frame'][1]}",
                    note,
                    "",
                ]
            )

        new_analysis = fix.get("analysis")
        if not new_analysis and "analysis_frame" in fix:
            video, index = fix["analysis_frame"]
            frame = load_frame(out_dir, video, index)
            new_analysis = clean_analysis(str(frame.get("clean_text") or frame.get("raw_text") or ""))
        if new_analysis and not en_ws.cell(row, 7).value:
            old = en_ws.cell(row, 7).value
            en_ws.cell(row, 7).value = new_analysis
            if analysis_evidence:
                set_hyperlink(en_ws.cell(row, 10), analysis_evidence)
            log_rows.append(
                [
                    "EN",
                    qno,
                    "解析",
                    old or "",
                    new_analysis[:220],
                    analysis_evidence,
                    f"{fix['analysis_frame'][0]}#{fix['analysis_frame'][1]}",
                    note,
                    "",
                ]
            )

        for col in range(1, en_ws.max_column + 1):
            en_ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")

        synced = 0
        for align_row in range(2, align_ws.max_row + 1):
            try:
                en_qno = int(align_ws.cell(align_row, 10).value)
            except (TypeError, ValueError):
                continue
            if en_qno != qno:
                continue
            if new_answer:
                align_ws.cell(align_row, 12).value = new_answer
            if new_analysis:
                align_ws.cell(align_row, 16).value = new_analysis
            align_ws.cell(align_row, 8).value = answer_status(
                align_ws.cell(align_row, 11).value,
                align_ws.cell(align_row, 12).value,
            )
            for col in range(1, align_ws.max_column + 1):
                align_ws.cell(align_row, col).alignment = Alignment(wrap_text=True, vertical="top")
            synced += 1
        for log_row in log_rows:
            if log_row[1] == qno:
                log_row[-1] = synced

    write_log(wb, log_rows)
    wb.save(target)
    print(target)
    print("records", len(log_rows))


if __name__ == "__main__":
    main()
