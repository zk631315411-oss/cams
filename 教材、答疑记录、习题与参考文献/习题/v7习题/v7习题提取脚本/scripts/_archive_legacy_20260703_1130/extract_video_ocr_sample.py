from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass, field
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rapidocr_onnxruntime import RapidOCR


NOISE_PATTERNS = [
    "答题",
    "背题",
    "语音",
    "纠错",
    "收藏",
    "分享",
    "已做题",
    "新题",
    "斩题",
    "点击设置",
    "主题模式",
    "备考护眼",
    "向往的生活",
    "向往的生",
    "治愈系田园经营",
    "我们在一起",
    "由赞助商提供的内容",
    "广告",
    "本人作答",
    "全站作答",
    "正确率",
    "难度",
    "使用",
    "获取深度思考过程",
    "2026冲刺题",
    "写笔记",
    "完善笔记",
    "来源",
    "主站",
]


@dataclass
class OcrLine:
    y: float
    x: float
    text: str
    score: float


@dataclass
class FrameOcr:
    index: int
    time_sec: int
    image: str
    page_no: str = ""
    answer: str = ""
    has_explanation: bool = False
    text: str = ""
    clean_text: str = ""
    lines: list[OcrLine] = field(default_factory=list)


@dataclass
class QuestionCluster:
    cluster_id: int
    page_no: str = ""
    answer: str = ""
    first_time: int = 0
    last_time: int = 0
    frames: list[FrameOcr] = field(default_factory=list)

    @property
    def best_question_frame(self) -> FrameOcr:
        return max(
            self.frames,
            key=lambda f: (
                score_question_text(f.clean_text),
                -abs(f.time_sec - self.first_time),
            ),
        )

    @property
    def best_explanation_frame(self) -> FrameOcr:
        return max(
            self.frames,
            key=lambda f: (
                int(f.has_explanation),
                len(f.clean_text),
            ),
        )


def center(box: list[list[float]]) -> tuple[float, float]:
    return (
        sum(point[1] for point in box) / 4,
        sum(point[0] for point in box) / 4,
    )


def normalize_text(text: str) -> str:
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("，", ",").replace("。", ".")
    text = text.replace("：", ":").replace("；", ";")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def is_noise_line(text: str) -> bool:
    compact = re.sub(r"\s+", "", text)
    if not compact:
        return True
    if compact in {"A", "B", "C", "D"}:
        return False
    if any(pattern in compact for pattern in NOISE_PATTERNS):
        return True
    if re.fullmatch(r"\d+/\d+", compact):
        return True
    if re.search(r"\d+次.*正确率", compact):
        return True
    return False


def clean_lines(lines: list[OcrLine]) -> list[str]:
    cleaned: list[str] = []
    for line in lines:
        if line.y < 72 or line.y > 900:
            continue
        text = normalize_text(line.text)
        if line.score < 0.58:
            continue
        if is_noise_line(text):
            continue
        cleaned.append(text)
    return cleaned


def detect_page_no(lines: list[OcrLine]) -> str:
    # The app footer also contains progress like "4/395"; only trust the
    # current-question number near the upper-left content area.
    candidates: list[int] = []
    for line in lines:
        if not (75 <= line.y <= 280 and line.x <= 180):
            continue
        compact = re.sub(r"\s+", "", line.text)
        for match in re.findall(r"(\d{1,3})\s*/\s*(?:395|396)", compact):
            num = int(match)
            if 1 <= num <= 396:
                candidates.append(num)
        # OCR sometimes drops the slash in compact page labels such as "6/395".
        # Only accept the no-slash form when the line is essentially just that
        # page label; otherwise numbers inside the question text create false
        # positives such as 143/395.
        if len(compact) <= 8:
            no_slash_matches = re.findall(r"^(\d{1,3})(?:395|396)$", compact)
        else:
            no_slash_matches = []
        for match in no_slash_matches:
            num = int(match)
            if 1 <= num <= 396:
                candidates.append(num)
    if not candidates:
        return ""
    counts = Counter(candidates)
    best_count = max(counts.values())
    best = min(num for num, count in counts.items() if count == best_count)
    return str(best)


def cluster_page_no(cluster: QuestionCluster) -> str:
    values = [frame.page_no for frame in cluster.frames if frame.page_no]
    if not values:
        return cluster.page_no
    counts = Counter(values)
    best_count = max(counts.values())
    best = min(value for value, count in counts.items() if count == best_count)
    return best


def cluster_answer(cluster: QuestionCluster) -> str:
    values = [frame.answer for frame in cluster.frames if frame.answer]
    if not values:
        return cluster.answer
    counts = Counter(values)
    best_count = max(counts.values())
    best = min(value for value, count in counts.items() if count == best_count)
    return best


def detect_answer(text: str) -> str:
    normalized = text.replace(" ", "")
    patterns = [
        r"正确答案[:：为]?\s*([A-E]{1,5})",
        r"正确答案是([A-E]{1,5})",
        r"答案[:：为]?\s*([A-E]{1,5})",
    ]
    patterns = [pattern.replace("[A-E]{1,5}", "[A-F]{1,6}") for pattern in patterns]
    for pattern in patterns:
        match = re.search(pattern, normalized)
        if match:
            return "".join(dict.fromkeys(match.group(1)))
    return ""


def score_question_text(text: str) -> int:
    compact = re.sub(r"\s+", "", text)
    score = min(len(text), 1200) // 2

    question_tokens = [
        "单选",
        "多选",
        "Which",
        "What",
        "When",
        "According",
        "Select",
        "Choose",
    ]
    if any(token in text for token in question_tokens):
        score += 650
    if "-[" in text:
        score += 250

    option_hits = 0
    for option in "ABCDEF":
        if re.search(rf"(^|[\s,，。:：;；]){option}([\s\w\u4e00-\u9fff]|$)", text):
            option_hits += 1
    if option_hits >= 2:
        score += 200 + option_hits * 40

    # Explanation screens are often longer than question screens. Penalize them
    # hard so the representative question image stays on the stem/options.
    for token in ["正确答案", "试题详解", "原解析", "本人作答", "全站作答"]:
        score -= text.count(token) * 260
    if compact.startswith(("正确答案", "试题详解", "原解析")):
        score -= 1200
    return score


def question_signature(text: str) -> str:
    text = re.sub(r"\b\d{1,3}\s*/\s*(?:395|396)\b", " ", text)
    text = re.sub(r"正确答案.*", " ", text)
    text = re.sub(r"试题详解.*", " ", text)
    text = re.sub(r"原解析.*", " ", text)
    text = re.sub(r"[A-D]\s+", " ", text)
    text = re.sub(r"[\W_]+", "", text, flags=re.UNICODE)
    return text[:80]


def similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    a_set = {a[i : i + 2] for i in range(max(1, len(a) - 1))}
    b_set = {b[i : i + 2] for i in range(max(1, len(b) - 1))}
    if not a_set or not b_set:
        return 0.0
    return len(a_set & b_set) / len(a_set | b_set)


def run_ocr(ocr: RapidOCR, image_path: Path, index: int) -> FrameOcr:
    result, _ = ocr(str(image_path))
    lines: list[OcrLine] = []
    for box, text, score in result or []:
        y, x = center(box)
        lines.append(OcrLine(y=y, x=x, text=text, score=float(score)))
    lines.sort(key=lambda item: (item.y, item.x))
    raw_text = "\n".join(line.text for line in lines if line.score >= 0.58)
    cleaned = "\n".join(clean_lines(lines))
    flat = normalize_text(raw_text)
    clean_flat = normalize_text(cleaned)
    return FrameOcr(
        index=index,
        time_sec=index - 1,
        image=str(image_path.resolve()),
        page_no=detect_page_no(lines),
        answer=detect_answer(flat),
        has_explanation=("试题详解" in flat or "原解析" in flat),
        text=raw_text,
        clean_text=clean_flat,
        lines=lines,
    )


def cluster_frames(frames: list[FrameOcr]) -> list[QuestionCluster]:
    clusters: list[QuestionCluster] = []
    current: QuestionCluster | None = None
    current_sig = ""

    for frame in frames:
        sig = question_signature(frame.clean_text)
        starts_new = False
        if frame.page_no and current and current.page_no and frame.page_no != current.page_no:
            starts_new = True
        elif (
            current
            and sig
            and current_sig
            and not frame.has_explanation
            and similarity(sig, current_sig) < 0.16
            and score_question_text(frame.clean_text) > 220
        ):
            starts_new = True

        if current is None or starts_new:
            current = QuestionCluster(
                cluster_id=len(clusters) + 1,
                page_no=frame.page_no,
                answer=frame.answer,
                first_time=frame.time_sec,
                last_time=frame.time_sec,
                frames=[],
            )
            clusters.append(current)
            current_sig = sig

        current.frames.append(frame)
        current.last_time = frame.time_sec
        if frame.page_no and not current.page_no:
            current.page_no = frame.page_no
        if frame.answer:
            current.answer = frame.answer
        if sig and (not current_sig or len(sig) > len(current_sig)):
            current_sig = sig

    return clusters


def write_excel(path: Path, frames: list[FrameOcr], clusters: list[QuestionCluster]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "题目聚合"
    headers = [
        "聚合ID",
        "识别页码",
        "答案",
        "起始秒",
        "结束秒",
        "帧数",
        "题目候选文本",
        "解析候选文本",
        "代表题目截图",
        "代表解析截图",
    ]
    ws.append(headers)
    for cluster in clusters:
        qf = cluster.best_question_frame
        ef = cluster.best_explanation_frame
        ws.append(
            [
                cluster.cluster_id,
                cluster_page_no(cluster),
                cluster_answer(cluster),
                cluster.first_time,
                cluster.last_time,
                len(cluster.frames),
                qf.clean_text,
                ef.clean_text,
                qf.image,
                ef.image,
            ]
        )

    ws_raw = wb.create_sheet("逐帧OCR")
    ws_raw.append(["帧序号", "时间秒", "识别页码", "答案", "有解析", "截图", "清洗文本", "原始OCR文本"])
    for frame in frames:
        ws_raw.append(
            [
                frame.index,
                frame.time_sec,
                frame.page_no,
                frame.answer,
                "是" if frame.has_explanation else "",
                frame.image,
                frame.clean_text,
                frame.text,
            ]
        )

    for sheet in [ws, ws_raw]:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        widths = {
            "A": 10,
            "B": 10,
            "C": 8,
            "D": 10,
            "E": 10,
            "F": 8,
            "G": 80,
            "H": 80,
            "I": 58,
            "J": 58,
        }
        for col, width in widths.items():
            if sheet.max_column >= ord(col) - 64:
                sheet.column_dimensions[col].width = width
        for idx in range(1, sheet.max_row + 1):
            sheet.row_dimensions[idx].height = 60 if idx > 1 else 24

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--frames-dir", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--start", type=int, default=1)
    parser.add_argument("--end", type=int, default=80)
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    frame_paths = [
        args.frames_dir / f"cn1_{index:04d}.jpg"
        for index in range(args.start, args.end + 1)
        if (args.frames_dir / f"cn1_{index:04d}.jpg").exists()
    ]

    ocr = RapidOCR()
    frames = [run_ocr(ocr, path, args.start + offset) for offset, path in enumerate(frame_paths)]
    clusters = cluster_frames(frames)

    raw_json = args.output_dir / f"cn1_frames_{args.start:04d}_{args.end:04d}.json"
    raw_json.write_text(
        json.dumps(
            [
                {
                    "index": frame.index,
                    "time_sec": frame.time_sec,
                    "image": frame.image,
                    "page_no": frame.page_no,
                    "answer": frame.answer,
                    "has_explanation": frame.has_explanation,
                    "clean_text": frame.clean_text,
                    "raw_text": frame.text,
                }
                for frame in frames
            ],
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    excel_path = args.output_dir / f"cn1_sample_{args.start:04d}_{args.end:04d}.xlsx"
    write_excel(excel_path, frames, clusters)
    print(f"frames={len(frames)} clusters={len(clusters)}")
    print(f"json={raw_json}")
    print(f"xlsx={excel_path}")


if __name__ == "__main__":
    main()
