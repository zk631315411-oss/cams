"""
极简测试：英文基准知识点提取与中英双语对齐导出。
CAMS v7 MinerU Markdown + PDF TOC → 英文切分 → 中文片段归集 → 审计型 xlsx。
"""
from __future__ import annotations

import html
import json
import re
import subprocess
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

# ── 路径常量 ──────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_EN = Path(
    "D:/守正公司工作区/cams考试/教材、答疑记录、习题与参考文献/"
    "教材原文/v7/mineru提取/英文/v7_en_mineru_merged.md"
)
INPUT_ZH = Path(
    "D:/守正公司工作区/cams考试/教材、答疑记录、习题与参考文献/"
    "教材原文/v7/mineru提取/中文/v7_zh_mineru_merged.md"
)
OUTPUT_XLSX = SCRIPT_DIR / "v7_knowledge_points.xlsx"
OUTPUT_ALIGNED_XLSX = SCRIPT_DIR / "v7_knowledge_points_aligned.xlsx"
ALIGNMENT_MANIFEST = SCRIPT_DIR / "bilingual_alignment.json"
INPUT_EN_PDF = INPUT_EN.parents[2] / "v7_en_split.pdf"
INPUT_ZH_PDF = INPUT_ZH.parents[2] / "v7_zh_split.pdf"

# The English extraction contains four headings that are not emitted verbatim by
# pdftotext on the table-of-contents pages. Their Chinese PDF counterparts are
# still explicit and provide stable anchors.
TOC_ALIGNMENT_OVERRIDES = {
    "Terrorism financing": ("资助恐怖主义", 36),
    "Case example: Mr. Wolfe’s scheme": ("病例示例：Wolfe先生的方案", 38),
    "AI regulations around the world": ("全球人工智能监管现状", 214),
    "Case example: Financial crime functions' structure at Global Finance, Corp.": (
        "案例分析：全球金融公司金融犯罪职能部门架构",
        253,
    ),
}

# ── 工具函数 ──────────────────────────────────────────────

def compact_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def protect_angle_placeholders(text: str) -> tuple[str, dict[str, str]]:
    placeholders: dict[str, str] = {}

    def repl(match: re.Match[str]) -> str:
        inner = match.group(1).strip()
        if not inner or len(inner) > 80:
            return match.group(0)
        if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9 /&._-]*", inner):
            return match.group(0)
        key = f"@@ANGLE_PLACEHOLDER_{len(placeholders)}@@"
        placeholders[key] = f"<{inner}>"
        return key

    return re.sub(r"<([^<>]+)>", repl, text), placeholders


def visible_text_from_md(text: str) -> str:
    text = re.sub(r"<!--.*?-->", " ", text, flags=re.S)
    text, placeholders = protect_angle_placeholders(text)
    text = re.sub(r"<[^>]+>", " ", text)
    for key, value in placeholders.items():
        text = text.replace(key, value)
    text = re.sub(r"!\[[^\]]*\]\([^)]+\)", " ", text)
    text = re.sub(r"\[[^\]]+\]\([^)]+\)", " ", text)
    text = html.unescape(text)
    return compact_spaces(text)


# ── 第一阶段：清洗 Markdown ──────────────────────────────

def fix_false_headings(md_text: str) -> tuple[str, int]:
    """去掉 MinerU 误标的假 ## 标题（bullet/编号被错误加了 ## 前缀）。"""
    # ## • xxx 或 ## 数字. xxx 或 ## 数字) xxx → 去掉 ##
    false_pattern = re.compile(r"^## (\s*(?:•|\d+[.)]\s*))", re.MULTILINE)
    fixed_text, count = re.subn(false_pattern, r"\1", md_text)
    # 特殊：## 1LoD（后面没有标点，但明显不是标题）
    fixed_text, n = re.subn(r"^## (1LoD)$", r"\1", fixed_text, flags=re.MULTILINE)
    count += n
    return fixed_text, count


def insert_missing_case_headings(md_text: str) -> tuple[str, int]:
    """补上 MinerU 丢失的 ## Case example: 标题。"""
    count = 0

    # (信号文本, 案例名称) — 信号文本需在正文中唯一匹配
    missing_cases: list[tuple[str, str]] = [
        (
            "According to a US Department of Justice press release, "
            "in December 2019, Yamel Guevara Tamayo",
            "Case example: Tamayo's money mules",
        ),
        (
            "Businessman Alexei Komarov amassed his fortune "
            "through Volkof Industries",
            "Case example: Komarov's tactics",
        ),
        (
            "Danske Bank, Denmark's largest financial institution, "
            "became embroiled in a significant money laundering case",
            "Case example: Estonian bank branch",
        ),
    ]

    for signal, case_title in missing_cases:
        if f"## {case_title}" in md_text:
            continue  # 已存在，跳过
        idx = md_text.find(signal)
        if idx == -1:
            print(f"  [WARN] 未找到案例信号: {case_title}")
            continue
        # 在信号文本所在行首插入 ## Case example: 标题
        line_start = md_text.rfind("\n", 0, idx) + 1
        md_text = md_text[:line_start] + f"## {case_title}\n" + md_text[line_start:]
        count += 1

    # 修复错误级别的案例标题：# Case example: → ## Case example:
    wrong_level_pattern = re.compile(r"^(# Case example:)", re.MULTILINE)
    fixed_text, n = re.subn(wrong_level_pattern, r"#\1", md_text)
    # 上面会把 # Case example: 变成 ## Case example:
    # 不对, 让我直接替换
    fixed_text, n = re.subn(
        r"^# (Case example:)", r"## \1", md_text, flags=re.MULTILINE
    )
    count += n

    return fixed_text, count


def strip_mineru_merge_markers(md_text: str) -> str:
    """去掉 MinerU 文件合并标记行。"""
    return re.sub(
        r"^<!-- MINERU_MERGE_SOURCE .*? -->\n", "", md_text, flags=re.MULTILINE
    )


def parse_toc(md_text: str, lang: str = "en") -> tuple[set[str], set[str]]:
    """从 TOC 区域解析模块和章标题列表。"""
    module_titles: set[str] = set()
    chapter_titles: set[str] = set()

    toc_anchor = "## Table of Contents" if lang == "en" else "## 目录"
    page_pattern = r"\(Page\s+\d+\)" if lang == "en" else r"（第\s+\d+\s+页）"

    toc_start = md_text.find(toc_anchor)
    if toc_start == -1:
        return module_titles, chapter_titles

    toc_text = md_text[toc_start:]

    toc_heading_end = toc_text.find("\n", toc_text.find(toc_anchor))
    search_start = toc_heading_end if toc_heading_end > 0 else 0
    body_start = None
    for m in re.finditer(r"^(#{1,2}) (.+)$", toc_text[search_start:], re.MULTILINE):
        if not re.search(page_pattern, m.group(0)):
            body_start = search_start + m.start()
            break

    if body_start is not None:
        toc_text = toc_text[:body_start]

    for line in toc_text.splitlines():
        raw = line.strip()
        if raw in (toc_anchor, ""):
            continue
        if not raw.startswith("#"):
            continue

        title = re.sub(r"\s*" + page_pattern + r"\s*$", "", raw)
        title = re.sub(r"^#{1,3}\s+", "", title).strip()
        if not title:
            continue

        if raw.startswith("# ") and not raw.startswith("## "):
            module_titles.add(title)
        elif raw.startswith("## ") and not raw.startswith("### "):
            chapter_titles.add(title)

    return module_titles, chapter_titles


# 从 TOC 自动解析（运行时填充）
MODULE_TITLES: set[str] = set()
CHAPTER_TITLES: set[str] = set()


def fix_heading_levels(md_text: str) -> tuple[str, int]:
    """修正三层标题层级：
    #   = 模块（MODULE_TITLES）
    ##  = 知识点（CHAPTER_TITLES，不论原始层级）
    ### = 节（其余全部）
    """
    changed = 0

    lines = md_text.splitlines()
    fixed_lines: list[str] = []
    h1_pattern = re.compile(r"^# (?!\#)(.+)$")
    h2_pattern = re.compile(r"^## (?!\#)(.+)$")

    for line in lines:
        h1_match = h1_pattern.match(line)
        h2_match = h2_pattern.match(line)

        if h1_match:
            title = h1_match.group(1).strip()
            if title in MODULE_TITLES:
                fixed_lines.append(line)  # # 模块
            elif title in CHAPTER_TITLES:
                fixed_lines.append(f"## {title}")  # ## 知识点
                changed += 1
            else:
                fixed_lines.append(f"### {title}")  # ### 节
                changed += 1

        elif h2_match:
            title = h2_match.group(1).strip()
            if title in MODULE_TITLES:
                fixed_lines.append(f"# {title}")  # 提升为模块
                changed += 1
            elif title in CHAPTER_TITLES:
                fixed_lines.append(line)  # 保留 ##（知识点在正文中已是 ##）
            else:
                fixed_lines.append(f"### {title}")  # ### 节
                changed += 1

        else:
            fixed_lines.append(line)

    return "\n".join(fixed_lines), changed


def insert_missing_chapter_headings(md_text: str) -> tuple[str, int]:
    """补上正文中缺失的章标题（TOC 有但正文无对应标题）。"""
    count = 0

    # 仅在正文区域搜索（TOC 之后）
    toc_marker = "## Table of Contents"
    toc_start = md_text.find(toc_marker)
    body_start = toc_start + len(toc_marker) if toc_start >= 0 else 0
    # 跳过 TOC 内容直到第一个不含 "(Page" 的标题行
    body_real_start = body_start
    for m in re.finditer(r"^(#{1,2}) (.+)$", md_text[body_start:], re.MULTILINE):
        if "Page" not in m.group(0):
            body_real_start = body_start + m.start()
            break

    body_text = md_text[body_real_start:]

    # (在正文中的特征文本, 章标题)
    missing: list[tuple[str, str]] = [
        (
            "\n# Financial Action Task Force\n",
            "## Global AFC Standards and Guidance",
        ),
        (
            "\n## Introduction: Components of an AFC program",
            "## Components of an AFC Program",
        ),
        (
            "\n# Types of risk assessment\n",
            "## Risk Assessment",
        ),
        (
            "\n# Data as an input for solutions\n",
            "## Data Collection and Preparation",
        ),
    ]
    for signal, heading in missing:
        if heading in body_text:
            continue
        idx = body_text.find(signal)
        if idx == -1:
            print(f"  [WARN] 未找到章信号: {heading}")
            continue
        insert_pos = idx + 1  # 在 \n 之后、信号行之前插入
        body_text = body_text[:insert_pos] + heading + "\n" + body_text[insert_pos:]
        count += 1

    # 组装回去
    if count > 0:
        md_text = md_text[:body_real_start] + body_text

    return md_text, count


def clean_markdown(md_text: str) -> str:
    """清洗全部 Markdown 格式问题，返回干净文本。"""
    print("=" * 60)
    print("第一阶段：清洗 Markdown 格式")
    print("=" * 60)

    md_text = strip_mineru_merge_markers(md_text)

    md_text, n_false = fix_false_headings(md_text)
    print(f"  [FIX] 去除假 ## 标题: {n_false} 处")

    md_text, n_missing = insert_missing_case_headings(md_text)
    print(f"  [FIX] 补缺/修正案例标题: {n_missing} 处")

    md_text, n_ch = insert_missing_chapter_headings(md_text)
    print(f"  [FIX] 补缺章标题: {n_ch} 处")

    md_text, n_level = fix_heading_levels(md_text)
    print(f"  [FIX] 修正标题层级: {n_level} 处")
    print(f"         #  = 模块({len(MODULE_TITLES)})  |  ## = 知识点({len(CHAPTER_TITLES)})  |  ### = 节")

    return md_text


# ── 第二阶段：提取知识点 ──────────────────────────────────

# 前置页面关键词 — heading_stack 含这些视为非正文
FRONT_MATTER_KEYWORDS = [
    "Credits", "Copyright", "Table of Contents", "Study Guide",
    "ACAMS Task Force", "ACAMS Product Staff",
]

# 需要跳过的 ## 标题（精确匹配）
SKIP_TITLES = {"Key takeaways", "关键要点"}


def split_by_h3(md_text: str) -> list[dict]:
    """按 ### 切分知识点，追踪 # 模块和 ## 章上下文。"""
    lines = md_text.splitlines()
    sections: list[dict] = []
    current_module = "Front Matter"
    current_chapter = ""
    current_title: str | None = None
    current_body_lines: list[str] = []
    in_body = False

    h1_pattern = re.compile(r"^# (?!\#)(.+)$")
    h2_pattern = re.compile(r"^## (?!\#)(.+)$")
    h3_pattern = re.compile(r"^### (.+)$")

    for line in lines:
        h1_match = h1_pattern.match(line)
        h2_match = h2_pattern.match(line)
        h3_match = h3_pattern.match(line)

        if h1_match:
            if current_title is not None:
                sections.append({
                    "title": current_title,
                    "body": "\n".join(current_body_lines).strip(),
                    "module": current_module,
                    "chapter": current_chapter,
                })
                current_body_lines = []
                in_body = False
            current_module = h1_match.group(1).strip()
            current_chapter = ""
            current_title = None

        elif h2_match:
            if current_title is not None:
                sections.append({
                    "title": current_title,
                    "body": "\n".join(current_body_lines).strip(),
                    "module": current_module,
                    "chapter": current_chapter,
                })
                current_body_lines = []
                in_body = False
            current_chapter = h2_match.group(1).strip()
            current_title = None

        elif h3_match:
            if current_title is not None:
                sections.append({
                    "title": current_title,
                    "body": "\n".join(current_body_lines).strip(),
                    "module": current_module,
                    "chapter": current_chapter,
                })
            current_title = h3_match.group(1).strip()
            current_body_lines = []
            in_body = True

        elif in_body:
            current_body_lines.append(line)

    if current_title is not None:
        sections.append({
            "title": current_title,
            "body": "\n".join(current_body_lines).strip(),
            "module": current_module,
            "chapter": current_chapter,
        })

    return sections


def clean_body(body: str) -> str:
    """去除 body 中的图片、HTML 残留，规整空白。"""
    # 去掉图片
    body = re.sub(r"!\[.*?\]\(.*?\)", "", body)
    # 去掉 HTML table 残留（含 table 标签的整段）
    body = re.sub(r"<table>.*?</table>", "", body, flags=re.S | re.I)
    # 去掉 <sub>, <td>, <tr> 等残留标签，保留内容
    body = re.sub(r"</?(?:sub|td|tr|th|col|colgroup|tbody|thead|caption)[^>]*>", "", body)
    # 规整空白
    body = re.sub(r"\n{3,}", "\n\n", body)
    body = compact_spaces(body)
    return body


def is_front_matter(chapter: str) -> bool:
    """判断是否属于前置页面。"""
    chapter_lower = chapter.lower()
    for kw in FRONT_MATTER_KEYWORDS:
        if kw.lower() in chapter_lower:
            return True
    return False


def filter_sections(sections: list[dict]) -> list[dict]:
    """过滤：去 Front Matter、Key takeaways、Glossary、空 body。"""
    filtered = []
    in_glossary = False

    for sec in sections:
        title = sec["title"]
        module = sec["module"]
        chapter = sec.get("chapter", "")

        if module == "Front Matter":
            continue

        if title.strip() == "Glossary":
            in_glossary = True
            break

        if title.strip() in SKIP_TITLES:
            continue

        # 跳过章级 Key takeaways（### 下可能有正常小节）
        if chapter.strip() in SKIP_TITLES:
            continue

        body = clean_body(sec["body"])

        if not body or len(body) < 10:
            continue

        filtered.append({
            "title": title.strip(),
            "body": body,
            "module": module.strip(),
            "chapter": chapter.strip(),
        })

    return filtered


def classify_section(title: str) -> str:
    """根据 ## 标题关键词标注知识点类型。"""
    t = title.lower()
    if "what is" in t:
        return "definition"
    if "case example" in t:
        return "case_example"
    if "key takeaways" in t:
        return "key_takeaway"
    if "risks" in t:
        return "risk_analysis"
    if "consequences" in t or "impact" in t:
        return "consequence"
    if "versus" in t or "compared" in t:
        return "comparison"
    if "accountability" in t or "compliance" in t:
        return "requirement"
    if "techniques" in t or "methods" in t or "types of" in t:
        return "technique_enum"
    if "introduction" in t:
        return "introduction"
    return "concept"


def build_numbering(sections: list[dict]) -> list[dict]:
    """为知识点分配模块/章/节编号。"""
    result: list[dict] = []
    # 按 (module, chapter) 分组，组内按出现顺序编号
    group_order: list[tuple[str, str]] = []
    group_kps: dict[tuple[str, str], list[dict]] = {}

    for sec in sections:
        key = (sec["module"], sec["chapter"])
        if key not in group_kps:
            group_order.append(key)
            group_kps[key] = []
        group_kps[key].append(sec)

    mod_idx = 0
    last_module = ""
    for mod_name, ch_name in group_order:
        if mod_name != last_module:
            mod_idx += 1
            last_module = mod_name
            ch_idx = 0
        ch_idx += 1
        kps = group_kps[(mod_name, ch_name)]
        for kp_seq, sec in enumerate(kps, start=1):
            kp_type = classify_section(sec["title"])
            result.append({
                "kp_id": f"M{mod_idx:02d}-{ch_idx:02d}-{kp_seq:02d}",
                "module": mod_name,
                "chapter": ch_name,
                "section_title": sec["title"],
                "kp_type": kp_type,
                "kp_content": sec["body"],
                "char_count": len(sec["body"]),
            })

    for i, item in enumerate(result, start=1):
        item["kp_index"] = i

    return result


# ── 第三阶段：导出 xlsx ────────────────────────────────────

def write_xlsx(kps: list[dict], path: Path) -> None:
    """将知识点列表写入 xlsx 文件。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "知识點"

    # 表头
    headers = [
        "KP_ID", "MODULE", "CHAPTER", "SECTION_TITLE", "KP_CONTENT",
    ]
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    header_font = Font(bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    wrap_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")

    for row_idx, kp in enumerate(kps, start=2):
        values = [
            kp["kp_id"], kp["module"], kp["chapter"],
            kp["section_title"], kp["kp_content"],
            kp["char_count"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 1:  # KP_ID 居中
                cell.alignment = center_align
            else:
                cell.alignment = wrap_align

    # 列宽
    col_widths = {1: 14, 2: 40, 3: 45, 4: 45, 5: 100}
    for col, width in col_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(kps) + 1}"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"\n写入 xlsx: {path}")


def write_xlsx_combined(kps: list[dict], path: Path) -> None:
    """写入中英双语 xlsx。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "知识点"

    headers = [
        "KP_ID",
        "MODULE_EN", "MODULE_ZH",
        "CHAPTER_EN", "CHAPTER_ZH",
        "SECTION_TITLE_EN", "SECTION_TITLE_ZH",
        "KP_CONTENT_EN", "KP_CONTENT_ZH",
    ]
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    header_font = Font(bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    wrap_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")

    for row_idx, kp in enumerate(kps, start=2):
        values = [
            kp["kp_id"],
            kp["module"], kp.get("module_zh", ""),
            kp["chapter"], kp.get("chapter_zh", ""),
            kp["section_title"], kp.get("section_title_zh", ""),
            kp["kp_content"], kp.get("kp_content_zh", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = center_align
            else:
                cell.alignment = wrap_align

    col_widths = {1: 14, 2: 40, 3: 40, 4: 45, 5: 45, 6: 45, 7: 45, 8: 100, 9: 100}
    for col, width in col_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(kps) + 1}"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"\n写入 xlsx: {path}")


# ── 汇总 ───────────────────────────────────────────────────

def print_summary(kps: list[dict]) -> None:
    """打印统计和抽样。"""
    print("\n" + "=" * 60)
    print("第二阶段：知识点提取结果")
    print("=" * 60)

    # 总览
    total = len(kps)
    total_chars = sum(k["char_count"] for k in kps)
    modules = sorted(set(k["module"] for k in kps))
    types: dict[str, int] = {}
    for k in kps:
        t = k["kp_type"]
        types[t] = types.get(t, 0) + 1

    print(f"\n知识点总数: {total}")
    print(f"总字符数: {total_chars:,}")
    print(f"涉及模块数: {len(modules)}")

    # 每模块分布
    print(f"\n--- 每模块/章分布 ---")
    mod_counts: dict[str, int] = {}
    for k in kps:
        key = f'{k["module"]} > {k["chapter"]}'
        mod_counts[key] = mod_counts.get(key, 0) + 1
    for key, c in sorted(mod_counts.items()):
        print(f"  {key}: {c} 知识点")

    # 每类型分布
    print(f"\n--- 类型分布 ---")
    for t, c in sorted(types.items(), key=lambda x: -x[1]):
        print(f"  {t}: {c}")

    # 抽样
    print(f"\n--- 前 10 条抽样 ---")
    for kp in kps[:10]:
        print(f"  [{kp['kp_id']}] {kp['kp_type']:20s} | {kp['section_title'][:70]}")
        content_preview = kp["kp_content"][:120].replace("\n", " ")
        print(f"    {content_preview}...")

    # 一致性检查
    ids = [k["kp_id"] for k in kps]
    if len(ids) != len(set(ids)):
        dupes = [id for id in ids if ids.count(id) > 1]
        print(f"\n[ERROR] 重复 KP_ID: {set(dupes)}")
    else:
        print(f"\n[OK] KP_ID 无重复")

    indices = [k["kp_index"] for k in kps]
    expected = list(range(1, len(kps) + 1))
    if indices != expected:
        print(f"[ERROR] KP_INDEX 不连续")
    else:
        print(f"[OK] KP_INDEX 连续 1..{len(kps)}")


# ── 主入口 ─────────────────────────────────────────────────

def clean_markdown_zh(md_text: str) -> str:
    """中文版清洗管线。"""
    print("=" * 60)
    print("第一阶段：清洗 Markdown 格式 (ZH)")
    print("=" * 60)

    md_text = strip_mineru_merge_markers(md_text)

    # 1. 去假 ##
    md_text, n_false = fix_false_headings(md_text)
    print(f"  [FIX] 去除假 ## 标题: {n_false} 处")

    # 2. 补缺案例标题
    n_case = 0
    zh_missing_cases = [
        ("美国司法部新闻稿披露，2019年12月", "案例示例：Tamayo 的金钱骡"),
        ("阿列克谢·科马罗夫通过沃尔科夫工业公司", "案例：科马罗夫战术"),
        ("丹麦最大金融机构", "案例示例：爱沙尼亚银行分行"),
        ("类型学报告，详细说明了", "案例示例：利用类型学报告加强 AML 管控"),
    ]
    for signal, case_title in zh_missing_cases:
        if f"## {case_title}" in md_text or f"### {case_title}" in md_text:
            continue
        idx = md_text.find(signal)
        if idx == -1:
            print(f"  [WARN] 未找到案例信号: {case_title}")
            continue
        line_start = md_text.rfind("\n", 0, idx) + 1
        md_text = md_text[:line_start] + f"## {case_title}\n" + md_text[line_start:]
        n_case += 1
    print(f"  [FIX] 补缺案例标题: {n_case} 处")

    # 3. 补缺章标题
    n_ch = 0
    zh_missing_chapters = [
        ("金融行动特别工作组（第", "## 全球 AFC 标准与指南"),
        ("AFC 计划的组成部分（第", "## AFC 计划的组成部分"),
        ("风险评估的类型（第", "## 风险评估"),
        ("数据收集与准备（第", "## 数据收集与准备"),
    ]
    for signal, heading in zh_missing_chapters:
        if heading in md_text:
            continue
        idx = md_text.find(signal)
        if idx == -1:
            print(f"  [WARN] 未找到章信号: {heading}")
            continue
        insert_pos = idx + 1
        md_text = md_text[:insert_pos] + heading + "\n" + md_text[insert_pos:]
        n_ch += 1
    print(f"  [FIX] 补缺章标题: {n_ch} 处")

    # 4. 修正层级
    md_text, n_level = fix_heading_levels(md_text)
    print(f"  [FIX] 修正标题层级: {n_level} 处")
    print(f"         #  = 模块({len(MODULE_TITLES)})  |  ## = 知识点({len(CHAPTER_TITLES)})  |  ### = 节")

    return md_text


def extract_kps(md_path: Path, lang: str = "en") -> list[dict]:
    """完整的知识点提取管线。"""
    global MODULE_TITLES, CHAPTER_TITLES

    raw_text = md_path.read_text(encoding="utf-8", errors="ignore")
    label = "EN" if lang == "en" else "ZH"
    print(f"\n{'=' * 60}")
    print(f"提取 {label} 知识点: {md_path.name}")
    print(f"{'=' * 60}")
    print(f"原始: {len(raw_text):,} 字符")

    MODULE_TITLES, CHAPTER_TITLES = parse_toc(raw_text, lang)
    print(f"TOC: {len(MODULE_TITLES)} 模块, {len(CHAPTER_TITLES)} 章")

    if lang == "en":
        cleaned = clean_markdown(raw_text)
    else:
        cleaned = clean_markdown_zh(raw_text)

    sections = split_by_h3(cleaned)
    filtered = filter_sections(sections)
    kps = build_numbering(filtered)
    print(f"结果: {len(kps)} 个知识点")
    return kps


# ── 第四阶段：以英文切分为主的双语对齐 ────────────────────

def normalize_alignment_title(text: str) -> str:
    """Normalize titles for deterministic TOC and heading matching."""
    text = unicodedata.normalize("NFKC", text).lower()
    text = text.replace("’", "'")
    text = text.replace("病例示例", "案例示例")
    text = text.replace("case study", "case example")
    text = text.replace("ai-based", "aibased")
    text = text.replace("cross-jurisdictional", "crossjurisdictional")
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", text)


def pdftotext(pdf_path: Path, first_page: int, last_page: int) -> str:
    result = subprocess.run(
        ["pdftotext", "-f", str(first_page), "-l", str(last_page), "-layout", str(pdf_path), "-"],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=True,
    )
    return result.stdout


def extract_pdf_toc_entries(pdf_path: Path, toc_anchor: str) -> list[dict]:
    """Extract ordered title/page entries from the PDF table of contents."""
    text = pdftotext(pdf_path, 5, 20)
    start = text.find(toc_anchor)
    if start == -1:
        raise ValueError(f"Could not locate TOC anchor {toc_anchor!r} in {pdf_path}")

    entries: list[dict] = []
    for match in re.finditer(r"([^\.\f]{1,250}?)\.{3,}\s*(\d+)", text[start:], re.S):
        title = re.sub(r"\s+", " ", match.group(1)).strip()
        title = re.sub(r"^(?:Table of Contents|目录)\s+", "", title)
        if title and len(title) < 180:
            entries.append({"title": title, "page": int(match.group(2))})
    return entries


def prepare_zh_fragments() -> list[dict]:
    """Keep every Chinese body fragment, including Key takeaways and false headings.

    The fragments are only source material. They do not define output row counts.
    """
    global MODULE_TITLES, CHAPTER_TITLES
    raw_text = INPUT_ZH.read_text(encoding="utf-8", errors="ignore")
    MODULE_TITLES, CHAPTER_TITLES = parse_toc(raw_text, "zh")
    cleaned = clean_markdown_zh(raw_text)
    fragments: list[dict] = []
    for sec in split_by_h3(cleaned):
        if sec["module"].strip() == "Front Matter":
            continue
        if sec["title"].strip() == "术语表":
            break
        body = clean_body(sec["body"])
        if not body:
            continue
        fragments.append(
            {
                "fragment_index": len(fragments) + 1,
                "module": sec["module"].strip(),
                "chapter": sec["chapter"].strip(),
                "title": sec["title"].strip(),
                "body": body,
            }
        )
    return fragments


def locate_english_toc_entries(en_kps: list[dict], toc_pairs: list[dict]) -> list[dict]:
    """Map each canonical English section to its ordered bilingual TOC entry."""
    located: list[dict] = []
    cursor = 0
    for kp in en_kps:
        title = kp["section_title"]
        override = TOC_ALIGNMENT_OVERRIDES.get(title)
        if override:
            zh_title, page = override
            located.append({"toc_index": None, "zh_title": zh_title, "page": page})
            continue

        needle = normalize_alignment_title(title)
        toc_index = None
        for i in range(cursor, len(toc_pairs)):
            if normalize_alignment_title(toc_pairs[i]["en_title"]) == needle:
                toc_index = i
                break
        if toc_index is None:
            raise ValueError(f"No ordered English TOC entry for {kp['kp_id']}: {title}")

        cursor = toc_index + 1
        pair = toc_pairs[toc_index]
        located.append(
            {"toc_index": toc_index, "zh_title": pair["zh_title"], "page": pair["page"]}
        )
    return located


def find_chinese_anchors(
    en_kps: list[dict], located: list[dict], fragments: list[dict]
) -> list[dict | None]:
    """Find ordered Chinese Markdown anchors for the canonical English rows."""
    anchors: list[dict | None] = []
    cursor = 0
    for kp, toc in zip(en_kps, located):
        needle = normalize_alignment_title(toc["zh_title"])
        found = None
        for i in range(cursor, len(fragments)):
            if normalize_alignment_title(fragments[i]["title"]) == needle:
                found = i
                break

        if found is None:
            candidates = []
            for i in range(cursor, min(cursor + 100, len(fragments))):
                score = SequenceMatcher(
                    None, needle, normalize_alignment_title(fragments[i]["title"])
                ).ratio()
                candidates.append((score, i))
            if candidates:
                score, candidate = max(candidates)
                if score >= 0.68:
                    found = candidate

        if found is None:
            anchors.append(None)
            continue

        fragment = fragments[found]
        anchors.append(
            {
                "fragment_offset": found,
                "match_title": fragment["title"],
                "match_kind": "exact"
                if normalize_alignment_title(fragment["title"]) == needle
                else "fuzzy",
            }
        )
        cursor = found + 1
    return anchors


def infer_direct_chapter(kp: dict) -> str:
    """Fill the four English rows whose MinerU body lacks an explicit chapter heading."""
    if kp["chapter"]:
        return kp["chapter"]
    if kp["module"] == "Global AFC Frameworks, Governance, and Regulations":
        return "Global AFC Standards and Guidance"
    if kp["module"] == "Use of Guidance and AFC Cooperation":
        return "Use of Guidance and AFC Cooperation"
    raise ValueError(f"Cannot infer chapter for {kp['kp_id']}")


def toc_translation_for_context(
    english_title: str, toc_pairs: list[dict], before: int | None
) -> str:
    needle = normalize_alignment_title(english_title)
    limit = len(toc_pairs) if before is None else before + 1
    matches = [
        pair for pair in toc_pairs[:limit]
        if normalize_alignment_title(pair["en_title"]) == needle
    ]
    if not matches:
        raise ValueError(f"No Chinese TOC translation for context: {english_title}")
    return matches[-1]["zh_title"]


def clean_pdf_fallback(text: str, title: str) -> str:
    """Remove repeated PDF furniture while retaining source-language wording."""
    text = re.sub(r"反洗钱专家认证证书\s*第\s*\d+页", " ", text)
    text = re.sub(r"7\.0版", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    normalized_title = normalize_alignment_title(title)
    normalized_text = normalize_alignment_title(text)
    if normalized_text.startswith(normalized_title):
        # The title itself is metadata and should not be duplicated in the body.
        title_pos = text.find(title)
        if title_pos >= 0:
            text = text[title_pos + len(title):].strip()
    return text


def pdf_fallback_content(
    printed_page: int, next_printed_page: int | None, title: str
) -> tuple[str, str]:
    """Extract a Chinese PDF page interval for source headings absent from Markdown."""
    # Printed study-guide page 13 is PDF page 18 in both split PDFs.
    first_pdf_page = printed_page + 5
    if next_printed_page and next_printed_page > printed_page:
        last_pdf_page = next_printed_page + 4
    else:
        last_pdf_page = first_pdf_page
    text = pdftotext(INPUT_ZH_PDF, first_pdf_page, last_pdf_page)
    return clean_pdf_fallback(text, title), f"PDF pages {first_pdf_page}-{last_pdf_page}"


def write_aligned_workbook(rows: list[dict], audit_rows: list[dict], path: Path) -> None:
    """Write the bilingual primary table and an auditable alignment sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "知识点"
    audit = wb.create_sheet("对齐审计")
    headers = [
        "KP_ID",
        "MODULE_EN", "MODULE_ZH",
        "CHAPTER_EN", "CHAPTER_ZH",
        "SECTION_TITLE_EN", "SECTION_TITLE_ZH",
        "KP_CONTENT_EN", "KP_CONTENT_ZH",
    ]
    audit_headers = [
        "KP_ID", "SECTION_TITLE_EN", "SECTION_TITLE_ZH", "ZH_SOURCE_FRAGMENTS",
        "ZH_SOURCE_LOCATION", "CONTENT_SOURCE", "ALIGNMENT_STATUS",
        "MANUAL_REVIEW_REQUIRED", "EN_CHAR_COUNT", "ZH_CHAR_COUNT", "ZH_EN_RATIO",
    ]
    fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    header_font = Font(bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    side = Side(style="thin")
    border = Border(left=side, right=side, top=side, bottom=side)

    for target, target_headers in ((ws, headers), (audit, audit_headers)):
        for col, header in enumerate(target_headers, start=1):
            cell = target.cell(1, col, header)
            cell.fill = fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

    for row_number, row in enumerate(rows, start=2):
        values = [row[h] for h in headers]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_number, col, value)
            cell.alignment = center_align if col == 1 else wrap_align
            cell.border = border

    for row_number, row in enumerate(audit_rows, start=2):
        values = [row[h] for h in audit_headers]
        for col, value in enumerate(values, start=1):
            cell = audit.cell(row_number, col, value)
            cell.alignment = center_align if col in (1, 7, 8) else wrap_align
            cell.border = border

    for target in (ws, audit):
        target.freeze_panes = "A2"
        target.auto_filter.ref = target.dimensions
    for col, width in {1: 14, 2: 40, 3: 40, 4: 45, 5: 45, 6: 48, 7: 48, 8: 100, 9: 100}.items():
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width
    for col, width in {1: 14, 2: 45, 3: 45, 4: 72, 5: 26, 6: 18, 7: 20, 8: 24, 9: 15, 10: 15, 11: 14}.items():
        audit.column_dimensions[audit.cell(1, col).column_letter].width = width
    wb.save(path)


def build_bilingual_alignment(en_kps: list[dict]) -> tuple[list[dict], list[dict], list[dict]]:
    """Build 308 English-anchored bilingual records and their audit manifest."""
    en_toc = extract_pdf_toc_entries(INPUT_EN_PDF, "Table of Contents")
    zh_toc = extract_pdf_toc_entries(INPUT_ZH_PDF, "目录")
    if len(en_toc) != len(zh_toc):
        raise ValueError(f"TOC count mismatch: EN={len(en_toc)} ZH={len(zh_toc)}")
    toc_pairs = [
        {"en_title": en_entry["title"], "zh_title": zh_entry["title"], "page": en_entry["page"]}
        for en_entry, zh_entry in zip(en_toc, zh_toc)
        if en_entry["page"] == zh_entry["page"]
    ]
    if len(toc_pairs) != len(en_toc):
        raise ValueError("The bilingual PDF TOC page sequences are not aligned")

    located = locate_english_toc_entries(en_kps, toc_pairs)
    fragments = prepare_zh_fragments()
    anchors = find_chinese_anchors(en_kps, located, fragments)
    next_anchor = [None] * len(anchors)
    upcoming = None
    for i in range(len(anchors) - 1, -1, -1):
        next_anchor[i] = upcoming
        if anchors[i] is not None:
            upcoming = anchors[i]["fragment_offset"]

    rows: list[dict] = []
    audit_rows: list[dict] = []
    manifest: list[dict] = []
    assigned_fragments: set[int] = set()
    for i, (kp, toc, anchor) in enumerate(zip(en_kps, located, anchors)):
        chapter_en = infer_direct_chapter(kp)
        before = toc["toc_index"]
        module_zh = toc_translation_for_context(kp["module"], toc_pairs, before)
        chapter_zh = toc_translation_for_context(chapter_en, toc_pairs, before)
        zh_title = toc["zh_title"]
        next_page = located[i + 1]["page"] if i + 1 < len(located) else None

        if anchor is not None:
            start = anchor["fragment_offset"]
            end = next_anchor[i] if next_anchor[i] is not None else len(fragments)
            source_fragments = fragments[start:end]
            for fragment in source_fragments:
                assigned_fragments.add(fragment["fragment_index"])
            zh_content = "\n\n".join(fragment["body"] for fragment in source_fragments).strip()
            source_location = ", ".join(
                f"#{fragment['fragment_index']}:{fragment['title']}" for fragment in source_fragments
            )
            content_source = "Markdown fragments"
            status = "auto_aligned" if anchor["match_kind"] == "exact" else "needs_manual_review"
            review = "No" if status == "auto_aligned" else "Yes"
        else:
            zh_content, source_location = pdf_fallback_content(toc["page"], next_page, zh_title)
            source_fragments = []
            content_source = "Chinese PDF fallback"
            status = "needs_manual_review"
            review = "Yes"

        if not zh_content:
            raise ValueError(f"Empty Chinese content for {kp['kp_id']}")
        row = {
            "KP_ID": kp["kp_id"],
            "MODULE_EN": kp["module"],
            "MODULE_ZH": module_zh,
            "CHAPTER_EN": chapter_en,
            "CHAPTER_ZH": chapter_zh,
            "SECTION_TITLE_EN": kp["section_title"],
            "SECTION_TITLE_ZH": zh_title,
            "KP_CONTENT_EN": kp["kp_content"],
            "KP_CONTENT_ZH": zh_content,
        }
        if any(value in (None, "") for value in row.values()):
            raise ValueError(f"Blank bilingual field for {kp['kp_id']}")
        rows.append(row)
        ratio = round(len(zh_content) / len(kp["kp_content"]), 3)
        if len(source_fragments) > 1 or ratio < 0.10 or ratio > 0.90:
            status = "needs_manual_review"
            review = "Yes"
        audit_rows.append(
            {
                "KP_ID": kp["kp_id"],
                "SECTION_TITLE_EN": kp["section_title"],
                "SECTION_TITLE_ZH": zh_title,
                "ZH_SOURCE_FRAGMENTS": "; ".join(fragment["title"] for fragment in source_fragments),
                "ZH_SOURCE_LOCATION": source_location,
                "CONTENT_SOURCE": content_source,
                "ALIGNMENT_STATUS": status,
                "MANUAL_REVIEW_REQUIRED": review,
                "EN_CHAR_COUNT": len(kp["kp_content"]),
                "ZH_CHAR_COUNT": len(zh_content),
                "ZH_EN_RATIO": ratio,
            }
        )
        manifest.append(
            {
                "kp_id": kp["kp_id"],
                "module_en": kp["module"], "module_zh": module_zh,
                "chapter_en": chapter_en, "chapter_zh": chapter_zh,
                "section_title_en": kp["section_title"], "section_title_zh": zh_title,
                "printed_page": toc["page"],
                "zh_fragments": [
                    {"index": fragment["fragment_index"], "title": fragment["title"]}
                    for fragment in source_fragments
                ],
                "content_source": content_source,
                "source_location": source_location,
                "alignment_status": status,
            }
        )

    if len(rows) != 308 or len({row['KP_ID'] for row in rows}) != len(rows):
        raise ValueError("Canonical English KP_ID validation failed")
    unassigned = [fragment["fragment_index"] for fragment in fragments if fragment["fragment_index"] not in assigned_fragments]
    if unassigned:
        raise ValueError(f"Unassigned Chinese source fragments: {unassigned[:20]}")
    return rows, audit_rows, manifest


def main() -> None:
    en_kps = extract_kps(INPUT_EN, "en")
    print(f"\n{'=' * 60}")
    print("按英文切分重建中英双语知识点")
    print(f"{'=' * 60}")
    rows, audit_rows, manifest = build_bilingual_alignment(en_kps)

    ALIGNMENT_MANIFEST.write_text(
        json.dumps(
            {"version": 1, "english_kp_count": len(rows), "entries": manifest},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    write_aligned_workbook(rows, audit_rows, OUTPUT_ALIGNED_XLSX)

    review_count = sum(row["MANUAL_REVIEW_REQUIRED"] == "Yes" for row in audit_rows)
    print(f"已生成: {OUTPUT_ALIGNED_XLSX}")
    print(f"已生成: {ALIGNMENT_MANIFEST}")
    print(f"主表: {len(rows)} 条英文基准知识点；待人工复核: {review_count} 条")


if __name__ == "__main__":
    main()
