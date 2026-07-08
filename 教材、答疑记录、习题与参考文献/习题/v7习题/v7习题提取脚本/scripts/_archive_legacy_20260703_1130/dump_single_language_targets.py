from __future__ import annotations

import json
from pathlib import Path
import sys

from openpyxl import load_workbook


def dump_workbook(workbook_path: Path) -> dict:
    wb = load_workbook(workbook_path, data_only=True)
    targets = {
        0: [101, 132, 164, 168, 378],
        1: [124, 221, 284, 308],
    }
    out = {}
    for sheet_index, seqs in targets.items():
        ws = wb.worksheets[sheet_index]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value in seqs:
                rows.append(
                    {
                        "excel_row": r,
                        "values": {
                            headers[c - 1]: ws.cell(r, c).value
                            for c in range(1, ws.max_column + 1)
                        },
                    }
                )
        out[ws.title] = rows
    return out


def dump_frames(json_dir: Path) -> dict:
    specs = {
        "cn101": ("中文版2_frames.json", range(1, 13)),
        "cn132": ("中文版2_frames.json", range(79, 94)),
        "cn164": ("中文版2_frames.json", range(168, 174)),
        "cn168": ("中文版2_frames.json", range(173, 181)),
        "cn378": ("中文版4_frames.json", range(214, 220)),
        "en124": ("英文版2_frames.json", range(68, 76)),
        "en221": ("英文版3_frames.json", range(54, 60)),
        "en284": ("英文版3_frames.json", range(250, 265)),
        "en308": ("英文版4_frames.json", range(16, 31)),
    }
    out = {}
    for key, (name, indices) in specs.items():
        path = json_dir / name
        if not path.exists():
            continue
        frames = json.loads(path.read_text(encoding="utf-8"))
        by_index = {int(item["index"]): item for item in frames}
        selected = []
        for index in indices:
            item = by_index.get(index)
            if not item:
                continue
            selected.append(
                {
                    "index": item.get("index"),
                    "time_sec": item.get("time_sec"),
                    "page_no": item.get("page_no"),
                    "answer": item.get("answer"),
                    "has_explanation": item.get("has_explanation"),
                    "clean_text": item.get("clean_text"),
                    "image": item.get("image"),
                }
            )
        out[key] = selected
    return out


def main() -> None:
    out_dir = Path(sys.argv[1])
    key = sys.argv[2] if len(sys.argv) > 2 else ""
    workbook = out_dir / "semantic_aligned_cn_en.xlsx"
    result = {
        "workbook": dump_workbook(workbook),
        "frames": dump_frames(out_dir / "json"),
    }
    if key:
        if key in result["frames"]:
            result = {key: result["frames"][key]}
        else:
            result = {
                sheet: [
                    row
                    for row in rows
                    if str(row["values"].get("题号")) == key
                ]
                for sheet, rows in result["workbook"].items()
            }
    print(json.dumps(result, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
