from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def compact(value) -> str:
    return " ".join(str(value or "").split())


def norm_answer(value) -> str:
    return "".join(ch for ch in str(value or "").upper() if ch in "ABCDEF")


def answer_status(cn_answer, en_answer) -> str:
    cn = norm_answer(cn_answer)
    en = norm_answer(en_answer)
    if cn and en and cn == en:
        return "一致"
    if cn and en:
        return f"不一致: 中文={cn}; 英文={en}"
    if cn or en:
        return "单侧识别"
    return "未识别"


def clean_analysis(text: str) -> str:
    text = compact(text)
    if not text:
        return ""
    starts = [text.find(marker) for marker in ("试题详解", "原解析") if text.find(marker) >= 0]
    if starts:
        text = text[min(starts) :]
    for marker in (
        "使用【深度解题】",
        "使用【相",
        "2026新真题",
        "2026新直",
        "来 源",
        "来源",
        "考友笔记",
        "写笔记",
        "0/395",
        "分享",
        "已做题",
        "收藏",
    ):
        pos = text.find(marker)
        if pos > 60:
            text = text[:pos]
    return compact(text)


def load_frames(out_dir: Path) -> dict[str, list[dict]]:
    frames_by_video = {}
    for path in (out_dir / "json").glob("*_frames.json"):
        frames_by_video[path.stem.replace("_frames", "")] = json.loads(path.read_text(encoding="utf-8"))
    return frames_by_video


def get_frame(frames_by_video: dict[str, list[dict]], video_stem: str, index: int) -> dict:
    for frame in frames_by_video[video_stem]:
        if int(frame.get("index") or -1) == index:
            return frame
    raise KeyError(f"{video_stem}#{index} not found")


def frame_text(frames_by_video: dict[str, list[dict]], video_stem: str, index: int) -> str:
    frame = get_frame(frames_by_video, video_stem, index)
    return compact(frame.get("clean_text") or frame.get("raw_text") or "")


def frame_image(frames_by_video: dict[str, list[dict]], video_stem: str, index: int) -> str:
    return str(get_frame(frames_by_video, video_stem, index).get("image") or "")


def set_link(cell, path: str) -> None:
    if path:
        cell.value = path
        cell.hyperlink = path
        cell.style = "Hyperlink"


def write_sheet(wb, title: str, headers: list[str], rows: list[list]) -> None:
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [12, 12, 12, 14, 20, 46, 46, 72, 72, 18]
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1] if col <= len(widths) else 24
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


ANSWER_FIXES = {
    ("中文题库", 57): "B",
    ("中文题库", 243): "ABDE",
    ("中文题库", 295): "BC",
    ("中文题库", 352): "BC",
    ("中文题库", 354): "BCE",
    ("中文题库", 355): "AD",
    ("英文题库", 76): "A",
    ("英文题库", 103): "B",
    ("英文题库", 131): "B",
    ("英文题库", 156): "B",
    ("英文题库", 206): "D",
    ("英文题库", 238): "BE",
    ("英文题库", 271): "ACF",
    ("英文题库", 272): "ACF",
    ("英文题库", 298): "BC",
    ("英文题库", 312): "AD",
    ("英文题库", 346): "BC",
    ("英文题库", 350): "ACDE",
    ("英文题库", 353): "BCE",
    ("英文题库", 357): "BCDE",
    ("英文题库", 360): "AC",
    ("英文题库", 370): "BCE",
}


QUESTION_FRAME_FIXES = {
    ("中文题库", 377): ("中文版4", 216),
    ("英文题库", 131): ("英文版2", 95),
    ("英文题库", 156): ("英文版2", 180),
    ("英文题库", 157): ("英文版2", 184),
    ("英文题库", 354): ("英文版4", 149),
    ("英文题库", 371): ("英文版4", 205),
}


ANALYSIS_FRAME_FIXES = {
    ("中文题库", 55): ("中文版1", 154),
    ("中文题库", 377): ("中文版4", 217),
    ("英文题库", 76): ("英文版1", 253),
    ("英文题库", 88): ("英文版1", 295),
    ("英文题库", 103): ("英文版2", 7),
    ("英文题库", 120): ("英文版2", 64),
    ("英文题库", 121): ("英文版2", 64),
    ("英文题库", 123): ("英文版2", 72),
    ("英文题库", 124): ("英文版2", 74),
    ("英文题库", 130): ("英文版2", 94),
    ("英文题库", 131): ("英文版2", 95),
    ("英文题库", 156): ("英文版2", 183),
    ("英文题库", 157): ("英文版2", 185),
    ("英文题库", 179): ("英文版2", 245),
    ("英文题库", 206): ("英文版3", 21),
    ("英文题库", 218): ("英文版3", 50),
    ("英文题库", 238): ("英文版3", 113),
    ("英文题库", 271): ("英文版3", 225),
    ("英文题库", 284): ("英文版3", 259),
    ("英文题库", 285): ("英文版3", 262),
    ("英文题库", 298): ("英文版3", 312),
    ("英文题库", 340): ("英文版4", 108),
    ("英文题库", 346): ("英文版4", 122),
    ("英文题库", 350): ("英文版4", 139),
    ("英文题库", 354): ("英文版4", 150),
    ("英文题库", 357): ("英文版4", 157),
    ("英文题库", 360): ("英文版4", 165),
    ("英文题库", 363): ("英文版4", 173),
    ("英文题库", 364): ("英文版4", 177),
    ("英文题库", 370): ("英文版4", 204),
    ("英文题库", 371): ("英文版4", 207),
}


MANUAL_REVIEW_ROWS = [
    ["中文题库", 4, "视频源内答案/解析冲突", "题干答案帧明确显示正确答案 D；同题解析帧却写第319(b)条款并称正确选项为 C。已按视频答案字段保留 D，解析不自动改。", "中文版1#19-20"],
    ["英文题库", 120, "疑似跨屏拆题", "EN120/EN121 是同一道 PPP 题的跨屏内容；已统一回填同一解析，但是否保留两行需人工决策。", "英文版2#58-64"],
    ["英文题库", 121, "疑似跨屏拆题", "EN120/EN121 是同一道 PPP 题的跨屏内容；已统一回填同一解析，但是否保留两行需人工决策。", "英文版2#58-64"],
    ["英文题库", 271, "疑似跨屏拆题", "EN271/EN272 是同一道阈值调整题的跨屏内容；EN271 已修为完整题和答案 ACF，但是否合并/删除 EN272 需人工决策。", "英文版3#222-225"],
    ["英文题库", 272, "疑似跨屏拆题", "EN271/EN272 是同一道阈值调整题的跨屏内容；EN272 已修为 ACF，但是否保留两行需人工决策。", "英文版3#222-225"],
    ["英文题库", 322, "答案/解析冲突", "视频答案帧显示 ABCDE，但后续解析文字又写 AD，且题干写 Select Five；需人工按截图决定。", "英文版4#61-63"],
]


MANUAL_TEXT_FIXES = {
    ("英文题库", 156, "解析"): (
        "试题详解 原解析 在反洗钱和反恐融资(AML/CFT)合规计划中,员工怀疑同事涉及金融投资诈骗时,"
        "应通过金融机构的举报渠道报告。直接质问同事、警告同事或客户、或仅向直属经理报告,"
        "均可能干扰调查或超出员工权限并涉及法律风险。通过举报渠道报告既保护举报人,又有利于机构按程序调查。"
        "因此,正确答案为B."
    ),
    ("英文题库", 157, "题干"): (
        "单选 -[AML/CFTCompliancePrograms] Which key metric provides senior management information about "
        "the effectiveness of its AML controls? The ratio of true positives to false positives generated by "
        "the automated monitoring system. The number of money laundering alerts generated by the watchlist "
        "screening system. The number of high-risk customers onboarded each month. The number of clients "
        "exited for commercial reasons. 正确答案A您选择/ 试题详解 原解析"
    ),
    ("中文题库", 13, "解析"): (
        "试题详解 原解析 该律师事务所设立复杂离岸公司架构，对资金来源文件记录很少，且不质疑客户交易、"
        "不报告可疑活动。选项A规避费用与洗钱风险无关；选项B税务筹划合法避税，非洗钱风险；"
        "选项C称常规法律服务无需审查，与事实不符，未体现洗钱风险。选项D指出律师事务所充当“看门人”，"
        "通过复杂公司架构为非法资金转移和隐匿提供便利，符合洗钱风险特征，故答案选D。"
        "易错提醒：注意区分合法税务筹划与非法洗钱行为。"
    ),
    ("中文题库", 36, "解析"): (
        "试题详解 原解析 选项A中，私人控股公司通过赌场将资金转入多个赌客的投注账户，"
        "此行为涉及资金分散转移，意图掩盖资金来源和流向，符合洗钱行为中“放置-离析-融合”的离析阶段特征，"
        "通过复杂交易使资金难以追溯，是洗钱的最有力迹象。选项B仅涉及资金转移至另一运营商，无复杂交易掩盖；"
        "选项C用现金买筹码属正常行为；选项D多次下注同一赛事也属正常。故选项A正确。"
        "易错提醒：注意区分正常交易与洗钱行为的界限，关注资金来源和流向的复杂性。"
    ),
    ("英文题库", 130, "题干"): (
        "单选 -[Risks and Methods of Money Laundering and Terrorist Financing] Correspondent banking is considered "
        "a high-risk area because correspondent banking transactions: A Are made primarily to and from high-risk jurisdictions "
        "B Can be made anonymously and without beneficial ownership information C Typically include less information than domestic payments "
        "D Are made cross-border and on behalf of third parties 正确答案D您选择/ 试题详解 原解析"
    ),
    ("英文题库", 131, "解析"): (
        "试题详解 原解析 在反洗钱和反恐融资(AML/CFT)合规计划中，与政治公众人物(Politically Exposed Person, PEP)"
        "建立业务关系需严格审批。KYC分析师负责客户身份识别，关系经理负责客户关系维护，强化尽职调查合规官负责深入调查，"
        "但均不承担最终审批责任。根据合规管理原则，此类高风险决策需由具备足够权限和责任心的主体承担，"
        "高级管理层(Senior management)具备整体风险控制和决策权，是最终审批者。因此，答案选B。"
    ),
    ("英文题库", 132, "解析"): (
        "试题详解 原解析 在执行反洗钱(AML)调查时，审计员需遵循保护隐私和数据的原则。"
        "选项A指出AML和数据保护隐私法不应相互排斥，这符合数据保护和合规调查的平衡原则；"
        "选项B称证据收集时隐私法不如本地AML法重要，忽视了隐私保护的必要性；"
        "选项C认为恐怖主义融资在数据保护中更相关且优先于法律，违背了法律平等原则；"
        "选项D建议国家明确AML和数据保护法不平衡之处，未直接回答审计人员调查时应遵循的原则。"
        "因此，A项最符合审计员在执行AML调查时应遵循的隐私和数据保护原则。"
    ),
    ("英文题库", 175, "题干"): (
        "单选 -[AML/CFTCompliancePrograms] During customer due diligence, an insurance company discovers that "
        "the policyholder's business address and the ultimate beneficial owner (UBO) changed two weeks ago. "
        "What actions should be taken immediately? A Investigate the changes of address and change of UBO and "
        "in the meantime decline payment and withdrawal instructions from the policy until completion of the investigation "
        "and next steps are agreed upon B Request the relationship manager set up a meeting with the policyholder to "
        "update their address and submit details of the new UBO in the name of good customer service C Investigate the "
        "changes of address and change of UBO and in the meantime freeze the client's policy D File a suspicious transaction "
        "report because the insurance company was not made aware of the business' change of UBO 正确答案A您选择/ 试题详解 原解析"
    ),
    ("英文题库", 175, "解析"): (
        "试题详解 原解析 在反洗钱和反恐融资合规计划中，保险公司在客户尽职调查(CDD)过程中发现保单持有人的业务地址"
        "且最终受益所有人(UBO)发生变更，可能涉及洗钱风险。根据合规要求，应立即调查地址变更和UBO变更的原因，"
        "同时暂停支付和提款指令，直至调查完成并确定后续步骤。选项A既进行了必要调查，又采取临时控制措施以防止潜在风险。"
        "其他选项要么措施不足(B)，要么过于激进(C)，要么忽视风险(D)。"
    ),
    ("英文题库", 265, "题干"): (
        "多选 -[Sanctions Compliance and Screening] Using artificial intelligence (AI) and machine learning-based techniques "
        "in adverse media screening can: (Choose three.) A ensure that all adverse media sources are comprehensively analyzed "
        "without the need for human review. B significantly reduce human errors arising from repetitive tasks by delivering "
        "consistent and highly accurate analysis. C instantly identify intent behind media articles, allowing for more effective "
        "risk scoring. D automate the process of identifying new information and distinguishing it from previously encountered data. "
        "E cover multiple languages and scripts, surpassing the limitations of human linguistics. 正确答案BDE您选择/ 试题详解 原解析"
    ),
    ("英文题库", 265, "解析"): (
        "试题详解 原解析 选项B，人工智能和机器学习能通过一致且高度准确的分析，显著减少重复性任务中的人为错误。"
        "选项D，其可自动化识别新信息，并区分于已遇数据，提升筛选效率。选项E，人工智能和机器学习能处理多语言和脚本，"
        "突破人类语言能力的局限。选项A错误，因为人工智能虽能分析大量数据，但仍需人类审阅以确保准确性。"
        "选项C错误，因为人工智能能辅助分析文本，但准确识别意图仍具挑战，需结合其他方法。因此，正确答案为BDE。"
        "易错提醒：注意人工智能和机器学习在处理复杂任务时的局限性。"
    ),
    ("英文题库", 347, "解析"): (
        "试题详解 原解析 FATF在关于环境犯罪洗钱风险的研究报告中指出，环境犯罪常通过合法行业(如伐木、采矿)的异常资金流动掩盖非法所得。"
        "选项B涉及非法收益通过频繁支付给无关方实现转移，与虚假交易模式吻合；选项C指向利用职位管理自然资源的官员隐藏非法财富，"
        "符合腐败与资源类犯罪关联的风险特征。两者均直接关联FATF报告列举的典型洗钱手段。"
        "选项A描述的交易与授权范围一致，缺乏异常性；选项D的小额现金流动可能属常规业务，未被列为高风险信号。"
    ),
    ("英文题库", 353, "题干"): (
        "多选 -[Risks and Methods of Money Laundering and Terrorist Financing] Professional service providers, including lawyers, "
        "accountants, investment brokers, and other third parties, may abuse their positions to facilitate money laundering. "
        "Which financial crime risks are associated with this type of abuse? (Select Three.) A Opening an account to settle an estate "
        "on behalf of a client B Establishing shell companies to enable money laundering activities, including placement or layering "
        "C Opening third-party accounts for the primary purpose of masking the underlying client's identity D Opening a trust account "
        "to facilitate a legitimate real estate transaction E Directing or facilitating the laundering of illicit funds, including structuring transactions "
        "正确答案BCE您选择/ 试题详解 原解析"
    ),
    ("英文题库", 353, "解析"): (
        "试题详解 原解析 本题聚焦专业服务提供者滥用职位助长洗钱时涉及的金融犯罪风险。"
        "选项B，建立空壳公司进行洗钱活动如放置或分层，属于相关风险；"
        "选项C，以掩盖客户身份为主要目的开设第三方账户，说明犯罪意图和不透明资金来源、流向，与洗钱风险紧密相关；"
        "选项E，指导或促成非法资金洗钱，包括结构化交易，直接体现了利用职位便利进行洗钱犯罪。"
        "选项A为客户代理账户用于遗产，选项D开立信托账户用于合法房地产交易，均属正常业务，不体现洗钱风险。"
        "因此，正确答案BCE。易错提醒：注意区分正常业务与滥用职位助长洗钱的行为。"
    ),
    ("英文题库", 370, "题干"): (
        "多选 -[AML/CFTCompliancePrograms] A compliance officer at a financial institution (FI) is reviewing a new client application "
        "for a Virtual Asset Service Provider (VASP). Which details should be part of the risk assessment to determine whether the customer "
        "falls within the FI's financial crimes risk appetite? (Select Three.) A Whether the VASP's procedures are sufficient for protecting "
        "client personally identifiable information B What percentage of the VASP's clients are classified as higher-risk C Which registered "
        "institutions act on behalf of the VASP as operators of virtual asset wallets and virtual asset exchange offices D Whether the VASP "
        "is utilizing central bank digital currencies E Who the VASP's clients are, including the breakdown of foreign and domestic individuals "
        "正确答案BCE您选择/ 试题详解 原解析"
    ),
    ("英文题库", 370, "解析"): (
        "试题详解 原解析 在评估虚拟资产服务提供商(VASP)是否符合金融机构的反洗钱和反恐融资(AML/CFT)风险偏好时，"
        "需关注其客户风险特征及业务关联性。选项B(高风险客户占比)可反映VASP整体风险水平；"
        "选项C(合作注册机构信息)有助于识别潜在风险传导路径；选项E(客户身份及地域分布)可帮助判断客户基础风险。"
        "选项A关注数据保护，选项D关注是否使用央行数字货币，均非本题金融犯罪风险评估的核心要素。"
        "因此，正确答案为BCE。易错提醒：需区分数据安全与金融犯罪风险评估的核心要素。"
    ),
    ("英文题库", 371, "题干"): (
        "多选 -[AML/CFTCompliancePrograms] A deficiency in the design of a bank's AML/CFT compliance program could result in placing "
        "individual accountability on which part of the regulated entity? (Select Two.) A The compliance department B Board of directors "
        "C Product oversight committee D Senior management 正确答案BD您选择/ 试题详解 原解析"
    ),
}


SAMPLE_REVIEW_ROWS = [
    ["中文题库", 57, "答案抽测", "视频解析明确为选项B；OCR 将 B 误读成“日”，已修复为 B。", "通过/已修"],
    ["中文题库", 295, "答案抽测", "解析明确“因此选BC”，原答案 ADE 串入邻题，已修复为 BC。", "通过/已修"],
    ["中文题库", 377, "题干抽测", "原题干混入上一题联合国制裁内容；用清晰题干帧重建。", "通过/已修"],
    ["英文题库", 76, "答案抽测", "解析明确选项A正确，原答案 D 错；已修复为 A。", "通过/已修"],
    ["英文题库", 103, "答案抽测", "API 题答案帧为 B，原答案 A 错；已修复为 B。", "通过/已修"],
    ["英文题库", 131, "题干/答案抽测", "原行串到下一题隐私调查题；用 PEP 审批责任题帧重建，答案 B。", "通过/已修"],
    ["英文题库", 206, "答案抽测", "负面媒体 AI/ML 筛查题答案帧为 D，原答案 B 错；已修复为 D。", "通过/已修"],
    ["英文题库", 238, "答案抽测", "会计师 AML/CFT 风险识别题答案帧为 BE，原答案 AC 错；已修复为 BE。", "通过/已修"],
    ["英文题库", 264, "视觉抽测", "标准答案字段未识别，但截图解析写“因此，选择BC”。", "通过/已确认"],
    ["英文题库", 320, "视觉抽测", "标准答案字段未识别，但截图解析写“正确选项是B和C”。", "通过/已确认"],
    ["英文题库", 322, "冲突抽测", "答案帧显示 ABCDE，后续解析写 AD，且题干为 Select Five；不自动改。", "需人工审核"],
    ["英文题库", 271, "结构抽测", "EN271/EN272 同一阈值调整题被拆成两行；不自动合并。", "需人工审核"],
    ["中文题库", 13, "细抽帧复核", "原 2 秒帧解析混入第14题；0.25秒细抽帧抓到完整解析，已修复。", "通过/已修"],
    ["中文题库", 36, "细抽帧复核", "原 2 秒帧解析混入第37题；0.25秒细抽帧抓到完整解析，已修复。", "通过/已修"],
    ["英文题库", 175, "细抽帧复核", "题目为保险CDD中地址和UBO变更，原解析串到会计师题；已按视频解析修复。", "通过/已修"],
    ["英文题库", 265, "细抽帧复核", "AI/ML负面媒体筛查题题干开头和解析缺失；细抽帧补齐，答案BDE。", "通过/已修"],
    ["英文题库", 347, "细抽帧复核", "环境犯罪洗钱风险题答案BC，原解析串到加密货币题；已修复解析。", "通过/已修"],
    ["英文题库", 353, "错位复核", "专业服务提供者滥用职位题答案BCE，原解析串到354题；已重建题干和解析。", "通过/已修"],
    ["英文题库", 354, "错位复核", "高级管理层参与题答案AD，原题干抓到353题尾部；已重建题干和解析。", "通过/已修"],
    ["英文题库", 370, "错位复核", "VASP风险评估题答案BCE，原解析串到371题；已修复题干、答案和解析。", "通过/已修"],
    ["英文题库", 371, "错位复核", "AML/CFT计划设计缺陷责任题答案BD，原题干抓到370题尾部；已重建题干。", "通过/已修"],
]


def sync_alignment(wb, sheet_name: str, qno: int, answer: str | None, analysis: str | None, question: str | None) -> int:
    ws = wb["语义对齐"]
    is_cn = sheet_name == "中文题库"
    q_col = 9 if is_cn else 10
    answer_col = 11 if is_cn else 12
    question_col = 13 if is_cn else 14
    analysis_col = 15 if is_cn else 16
    count = 0
    for row in range(2, ws.max_row + 1):
        try:
            if int(ws.cell(row, q_col).value) != qno:
                continue
        except (TypeError, ValueError):
            continue
        if answer:
            ws.cell(row, answer_col).value = answer
        if question:
            ws.cell(row, question_col).value = question
        if analysis:
            ws.cell(row, analysis_col).value = analysis
        ws.cell(row, 8).value = answer_status(ws.cell(row, 11).value, ws.cell(row, 12).value)
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
        count += 1
    return count


def main() -> None:
    base_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parents[1]
    out_dir = base_dir / "output_2s"
    source = out_dir / "semantic_aligned_cn_en_filled_missing_v2.xlsx"
    target = out_dir / "semantic_aligned_cn_en_reviewed_v4.xlsx"
    frames_by_video = load_frames(out_dir)
    wb = load_workbook(source)

    fix_rows: list[list] = []

    for (sheet_name, qno), new_answer in ANSWER_FIXES.items():
        ws = wb[sheet_name]
        row = qno + 1
        old = ws.cell(row, 3).value
        if norm_answer(old) != norm_answer(new_answer):
            ws.cell(row, 3).value = new_answer
            synced = sync_alignment(wb, sheet_name, qno, new_answer, None, None)
            fix_rows.append([sheet_name, qno, "答案", old or "", new_answer, "", "", "解析/正确答案帧明确指向该答案", synced])

    for (sheet_name, qno), (video, frame_idx) in QUESTION_FRAME_FIXES.items():
        ws = wb[sheet_name]
        row = qno + 1
        old = ws.cell(row, 6).value
        new = frame_text(frames_by_video, video, frame_idx)
        if new and compact(old) != new:
            ws.cell(row, 6).value = new
            image = frame_image(frames_by_video, video, frame_idx)
            set_link(ws.cell(row, 9), image)
            synced = sync_alignment(wb, sheet_name, qno, None, None, new)
            fix_rows.append([sheet_name, qno, "题干", compact(old)[:180], new[:180], image, f"{video}#{frame_idx}", "用清晰题干帧替换串题/半截题干", synced])

    for (sheet_name, qno), (video, frame_idx) in ANALYSIS_FRAME_FIXES.items():
        ws = wb[sheet_name]
        row = qno + 1
        old = ws.cell(row, 7).value
        text = frame_text(frames_by_video, video, frame_idx)
        new = clean_analysis(text) or text
        if new and compact(old) != new:
            ws.cell(row, 7).value = new
            image = frame_image(frames_by_video, video, frame_idx)
            set_link(ws.cell(row, 10), image)
            synced = sync_alignment(wb, sheet_name, qno, None, new, None)
            fix_rows.append([sheet_name, qno, "解析", compact(old)[:180], new[:180], image, f"{video}#{frame_idx}", "用对应解析帧替换串题/错位解析", synced])

    for (sheet_name, qno, field), new in MANUAL_TEXT_FIXES.items():
        ws = wb[sheet_name]
        row = qno + 1
        col = 6 if field == "题干" else 7
        old = ws.cell(row, col).value
        if compact(old) != compact(new):
            ws.cell(row, col).value = new
            synced = sync_alignment(
                wb,
                sheet_name,
                qno,
                None,
                new if field == "解析" else None,
                new if field == "题干" else None,
            )
            fix_rows.append([sheet_name, qno, field, compact(old)[:180], new[:180], "", "人工复核帧文本", "根据相邻清晰帧和解析文字去除串题残留", synced])

    for sheet_name in ("中文题库", "英文题库"):
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")

    # Recalculate all answer status labels after fixes, including rows not touched above.
    align_ws = wb["语义对齐"]
    for row in range(2, align_ws.max_row + 1):
        align_ws.cell(row, 8).value = answer_status(align_ws.cell(row, 11).value, align_ws.cell(row, 12).value)
        for col in range(1, align_ws.max_column + 1):
            align_ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")

    risk_rows: list[list] = []
    for row in range(2, align_ws.max_row + 1):
        confidence = align_ws.cell(row, 1).value
        status = align_ws.cell(row, 8).value
        try:
            qdiff = int(align_ws.cell(row, 6).value or 0)
        except (TypeError, ValueError):
            qdiff = 0
        reasons = []
        if confidence == "低":
            reasons.append("低置信")
        if status != "一致":
            reasons.append("答案状态不一致")
        if qdiff > 5:
            reasons.append("题号差>5")
        if reasons:
            risk_rows.append(
                [
                    align_ws.cell(row, 9).value,
                    align_ws.cell(row, 10).value,
                    confidence,
                    align_ws.cell(row, 3).value,
                    qdiff,
                    status,
                    "；".join(reasons),
                    compact(align_ws.cell(row, 13).value)[:220],
                    compact(align_ws.cell(row, 14).value)[:220],
                ]
            )

    audit_rows = [
        ["完整性", "中文题库", "395题，题干/答案/解析空值=0"],
        ["完整性", "英文题库", "395题，题干/答案/解析空值=0"],
        ["单语答案证据", "中文题库", "395/395 可在视频 OCR 帧附近找到答案证据"],
        ["单语答案证据", "英文题库", "393/395 标准答案帧直接支持；EN264/EN320 经视觉复核由解析文字确认"],
        ["本轮修复", "答案/题干/解析", f"写入修复记录 {len(fix_rows)} 条"],
        ["仍需人工审核", "结构性问题", f"列入 {len(MANUAL_REVIEW_ROWS)} 条"],
        ["中英对齐风险", "语义对齐", f"低置信/答案不一致/题号差大共 {len(risk_rows)} 条"],
    ]

    write_sheet(wb, "质量复核总览", ["类别", "范围", "结论"], audit_rows)
    write_sheet(wb, "质量复核修复记录", ["工作表", "题号", "字段", "旧值摘要", "新值摘要", "证据截图", "来源帧", "备注", "同步行数"], fix_rows)
    write_sheet(wb, "需人工审核清单", ["工作表", "题号", "问题类型", "说明", "证据帧"], MANUAL_REVIEW_ROWS)
    write_sheet(wb, "抽测记录", ["工作表", "题号", "抽测类型", "结论", "状态"], SAMPLE_REVIEW_ROWS)
    write_sheet(wb, "中英对齐风险清单", ["中文题号", "英文题号", "置信", "分数", "题号差", "答案状态", "原因", "中文题干摘要", "英文题干摘要"], risk_rows)

    wb.save(target)
    print(target)
    print("fix_rows", len(fix_rows))
    print("manual_review", len(MANUAL_REVIEW_ROWS))
    print("alignment_risk", len(risk_rows))


if __name__ == "__main__":
    main()
