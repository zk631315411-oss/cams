from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(r"D:\守正公司工作区\cams考试\教材、答疑记录、习题与参考文献\习题\v7习题")
AUDIT_DIR = BASE / "output_2s" / "v8_ocr_audit"
OUTPUT = AUDIT_DIR / "v8_ocr_audit_summary.xlsx"


def compact(value: Any, limit: int | None = None) -> str:
    text = " ".join(str(value or "").split())
    return text if limit is None else text[:limit]


def parse_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not path.exists():
        return rows
    for line_no, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        line = line.strip()
        if not line:
            continue
        try:
            item = json.loads(line)
            item["_source_file"] = path.name
            item["_line_no"] = line_no
            rows.append(item)
        except json.JSONDecodeError:
            rows.append(
                {
                    "cn_qno": "",
                    "en_qno": "",
                    "side": "",
                    "field": "",
                    "severity": "parse_error",
                    "issue_type": "报告JSON解析失败",
                    "evidence_excerpt": line[:500],
                    "suggested_fix": "",
                    "confidence": "",
                    "needs_screenshot": True,
                    "_source_file": path.name,
                    "_line_no": line_no,
                }
            )
    return rows


def main() -> None:
    report_files = sorted(AUDIT_DIR.glob("report_cn*.jsonl"))
    # The first CN101-200 audit was read with a wrong code page by a sub-agent,
    # so it over-reported normal Chinese text as mojibake. Prefer the rerun.
    if (AUDIT_DIR / "report_cn101_200_rerun.jsonl").exists():
        report_files = [p for p in report_files if p.name != "report_cn101_200.jsonl"]
    rows: list[dict[str, Any]] = []
    for path in report_files:
        rows.extend(parse_jsonl(path))
    rows = [row for row in rows if compact(row.get("issue_type")) != "qno_misalignment"]

    wb = Workbook()
    ws = wb.active
    ws.title = "子代理问题汇总"
    headers = [
        "中文题号",
        "英文题号",
        "侧别",
        "字段",
        "严重度",
        "问题类型",
        "证据摘录",
        "建议修法",
        "置信度",
        "需截图复核",
        "来源报告",
        "报告行号",
    ]
    ws.append(headers)
    for item in rows:
        ws.append(
            [
                item.get("cn_qno", ""),
                item.get("en_qno", ""),
                item.get("side", ""),
                item.get("field", ""),
                item.get("severity", ""),
                item.get("issue_type", ""),
                compact(item.get("evidence_excerpt", ""), 1000),
                compact(item.get("suggested_fix", ""), 1000),
                item.get("confidence", ""),
                item.get("needs_screenshot", ""),
                item.get("_source_file", ""),
                item.get("_line_no", ""),
            ]
        )

    fill = PatternFill("solid", fgColor="D9EAF7")
    widths = [10, 10, 10, 14, 12, 24, 70, 70, 10, 14, 28, 10]
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("统计")
    summary.append(["项目", "数量"])
    summary.append(["报告文件数", len(report_files)])
    summary.append(["问题总数", len(rows)])
    for key, count in Counter(compact(row.get("severity")) for row in rows).most_common():
        summary.append([f"严重度={key}", count])
    for key, count in Counter(compact(row.get("issue_type")) for row in rows).most_common():
        summary.append([f"问题类型={key}", count])
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    summary.column_dimensions["A"].width = 40
    summary.column_dimensions["B"].width = 12

    wb.save(OUTPUT)
    print(f"reports: {len(report_files)}")
    print(f"issues: {len(rows)}")
    print(f"written: {OUTPUT}")


if __name__ == "__main__":
    main()
