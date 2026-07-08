from __future__ import annotations

import argparse
import math
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


CN_SHEET = "中文合并"
EN_SHEET = "英文合并"

NOISE_TOKENS = [
    "正确答案",
    "您选择",
    "试题详解",
    "原解析",
    "难度",
    "来源",
    "考友笔记",
    "纠错",
    "2026新真题",
    "中文",
    "英文",
    "单选",
    "多选",
    "折题",
    "新题",
    "翻译",
    "选择两项",
    "选择三项",
]

CONCEPTS: dict[str, list[str]] = {
    "AML": ["反洗钱", "AML", "ANTI-MONEY"],
    "CFT": ["恐怖主义融资", "反恐怖融资", "资助恐怖主义", "CFT", "TERRORIST"],
    "CDD": ["客户尽职调查", "尽职调查", "CDD", "DUE"],
    "EDD": ["强化尽职调查", "EDD", "ENHANCED"],
    "KYC": ["了解您的客户", "KYC"],
    "SAR": ["可疑活动报告", "SAR", "STR"],
    "FIU": ["金融情报", "FIU", "FINANCIALINTELLIGENCE"],
    "FATF": ["金融行动", "FATF"],
    "OFAC": ["海外资产控制", "OFAC"],
    "PEP": ["政治公众", "PEP"],
    "BSA": ["银行保密法", "BSA"],
    "FINCEN": ["金融犯罪执法", "FINCEN"],
    "EGMONT": ["埃格蒙特", "EGMONT"],
    "WOLFSBERG": ["沃尔夫斯堡", "狼堡", "WOLFSBERG"],
    "SANCTIONS": ["制裁", "SANCTION"],
    "CRYPTO": ["加密资产", "虚拟货币", "区块链", "CRYPTO", "VIRTUALCURRENCY", "BLOCKCHAIN"],
    "RISK": ["风险评估", "风险偏好", "RISK"],
    "PRIVACY": ["隐私", "数据保护", "PRIVACY", "DATA"],
    "PROLIFERATION": ["扩散融资", "PROLIFERATION"],
    "INSURANCE": ["保险", "INSURANCE"],
    "CASINO": ["赌场", "赌博", "CASINO", "GAMBL"],
    "TCSP": ["信托和公司服务", "TCSP", "TRUST"],
    "MSB": ["货币服务", "MSB"],
    "OSINT": ["开源情报", "OSINT"],
    "EMPLOYEE": ["员工", "雇员", "EMPLOYEE", "STAFF"],
    "VENDOR": ["供应商", "VENDOR", "SUPPLIER"],
    "THIRD_PARTY": ["第三方", "THIRDPARTY", "THIRD-PARTY"],
    "UNAUTHORIZED": ["未经授权", "UNAUTHORIZED"],
    "SENSITIVE_INFO": ["敏感客户信息", "SENSITIVE"],
    "API": ["应用程序编程接口", "API"],
    "CORRESPONDENT": ["代理银行", "CORRESPONDENT"],
    "BENEFICIAL_OWNER": ["受益所有", "BENEFICIALOWNERSHIP", "BENEFICIALOWNER"],
    "NOMINEE": ["代名", "NOMINEE"],
    "OFFSHORE": ["离岸", "OFFSHORE"],
    "LIFE_INSURANCE": ["人寿保险", "LIFEINSURANCE"],
    "AUDIT": ["审计", "AUDIT"],
    "SUBPOENA": ["传票", "SUBPOENA"],
    "BRIBERY": ["贿赂", "BRIBERY"],
    "CORRUPTION": ["腐败", "CORRUPTION"],
    "PPP": ["公私合作", "PPP", "PUBLIC-PRIVATE"],
    "MODEL": ["模型", "MODEL"],
    "PERIODIC_REFRESH": ["定期客户尽职调查", "定期", "周期性", "PERIODIC", "REFRESH"],
    "DATA_SHARING": ["共享数据", "数据共享", "DATASHARING"],
    "WIRE_TRANSFER": ["电汇", "WIRETRANSFER"],
    "ORIGINATOR_BENEFICIARY": ["发起人", "受益人信息", "ORIGINATOR", "BENEFICIARY"],
}

VISUAL_OVERRIDES: dict[int, int] = {
    4: 4,
    69: 69,
    82: 288,
    83: 82,
    89: 88,
    91: 90,
    100: 99,
    101: 100,
    104: 103,
    125: 124,
    132: 130,
    140: 139,
    158: 157,
    164: 163,
    177: 176,
    194: 193,
    199: 198,
    200: 199,
    201: 200,
    261: 259,
    300: 299,
    301: 300,
    302: 301,
}

VISUAL_REVIEW_NOTES: list[tuple[int, str, str]] = [
    (4, "同题但答案字母冲突", "USA PATRIOT Act correspondent-account seizure question; keep answer conflict for review."),
    (82, "同题但答案冲突", "Chinese screenshot shows B; English screenshot shows BC."),
    (91, "同题但答案字母冲突", "Wolfsberg monitoring question; options/answer letters differ between versions."),
    (104, "同题但答案字母冲突", "API question; Chinese answer B, English OCR answer A while selected option text appears to be B."),
    (122, "未覆盖-需人工确认", "Chinese fuzzy-logic question differs from closest English fuzzy-matching/name-screening question."),
    (125, "同题但截图过渡污染", "English screenshot contains previous explanation on the left and target question on the right."),
]


def read_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({header: row[idx] for idx, header in enumerate(headers)})
    return rows


def clean_display(text: Any) -> str:
    value = re.sub(r"\s+", " ", str(text or "")).strip()
    value = value.replace(" ,", ",").replace(" .", ".")
    return value


def cjk_text(text: Any) -> str:
    value = str(text or "")
    for token in NOISE_TOKENS:
        value = value.replace(token, " ")
    return "".join(re.findall(r"[\u4e00-\u9fff]+", value))


def ngrams(text: str) -> Counter[str]:
    if not text:
        return Counter()
    pieces: list[str] = []
    for size in (2, 3, 4):
        pieces.extend(text[idx : idx + size] for idx in range(max(0, len(text) - size + 1)))
    return Counter(pieces or [text])


def cosine(left: Counter[str], right: Counter[str]) -> float:
    if not left or not right:
        return 0.0
    numerator = sum(min(left[key], right[key]) for key in left.keys() & right.keys())
    left_norm = math.sqrt(sum(value * value for value in left.values()))
    right_norm = math.sqrt(sum(value * value for value in right.values()))
    if not left_norm or not right_norm:
        return 0.0
    return numerator / (left_norm * right_norm)


def concept_set(text: Any) -> set[str]:
    compact = re.sub(r"\s+", "", str(text or "").upper())
    found: set[str] = set()
    for concept, values in CONCEPTS.items():
        for value in values:
            if value.upper().replace(" ", "") in compact:
                found.add(concept)
                break
    return found


def jaccard(left: set[str], right: set[str]) -> float:
    if not left and not right:
        return 0.0
    return len(left & right) / len(left | right)


def page_int(row: dict[str, Any], key: str) -> int:
    value = row.get(key)
    try:
        return int(value)
    except (TypeError, ValueError):
        return 9999


def row_text(row: dict[str, Any], question_key: str, explanation_key: str) -> str:
    return f"{row.get(question_key) or ''} {row.get(explanation_key) or ''}"


def near_bonus(diff: int) -> float:
    if diff == 0:
        return 0.150
    if diff <= 2:
        return 0.090
    if diff <= 10:
        return 0.040
    return 0.0


def answer_adjustment(cn_answer: Any, en_answer: Any) -> float:
    cn_value = str(cn_answer or "").strip()
    en_value = str(en_answer or "").strip()
    if not cn_value or not en_value:
        return 0.0
    if cn_value == en_value:
        return 0.040
    cn_set = set(cn_value)
    en_set = set(en_value)
    if cn_set & en_set and (cn_set <= en_set or en_set <= cn_set):
        return 0.010
    return -0.050


def answer_is_disjoint(cn_answer: Any, en_answer: Any) -> bool:
    cn_value = str(cn_answer or "").strip()
    en_value = str(en_answer or "").strip()
    return bool(cn_value and en_value and not (set(cn_value) & set(en_value)))


def fallback_score(
    sim: float,
    concept_overlap: float,
    diff: int,
    answer_adjust: float,
) -> float:
    concept_weight = 0.03 if diff > 20 and sim < 0.02 else 0.12
    return sim + concept_weight * concept_overlap + near_bonus(diff) + answer_adjust


def confidence_label(method: str, sim: float, score: float, margin: float, diff: int) -> str:
    if method == "内容相似" and sim >= 0.12 and margin >= 0.025:
        return "高"
    if method == "内容相似" and sim >= 0.075 and margin >= 0.012:
        return "中"
    if diff <= 2 and score >= 0.16:
        return "中"
    return "低"


def answer_status(cn_answer: Any, en_answer: Any) -> tuple[str, str]:
    cn_value = str(cn_answer or "").strip()
    en_value = str(en_answer or "").strip()
    if cn_value and en_value:
        if cn_value == en_value:
            return cn_value, "一致"
        return "", f"不一致: 中文={cn_value}; 英文={en_value}"
    if cn_value or en_value:
        return cn_value or en_value, "单侧识别"
    return "", "未识别"


def build_alignment(cn_rows: list[dict[str, Any]], en_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cn_vecs = [ngrams(cjk_text(row_text(row, "题目候选文本", "解析候选文本"))) for row in cn_rows]
    en_vecs = [ngrams(cjk_text(row_text(row, "Question Text", "Explanation Text"))) for row in en_rows]
    cn_concepts = [concept_set(row_text(row, "题目候选文本", "解析候选文本")) for row in cn_rows]
    en_concepts = [concept_set(row_text(row, "Question Text", "Explanation Text")) for row in en_rows]

    ranked_by_cn: list[list[dict[str, Any]]] = []
    for cn_idx, cn_row in enumerate(cn_rows):
        cn_page = page_int(cn_row, "题号")
        ranked: list[dict[str, Any]] = []
        for en_idx, en_row in enumerate(en_rows):
            en_page = page_int(en_row, "Question")
            sim = cosine(cn_vecs[cn_idx], en_vecs[en_idx])
            overlap = jaccard(cn_concepts[cn_idx], en_concepts[en_idx])
            diff = abs(cn_page - en_page)
            cn_answer = str(cn_row.get("答案") or "").strip()
            en_answer = str(en_row.get("Answer") or "").strip()
            answer_adjust = answer_adjustment(cn_answer, en_answer)
            score = fallback_score(sim, overlap, diff, answer_adjust)
            ranked.append(
                {
                    "idx": en_idx,
                    "page": en_page,
                    "sim": sim,
                    "concept": overlap,
                    "diff": diff,
                    "score": score,
                    "answer_disjoint": answer_is_disjoint(cn_answer, en_answer),
                }
            )

        by_content = sorted(ranked, key=lambda item: (item["sim"], item["score"]), reverse=True)
        best_content = by_content[0]
        same_page = next((item for item in ranked if item["page"] == cn_page), None)
        same_page_sim = float(same_page["sim"]) if same_page else 0.0
        top_score = max(item["score"] for item in ranked)
        for candidate in ranked:
            semantic_boost = 0.0
            if (
                candidate["sim"] >= 0.075
                and not candidate["answer_disjoint"]
                and (candidate["sim"] - same_page_sim >= 0.040 or candidate["score"] >= top_score - 0.020)
            ):
                semantic_boost = 0.10 + candidate["sim"]
            candidate["rank_score"] = candidate["score"] + semantic_boost
        ranked_by_cn.append(sorted(ranked, key=lambda item: (item["rank_score"], item["sim"], item["score"]), reverse=True))

    en_page_to_idx = {page_int(row, "Question"): idx for idx, row in enumerate(en_rows)}
    chosen_by_cn: dict[int, dict[str, Any]] = {}
    used_en: set[int] = set()
    for cn_idx, cn_row in enumerate(cn_rows):
        cn_page = page_int(cn_row, "题号")
        override_page = VISUAL_OVERRIDES.get(cn_page)
        if override_page is None:
            continue
        en_idx = en_page_to_idx.get(override_page)
        if en_idx is None:
            continue
        source_candidate = next((item for item in ranked_by_cn[cn_idx] if item["idx"] == en_idx), None)
        if source_candidate is None:
            continue
        candidate = dict(source_candidate)
        candidate["manual_override"] = True
        chosen_by_cn[cn_idx] = candidate
        used_en.add(en_idx)

    assignment_order = sorted(
        range(len(cn_rows)),
        key=lambda idx: (
            ranked_by_cn[idx][0]["rank_score"],
            ranked_by_cn[idx][0]["sim"],
            -ranked_by_cn[idx][0]["diff"],
        ),
        reverse=True,
    )
    for cn_idx in assignment_order:
        if cn_idx in chosen_by_cn:
            continue
        for candidate in ranked_by_cn[cn_idx]:
            if candidate["idx"] not in used_en:
                chosen_by_cn[cn_idx] = candidate
                used_en.add(candidate["idx"])
                break

    matches: list[dict[str, Any]] = []
    for cn_idx, cn_row in enumerate(cn_rows):
        chosen = chosen_by_cn[cn_idx]
        ranked = ranked_by_cn[cn_idx]
        next_best = next((item for item in ranked if item["idx"] != chosen["idx"]), None)
        if chosen.get("manual_override"):
            method = "视觉复核覆盖"
        else:
            method = "内容相似" if chosen["sim"] >= 0.075 and not chosen["answer_disjoint"] else "同号/近号+概念"
        margin = chosen["rank_score"] - (next_best["rank_score"] if next_best else 0.0)

        en_row = en_rows[chosen["idx"]]
        suggested_answer, status = answer_status(cn_row.get("答案"), en_row.get("Answer"))
        confidence = "高" if chosen.get("manual_override") else confidence_label(
            method,
            chosen["sim"],
            chosen["score"],
            margin,
            int(chosen["diff"]),
        )
        alternatives = [
            f'{item["page"]}({item["rank_score"]:.3f}/{item["sim"]:.3f})'
            for item in ranked[:5]
            if item["idx"] != chosen["idx"]
        ]
        matches.append(
            {
                "confidence": confidence,
                "method": method,
                "score": chosen["score"],
                "sim": chosen["sim"],
                "concept": chosen["concept"],
                "diff": chosen["diff"],
                "alternatives": "; ".join(alternatives[:4]),
                "suggested_answer": suggested_answer,
                "answer_status": status,
                "cn": cn_row,
                "en": en_row,
            }
        )
    return matches


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def write_output(cn_rows: list[dict[str, Any]], en_rows: list[dict[str, Any]], matches: list[dict[str, Any]], output: Path) -> None:
    wb = Workbook()
    ws_cn = wb.active
    ws_cn.title = "中文题库"
    ws_en = wb.create_sheet("英文题库")
    ws_align = wb.create_sheet("语义对齐")
    ws_unmatched = wb.create_sheet("英文未匹配")
    ws_visual = wb.create_sheet("视觉复核记录")

    ws_cn.append(["题号", "分卷", "答案", "时间", "帧数", "题目文本", "解析文本", "来源视频", "题目截图", "解析截图"])
    for row in cn_rows:
        ws_cn.append(
            [
                row.get("题号"),
                row.get("分卷"),
                row.get("答案"),
                row.get("起止秒"),
                row.get("帧数"),
                clean_display(row.get("题目候选文本")),
                clean_display(row.get("解析候选文本")),
                row.get("来源视频"),
                row.get("题目截图"),
                row.get("解析截图"),
            ]
        )

    ws_en.append(["题号", "分卷", "答案", "时间", "帧数", "题目文本", "解析文本", "来源视频", "题目截图", "解析截图"])
    for row in en_rows:
        ws_en.append(
            [
                row.get("Question"),
                row.get("Part"),
                row.get("Answer"),
                row.get("Time"),
                row.get("Frames"),
                clean_display(row.get("Question Text")),
                clean_display(row.get("Explanation Text")),
                row.get("Source"),
                row.get("Question Image"),
                row.get("Explanation Image"),
            ]
        )

    ws_align.append(
        [
            "匹配置信",
            "匹配方法",
            "匹配分数",
            "中文相似度",
            "概念重合",
            "题号差",
            "推荐答案",
            "答案状态",
            "中文题号",
            "英文题号",
            "中文答案",
            "英文答案",
            "中文题目",
            "英文题目",
            "中文解析",
            "英文解析",
            "中文来源",
            "英文来源",
            "中文时间",
            "英文时间",
            "中文题目截图",
            "英文题目截图",
            "备选英文题号(分数/中文相似度)",
        ]
    )
    used_en_pages: Counter[int] = Counter()
    for item in matches:
        cn = item["cn"]
        en = item["en"]
        used_en_pages[page_int(en, "Question")] += 1
        ws_align.append(
            [
                item["confidence"],
                item["method"],
                round(item["score"], 4),
                round(item["sim"], 4),
                round(item["concept"], 4),
                item["diff"],
                item["suggested_answer"],
                item["answer_status"],
                cn.get("题号"),
                en.get("Question"),
                cn.get("答案"),
                en.get("Answer"),
                clean_display(cn.get("题目候选文本")),
                clean_display(en.get("Question Text")),
                clean_display(cn.get("解析候选文本")),
                clean_display(en.get("Explanation Text")),
                cn.get("来源视频"),
                en.get("Source"),
                cn.get("起止秒"),
                en.get("Time"),
                cn.get("题目截图"),
                en.get("Question Image"),
                item["alternatives"],
            ]
        )

    ws_unmatched.append(["英文题号", "答案", "题目文本", "解析文本", "来源视频", "时间", "题目截图", "使用次数"])
    for row in en_rows:
        page = page_int(row, "Question")
        if used_en_pages[page] == 0 or used_en_pages[page] > 1:
            ws_unmatched.append(
                [
                    row.get("Question"),
                    row.get("Answer"),
                    clean_display(row.get("Question Text")),
                    clean_display(row.get("Explanation Text")),
                    row.get("Source"),
                    row.get("Time"),
                    row.get("Question Image"),
                    used_en_pages[page],
                ]
            )

    ws_visual.append(["中文题号", "英文题号", "复核结论", "备注"])
    note_by_page = {page: (status, note) for page, status, note in VISUAL_REVIEW_NOTES}
    for cn_page, en_page in VISUAL_OVERRIDES.items():
        status, note = note_by_page.get(cn_page, ("视觉确认覆盖", ""))
        ws_visual.append([cn_page, en_page, status, note])
    for cn_page, status, note in VISUAL_REVIEW_NOTES:
        if cn_page not in VISUAL_OVERRIDES:
            ws_visual.append([cn_page, "", status, note])

    for ws in (ws_cn, ws_en, ws_align, ws_unmatched, ws_visual):
        style_sheet(ws)
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 72

    set_widths(ws_cn, {"A": 8, "B": 8, "C": 8, "D": 14, "E": 8, "F": 70, "G": 70, "H": 18, "I": 55, "J": 55})
    set_widths(ws_en, {"A": 8, "B": 8, "C": 8, "D": 14, "E": 8, "F": 70, "G": 70, "H": 18, "I": 55, "J": 55})
    set_widths(
        ws_align,
        {
            "A": 10,
            "B": 14,
            "C": 10,
            "D": 10,
            "E": 10,
            "F": 8,
            "G": 10,
            "H": 24,
            "I": 8,
            "J": 8,
            "K": 8,
            "L": 8,
            "M": 58,
            "N": 58,
            "O": 58,
            "P": 58,
            "Q": 18,
            "R": 18,
            "S": 14,
            "T": 14,
            "U": 55,
            "V": 55,
            "W": 30,
        },
    )
    set_widths(ws_unmatched, {"A": 8, "B": 8, "C": 70, "D": 70, "E": 18, "F": 14, "G": 55, "H": 10})
    set_widths(ws_visual, {"A": 10, "B": 10, "C": 22, "D": 90})

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    cn_rows = read_sheet(args.workbook, CN_SHEET)
    en_rows = read_sheet(args.workbook, EN_SHEET)
    matches = build_alignment(cn_rows, en_rows)
    write_output(cn_rows, en_rows, matches, args.output)

    counts = Counter(item["confidence"] for item in matches)
    answer_counts = Counter(item["answer_status"].split(":")[0] for item in matches)
    print(f"rows={len(matches)}")
    print("confidence=" + ", ".join(f"{key}:{counts[key]}" for key in ("高", "中", "低")))
    print("answers=" + ", ".join(f"{key}:{value}" for key, value in answer_counts.items()))
    print(f"output={args.output}")


if __name__ == "__main__":
    main()
