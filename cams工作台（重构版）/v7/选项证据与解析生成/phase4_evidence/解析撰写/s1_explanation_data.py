# -*- coding: utf-8 -*-
"""s1 — 解析撰写专用：数据加载层。路径常量、题库加载、参考工作簿、KG 页码映射。"""

from __future__ import annotations

import json, os, re
from pathlib import Path
from typing import Any


HERE = Path(__file__).resolve().parent  # phase4_evidence/解析撰写/
PHASE4 = HERE.parent
V7_EVIDENCE_ROOT = PHASE4.parent  # 选项证据与解析生成/
CAMS_ROOT = PHASE4.parents[3]  # cams工作台（重构版）/
V7_ROOT = V7_EVIDENCE_ROOT.parent  # v7/
DEFAULT_OUTPUT_DIR = PHASE4 / "output"
KG_GRAPH_PATH = (
    V7_ROOT / "知识图谱提取" / "phases" / "phase06_kg_views" / "outputs" / "kg_retrieval_graph.json"
)
DEFAULT_QUESTIONS_PATH = (
    V7_EVIDENCE_ROOT / "phase3.5_questions" / "output" / "v7_questions.json"
)
DEFAULT_REFERENCE_WORKBOOK = (
    CAMS_ROOT
    / "教材、答疑记录、习题与参考文献"
    / "习题"
    / "v7习题"
    / "v7习题提取脚本"
    / "output_2s"
    / "CAMS_v7题库_中英对照_v8精修版.xlsx"
)

API_KEY_ENV_NAMES = ("DEEPSEEK_API_KEY", "DS_API_KEY", "DS_KEY")
DEFAULT_DEEPSEEK_BASE_URL = "https://api.deepseek.com/v1"
SCHEMA_VERSION = "s6_teacher_explanation_v3_1"
PROMPT_VERSION = "s6_teacher_prompt_v3_1"
INSUFFICIENT_TEXT = "现有材料不足以判断该项。"
INTERNAL_REVIEW_NEEDED = "需教研复核"
SOURCE_QUOTE_MIN_LENGTH = 40
SOURCE_QUOTE_MAX_LENGTH = 240
OPTION_SUPPLEMENT_CONTEXT_LIMIT = 2
TEXTBOOK_BASIS_TYPES = {"textbook_direct", "textbook_definition_application"}
STEM_BASIS_TYPES = {"stem_contrast"}

# 盲判 decision_basis → 解析 basis_type 的直接映射
_DECISION_TO_BASIS: dict[str, str] = {
    "direct_taxonomy": "textbook_direct",
    "definition_application": "textbook_definition_application",
    "domain_contrast": "textbook_direct",
    "stem_contrast": "stem_contrast",
    "insufficient": "insufficient",
}
INTERNAL_UNIT_ID_RE = re.compile(
    r"\s*[（(]?\s*v7u_[A-Za-z0-9_-]+\s*[）)]?", re.IGNORECASE
)


def get_llm_config() -> tuple[str, str, str]:
    for env_name in API_KEY_ENV_NAMES:
        value = os.environ.get(env_name)
        if value:
            base_url = (
                os.environ.get("DEEPSEEK_BASE_URL")
                or os.environ.get("DS_BASE_URL")
                or DEFAULT_DEEPSEEK_BASE_URL
            )
            return value, base_url, env_name
    names = " / ".join(API_KEY_ENV_NAMES)
    raise RuntimeError(f"{names} 环境变量均未设置，不能调用 LLM API。")


_UNIT_PAGE_MAP: dict[str, dict[str, Any]] | None = None


def _get_unit_page_map() -> dict[str, dict[str, Any]]:
    """懒加载 unit_id → 页码映射（PDF 页码 + 书内页码）。"""
    global _UNIT_PAGE_MAP
    if _UNIT_PAGE_MAP is not None:
        return _UNIT_PAGE_MAP
    if not KG_GRAPH_PATH.exists():
        _UNIT_PAGE_MAP = {}
        return _UNIT_PAGE_MAP
    with open(KG_GRAPH_PATH, "r", encoding="utf-8") as f:
        raw = json.load(f)
    page_map: dict[str, dict[str, Any]] = {}
    for unit in raw.get("units", []) or []:
        uid = str(unit.get("unit_id", "") or "").strip()
        if not uid:
            continue
        entry: dict[str, Any] = {}
        pdf_page = unit.get("pdf_page")
        printed_page = unit.get("printed_page")
        if pdf_page is not None:
            entry["pdf_page"] = pdf_page
        if printed_page:
            entry["printed_page"] = str(printed_page)
        if entry:
            page_map[uid] = entry
    _UNIT_PAGE_MAP = page_map
    return _UNIT_PAGE_MAP


def load_question_result(path: Path) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_standard_questions(path: str | Path) -> dict[str, dict[str, Any]]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return {row["question_id"]: row for row in data.get("items", [])}


def parse_answer_cell(value: Any) -> list[str]:
    text = str(value or "").upper()
    return sorted(set(re.findall(r"[A-F]", text)))


def load_reference_workbook(path: str | Path) -> dict[str, dict[str, Any]]:
    path = Path(path)
    if not path.exists():
        raise RuntimeError(f"参考题库工作簿不存在: {path}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，无法读取参考题库工作簿") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "中英对照表" not in workbook.sheetnames:
        raise RuntimeError('参考题库工作簿缺少"中英对照表"工作表')
    sheet = workbook["中英对照表"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    required = {"中文题号", "中文答案", "中文解析", "英文答案", "英文解析"}
    missing = required - set(headers)
    if missing:
        raise RuntimeError(f"参考题库工作簿缺少列: {', '.join(sorted(missing))}")

    lookup: dict[str, dict[str, Any]] = {}
    for values in rows:
        row = dict(zip(headers, values))
        try:
            number = int(row.get("中文题号"))
        except (TypeError, ValueError):
            continue
        qid = f"v7_q_{number:06d}"
        if qid in lookup:
            raise RuntimeError(f"参考题库题号重复: {qid}")
        lookup[qid] = {
            "cn_answer": parse_answer_cell(row.get("中文答案")),
            "cn_explanation": str(row.get("中文解析") or "").strip(),
            "en_answer": parse_answer_cell(row.get("英文答案")),
            "en_explanation": str(row.get("英文解析") or "").strip(),
        }
    workbook.close()
    return lookup


def build_reference_context(
    qid: str,
    predicted_answer: list[str],
    standard_question: dict[str, Any],
    workbook_row: dict[str, Any],
) -> dict[str, Any]:
    final_answer = [str(x).strip().upper() for x in standard_question.get("answer", [])]
    cn_answer = workbook_row.get("cn_answer", []) or []
    en_answer = workbook_row.get("en_answer", []) or []
    predicted = [str(x).strip().upper() for x in predicted_answer]
    cn_en_conflict = bool(cn_answer and en_answer and set(cn_answer) != set(en_answer))
    blind_final_conflict = bool(
        predicted and final_answer and set(predicted) != set(final_answer)
    )
    messages: list[str] = []
    if cn_en_conflict:
        messages.append(
            f"中英文参考答案冲突：中文={','.join(cn_answer)}，英文={','.join(en_answer)}"
        )
    if blind_final_conflict:
        messages.append(
            f"盲判与题库最终参考答案冲突：盲判={','.join(predicted)}，"
            f"题库最终={','.join(final_answer)}"
        )
    if not messages:
        messages.append("未发现答案冲突。")
    return {
        "question_id": qid,
        "final_answer": final_answer,
        "cn_answer": cn_answer,
        "cn_explanation": workbook_row.get("cn_explanation", ""),
        "en_answer": en_answer,
        "en_explanation": workbook_row.get("en_explanation", ""),
        "cn_en_conflict": cn_en_conflict,
        "blind_final_conflict": blind_final_conflict,
        "conflict_messages": messages,
        "tier": standard_question.get("tier", ""),
        "risk_flags": standard_question.get("risk_flags", []),
    }