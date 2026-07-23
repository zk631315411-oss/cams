# -*- coding: utf-8 -*-
"""从正式盲判结果生成有据可查的 V3.1 教研解析母版。

V3 正文仅使用盲判证据。参考答案和原始参考解析单独加载并确定性地追加；
它们绝不进入生成 prompt，也绝不替代 ``predicted_answer``。
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any


HERE = Path(__file__).resolve().parent  # phase4_evidence/解析撰写/
PHASE4 = HERE.parent
V7_EVIDENCE_ROOT = PHASE4.parent  # 选项证据与解析生成/
CAMS_ROOT = PHASE4.parents[3]  # cams工作台（重构版）/
V7_ROOT = V7_EVIDENCE_ROOT.parent  # v7/
DEFAULT_OUTPUT_DIR = PHASE4 / "output"
KG_GRAPH_PATH = (
    V7_ROOT / "知识图谱提取" / "phases" / "phase06_kg_views" / "outputs" / "kg_retrieval_graph.json"
)
DEFAULT_QUESTIONS_PATH = (
    V7_EVIDENCE_ROOT / "phase3.5_questions" / "output" / "v7_questions.json"
)
DEFAULT_REFERENCE_WORKBOOK = (
    CAMS_ROOT
    / "教材、答疑记录、习题与参考文献"
    / "习题"
    / "v7习题"
    / "v7习题提取脚本"
    / "output_2s"
    / "CAMS_v7题库_中英对照_v8精修版.xlsx"
)

API_KEY_ENV_NAMES = ("DEEPSEEK_API_KEY", "DS_API_KEY", "DS_KEY")
DEFAULT_DEEPSEEK_BASE_URL = "https://api.deepseek.com/v1"
SCHEMA_VERSION = "s6_teacher_explanation_v3_1"
PROMPT_VERSION = "s6_teacher_prompt_v3_1"
INSUFFICIENT_TEXT = "现有材料不足以判断该项。"
INTERNAL_REVIEW_NEEDED = "需教研复核"
SOURCE_QUOTE_MIN_LENGTH = 40
SOURCE_QUOTE_MAX_LENGTH = 240
OPTION_SUPPLEMENT_CONTEXT_LIMIT = 2
TEXTBOOK_BASIS_TYPES = {
    "textbook_direct",
    "textbook_definition_application",
}
STEM_BASIS_TYPES = {"stem_contrast"}

# 盲判 decision_basis → 解析 basis_type 的直接映射
_DECISION_TO_BASIS: dict[str, str] = {
    "direct_taxonomy": "textbook_direct",
    "definition_application": "textbook_definition_application",
    "domain_contrast": "textbook_direct",
    "stem_contrast": "stem_contrast",
    "insufficient": "insufficient",
}
INTERNAL_UNIT_ID_RE = re.compile(
    r"\s*[（(]?\s*v7u_[A-Za-z0-9_-]+\s*[）)]?", re.IGNORECASE
)


def get_llm_config() -> tuple[str, str, str]:
    for env_name in API_KEY_ENV_NAMES:
        value = os.environ.get(env_name)
        if value:
            base_url = (
                os.environ.get("DEEPSEEK_BASE_URL")
                or os.environ.get("DS_BASE_URL")
                or DEFAULT_DEEPSEEK_BASE_URL
            )
            return value, base_url, env_name
    names = " / ".join(API_KEY_ENV_NAMES)
    raise RuntimeError(f"{names} 环境变量均未设置，不能调用 LLM API。")


def strip_json_fence(text: str) -> str:
    text = (text or "").strip()
    text = re.sub(r"^```(?:json)?\s*\n?", "", text)
    text = re.sub(r"\n?```\s*$", "", text)
    return text.strip()


def parse_json_object(raw_text: str) -> dict[str, Any] | None:
    cleaned = strip_json_fence(raw_text)
    if not cleaned:
        return None
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass
    match = re.search(r"\{[\s\S]*\}", cleaned)
    if match:
        try:
            return json.loads(match.group(0))
        except json.JSONDecodeError:
            pass
    try:
        import json_repair

        return json.loads(json_repair.repair_json(cleaned))
    except Exception:
        return None


def call_llm(
    client: Any,
    prompt: str,
    model: str = "deepseek-v4-pro",
    max_tokens: int = 8000,
    timeout: float = 150.0,
    reasoning_effort: str = "high",
    enable_thinking: bool = True,
) -> str:
    kwargs: dict[str, Any] = dict(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=max_tokens,
        timeout=timeout,
    )
    if enable_thinking:
        kwargs["extra_body"] = {"thinking": {"type": "enabled"}}
        kwargs["reasoning_effort"] = reasoning_effort
    else:
        kwargs["temperature"] = 0.4
    response = client.chat.completions.create(**kwargs)
    return (response.choices[0].message.content or "").strip()


_UNIT_PAGE_MAP: dict[str, dict[str, Any]] | None = None


def _get_unit_page_map() -> dict[str, dict[str, Any]]:
    """懒加载 unit_id → 页码映射（PDF 页码 + 书内页码）。"""
    global _UNIT_PAGE_MAP
    if _UNIT_PAGE_MAP is not None:
        return _UNIT_PAGE_MAP
    if not KG_GRAPH_PATH.exists():
        _UNIT_PAGE_MAP = {}
        return _UNIT_PAGE_MAP
    with open(KG_GRAPH_PATH, "r", encoding="utf-8") as f:
        raw = json.load(f)
    page_map: dict[str, dict[str, Any]] = {}
    for unit in raw.get("units", []) or []:
        uid = str(unit.get("unit_id", "") or "").strip()
        if not uid:
            continue
        entry: dict[str, Any] = {}
        pdf_page = unit.get("pdf_page")
        printed_page = unit.get("printed_page")
        if pdf_page is not None:
            entry["pdf_page"] = pdf_page
        if printed_page:
            entry["printed_page"] = str(printed_page)
        if entry:
            page_map[uid] = entry
    _UNIT_PAGE_MAP = page_map
    return _UNIT_PAGE_MAP


def load_question_result(path: Path) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_standard_questions(path: str | Path) -> dict[str, dict[str, Any]]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return {row["question_id"]: row for row in data.get("items", [])}


def parse_answer_cell(value: Any) -> list[str]:
    text = str(value or "").upper()
    return sorted(set(re.findall(r"[A-F]", text)))


def load_reference_workbook(path: str | Path) -> dict[str, dict[str, Any]]:
    path = Path(path)
    if not path.exists():
        raise RuntimeError(f"参考题库工作簿不存在: {path}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，无法读取参考题库工作簿") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "中英对照表" not in workbook.sheetnames:
        raise RuntimeError("参考题库工作簿缺少“中英对照表”工作表")
    sheet = workbook["中英对照表"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    required = {
        "中文题号",
        "中文答案",
        "中文解析",
        "英文答案",
        "英文解析",
    }
    missing = required - set(headers)
    if missing:
        raise RuntimeError(f"参考题库工作簿缺少列: {', '.join(sorted(missing))}")

    lookup: dict[str, dict[str, Any]] = {}
    for values in rows:
        row = dict(zip(headers, values))
        try:
            number = int(row.get("中文题号"))
        except (TypeError, ValueError):
            continue
        qid = f"v7_q_{number:06d}"
        if qid in lookup:
            raise RuntimeError(f"参考题库题号重复: {qid}")
        lookup[qid] = {
            "cn_answer": parse_answer_cell(row.get("中文答案")),
            "cn_explanation": str(row.get("中文解析") or "").strip(),
            "en_answer": parse_answer_cell(row.get("英文答案")),
            "en_explanation": str(row.get("英文解析") or "").strip(),
        }
    workbook.close()
    return lookup


def build_reference_context(
    qid: str,
    predicted_answer: list[str],
    standard_question: dict[str, Any],
    workbook_row: dict[str, Any],
) -> dict[str, Any]:
    final_answer = [str(x).strip().upper() for x in standard_question.get("answer", [])]
    cn_answer = workbook_row.get("cn_answer", []) or []
    en_answer = workbook_row.get("en_answer", []) or []
    predicted = [str(x).strip().upper() for x in predicted_answer]
    cn_en_conflict = bool(cn_answer and en_answer and set(cn_answer) != set(en_answer))
    blind_final_conflict = bool(
        predicted and final_answer and set(predicted) != set(final_answer)
    )
    messages: list[str] = []
    if cn_en_conflict:
        messages.append(
            f"中英文参考答案冲突：中文={','.join(cn_answer)}，英文={','.join(en_answer)}"
        )
    if blind_final_conflict:
        messages.append(
            f"盲判与题库最终参考答案冲突：盲判={','.join(predicted)}，"
            f"题库最终={','.join(final_answer)}"
        )
    if not messages:
        messages.append("未发现答案冲突。")
    return {
        "question_id": qid,
        "final_answer": final_answer,
        "cn_answer": cn_answer,
        "cn_explanation": workbook_row.get("cn_explanation", ""),
        "en_answer": en_answer,
        "en_explanation": workbook_row.get("en_explanation", ""),
        "cn_en_conflict": cn_en_conflict,
        "blind_final_conflict": blind_final_conflict,
        "conflict_messages": messages,
        "tier": standard_question.get("tier", ""),
        "risk_flags": standard_question.get("risk_flags", []),
    }


def candidate_by_unit(result: dict[str, Any]) -> dict[str, dict[str, Any]]:
    by_unit = {
        str(candidate["unit_id"]): candidate
        for candidate in result.get("candidate_pool", []) or []
        if isinstance(candidate, dict) and candidate.get("unit_id")
    }
    for rows in (result.get("option_supplement_pool", {}) or {}).values():
        for candidate in rows or []:
            if isinstance(candidate, dict) and candidate.get("unit_id"):
                by_unit.setdefault(str(candidate["unit_id"]), candidate)
    return by_unit


def compact_text(value: Any, max_len: int = 520) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if len(text) <= max_len:
        return text
    return text[: max_len - 1].rstrip() + "…"


def _material_card(unit: dict[str, Any], uid: str, source_kind: str) -> dict[str, Any]:
    return {
        "unit_id": uid,
        "source_kind": source_kind,
        "knowledge_zh": unit.get("knowledge_zh", ""),
        "en_quote": unit.get("en_quote") or unit.get("knowledge_en", ""),
        "heading_context": unit.get("heading_context", []),
        "best_rank": unit.get("best_rank"),
        "routes": unit.get("routes", []),
        "languages": unit.get("languages", []),
        "content_type": unit.get("type", ""),
    }


def enriched_option_material(result: dict[str, Any]) -> list[dict[str, Any]]:
    options = result.get("options", {}) or {}
    unit_map = candidate_by_unit(result)
    supplements = result.get("option_supplement_pool", {}) or {}
    by_label = {
        str(row.get("option", "")).strip().upper(): row
        for row in result.get("option_analysis", []) or []
        if isinstance(row, dict)
    }
    rows: list[dict[str, Any]] = []
    for label, option_text in options.items():
        label = str(label).strip().upper()
        analysis = by_label.get(label, {})
        cards: list[dict[str, Any]] = []
        bound_ids: set[str] = set()
        for card in analysis.get("evidence_cards", []) or []:
            if not isinstance(card, dict):
                continue
            uid = str(card.get("unit_id", "")).strip()
            unit = unit_map.get(uid)
            if not unit or uid in bound_ids:
                continue
            bound_ids.add(uid)
            material = _material_card(unit, uid, "adjudicated")
            material["support_type"] = card.get("support_type", "")
            cards.append(material)

        supplement_cards: list[dict[str, Any]] = []
        for unit in supplements.get(label, []) or []:
            uid = str(unit.get("unit_id", "")).strip()
            if not uid or uid in bound_ids:
                continue
            supplement_cards.append(_material_card(unit, uid, "supplement_candidate"))
            if len(supplement_cards) >= OPTION_SUPPLEMENT_CONTEXT_LIMIT:
                break

        rows.append(
            {
                "option": label,
                "option_text": option_text,
                "judgement": analysis.get("judgement", ""),
                "evidence_status": analysis.get("evidence_status", ""),
                "decision_basis": analysis.get("decision_basis", ""),
                "evidence_cards": cards,
                "supplement_cards": supplement_cards,
            }
        )
    return rows


_TYPE_CN_MAP: dict[str, str] = {
    "definition": "概念定义",
    "rule": "规则/规定",
    "case": "案例",
    "fact": "事实陈述",
    "process": "流程描述",
    "risk_indicator": "风险指标",
    "classification": "分类说明",
    "context": "背景信息",
}


def _type_cn_label(content_type: str) -> str:
    return _TYPE_CN_MAP.get(str(content_type or "").strip(), content_type or "")


def _format_prompt_card(card: dict[str, Any]) -> str:
    retrieval = ""
    if card.get("source_kind") == "supplement_candidate":
        retrieval = (
            f" | best_rank={card.get('best_rank', '')}"
            f" | routes={','.join(card.get('routes', []) or [])}"
            f" | languages={','.join(card.get('languages', []) or [])}"
        )
    type_label = _type_cn_label(card.get("content_type", ""))
    type_str = f" | 教材类型：{type_label}" if type_label else ""
    page_info = _get_unit_page_map().get(card["unit_id"], {})
    printed_page = page_info.get("printed_page", "")
    page_str = f" | P{printed_page}" if printed_page else ""
    return (
        f"- {card['unit_id']} | {card['source_kind']}"
        f" | {card.get('support_type', '')}{type_str}{page_str}{retrieval}\n"
        f"  中文要点：{compact_text(card['knowledge_zh'])}\n"
        f"  英文原文：{compact_text(card['en_quote'])}\n"
        f"  章节：{' > '.join(card['heading_context'])}"
    )


# ── KG 教材原文连续上下文 ──────────────────────────────────────────

_KG_UNIT_CACHE: dict[str, dict[str, Any]] | None = None


def _load_kg_units() -> dict[str, dict[str, Any]]:
    """加载 KG 所有 unit，按 unit_id 索引。缓存，只加载一次。"""
    global _KG_UNIT_CACHE
    if _KG_UNIT_CACHE is not None:
        return _KG_UNIT_CACHE
    if not KG_GRAPH_PATH.exists():
        _KG_UNIT_CACHE = {}
        return _KG_UNIT_CACHE
    with open(KG_GRAPH_PATH, "r", encoding="utf-8") as f:
        kg = json.load(f)
    _KG_UNIT_CACHE = {}
    for unit in kg.get("units", []) or []:
        uid = str(unit.get("unit_id", "")).strip()
        if uid:
            _KG_UNIT_CACHE[uid] = unit
    return _KG_UNIT_CACHE


def _section_context_cards(
    unit_id: str,
    candidate_ids: set[str],
    context_range: int = 4,
) -> list[dict[str, Any]]:
    """返回 unit_id 同 Section 内 unit_order ±context_range 的连续材料卡。

    已检索到的用正常格式，未检索到的补占位（从 KG 取中文摘要 + 英文原文）。
    """
    kg_units = _load_kg_units()
    center = kg_units.get(unit_id)
    if not center:
        return []
    
    section_id = center.get("section_id", "")
    center_order = int(center.get("unit_order") or 0)
    if not section_id or not center_order:
        return []
    
    # 找出同 section 所有 unit，按 unit_order 排序
    siblings: list[dict[str, Any]] = []
    for uid, unit in kg_units.items():
        if unit.get("section_id") == section_id:
            siblings.append(unit)
    siblings.sort(key=lambda u: int(u.get("unit_order") or 0))
    
    # 取 ±context_range 范围内
    result: list[dict[str, Any]] = []
    for unit in siblings:
        order = int(unit.get("unit_order") or 0)
        if abs(order - center_order) <= context_range:
            uid = str(unit.get("unit_id", ""))
            is_candidate = uid in candidate_ids
            card = {
                "unit_id": uid,
                "knowledge_zh": unit.get("knowledge_zh", ""),
                "en_quote": unit.get("en_quote") or "",
                "heading_context": unit.get("heading_context") or [],
                "type": unit.get("type", ""),
                "printed_page": unit.get("printed_page", ""),
                "real_section": unit.get("real_section") or unit.get("section_id", ""),
                "unit_order": order,
                "is_candidate": is_candidate,
                "is_center": uid == unit_id,
            }
            result.append(card)
    return result


def _format_context_block(cards: list[dict[str, Any]]) -> str:
    """将一组连续材料卡（含上下文）格式化为提示文本块。"""
    if not cards:
        return ""
    
    section_label = cards[0].get("real_section", "")
    heading = " > ".join(cards[0].get("heading_context", []) or [])
    
    lines = [f"【教材原文连续段落 — {section_label} ({heading})】", ""]
    for card in cards:
        uid = card["unit_id"]
        zh = compact_text(card["knowledge_zh"])
        en = compact_text(card["en_quote"])
        page_str = f" | P{card['printed_page']}" if card.get("printed_page") else ""
        type_label = _type_cn_label(card.get("type", ""))
        type_str = f" | 教材类型：{type_label}" if type_label else ""
        
        if card["is_center"]:
            marker = "★ 命中"
        elif card["is_candidate"]:
            marker = "  已检索"
        else:
            marker = "  补充上下文"
        
        lines.append(f"[{uid}] {marker}{type_str}{page_str}")
        lines.append(f"  中文要点：{zh}")
        if en:
            lines.append(f"  英文原文：{en}")
        lines.append("")
    lines.append("-" * 60)
    return "\n".join(lines)


def _build_context_augmented_material(
    result: dict[str, Any],
) -> list[dict[str, Any]]:
    """与 enriched_option_material 同接口，但每张 evidence card 附加 ±2 上下文。"""
    options = result.get("options", {}) or {}
    unit_map = candidate_by_unit(result)
    supplements = result.get("option_supplement_pool", {}) or {}
    by_label = {
        str(row.get("option", "")).strip().upper(): row
        for row in result.get("option_analysis", []) or []
        if isinstance(row, dict)
    }
    
    rows: list[dict[str, Any]] = []
    for label, option_text in options.items():
        label = str(label).strip().upper()
        analysis = by_label.get(label, {})
        cards: list[dict[str, Any]] = []
        bound_ids: set[str] = set()
        for evidence_card in analysis.get("evidence_cards", []) or []:
            if not isinstance(evidence_card, dict):
                continue
            uid = str(evidence_card.get("unit_id", "")).strip()
            unit = unit_map.get(uid)
            if not unit or uid in bound_ids:
                continue
            bound_ids.add(uid)
            material = _material_card(unit, uid, "adjudicated")
            material["support_type"] = evidence_card.get("support_type", "")
            material["context_block"] = _section_context_cards(uid, bound_ids)
            cards.append(material)

        supplement_cards: list[dict[str, Any]] = []
        for unit in supplements.get(label, []) or []:
            uid = str(unit.get("unit_id", "")).strip()
            if not uid or uid in bound_ids:
                continue
            bound_ids.add(uid)
            sc = _material_card(unit, uid, "supplement_candidate")
            sc["context_block"] = _section_context_cards(uid, bound_ids)
            supplement_cards.append(sc)
            if len(supplement_cards) >= OPTION_SUPPLEMENT_CONTEXT_LIMIT:
                break

        judgement = analysis.get("judgement", "")
        rows.append({
            "option": label,
            "option_text": option_text,
            "judgement": judgement,
            "evidence_status": analysis.get("evidence_status", ""),
            "decision_basis": analysis.get("decision_basis", ""),
            "evidence_cards": cards,
            "supplement_cards": supplement_cards,
        })
    return rows


def build_prompt(
    result: dict[str, Any], standard_question: dict[str, Any] | None = None
) -> str:
    standard_question = standard_question or {}
    predicted = "、".join(result.get("predicted_answer", []) or []) or "未形成答案"
    option_lines = "\n".join(
        f"{label}. {text}" for label, text in (result.get("options", {}) or {}).items()
    )
    options_en = standard_question.get("options_en", {}) or {}
    option_en_lines = "\n".join(
        f"{label}. {options_en.get(label, '')}"
        for label in (result.get("options", {}) or {})
        if options_en.get(label)
    )

    # 使用上下文增强版材料卡（同 Section ±2 unit 连续展示）
    # 去重 key = (real_section, center_unit_order)，避免同 Section 不同窗口被合并
    shown_context_keys: set[tuple[str, int]] = set()
    material_lines: list[str] = []
    for row in _build_context_augmented_material(result):
        material_lines.append(
            f"选项{row['option']}：{row['option_text']}\n"
            f"盲判标签：{row['judgement']} | 证据状态：{row['evidence_status']} | "
            f"原判断类型：{row['decision_basis']}"
        )
        material_lines.append("已裁判证据：")
        for card in row["evidence_cards"]:
            context_block = card.get("context_block", [])
            if context_block:
                section = context_block[0].get("real_section", "")
                center_order = next((c["unit_order"] for c in context_block if c["is_center"]), 0)
                ctx_key = (section, center_order)
                if ctx_key not in shown_context_keys:
                    material_lines.append(_format_context_block(context_block))
                    shown_context_keys.add(ctx_key)
            else:
                material_lines.append(_format_prompt_card(card))
        if not row["evidence_cards"]:
            material_lines.append("- 无")
        material_lines.append("解析补充候选：")
        for card in row["supplement_cards"]:
            context_block = card.get("context_block", [])
            if context_block:
                section = context_block[0].get("real_section", "")
                center_order = next((c["unit_order"] for c in context_block if c["is_center"]), 0)
                ctx_key = (section, center_order)
                if ctx_key not in shown_context_keys:
                    material_lines.append(_format_context_block(context_block))
                    shown_context_keys.add(ctx_key)
            else:
                material_lines.append(_format_prompt_card(card))
        if not row["supplement_cards"]:
            material_lines.append("- 无")

    framework = result.get("decision_framework", {}) or {}
    unit_map = candidate_by_unit(result)
    framework_material: list[str] = []
    for uid in framework.get("cited_unit_ids", []) or []:
        uid = str(uid)
        unit = unit_map.get(uid)
        if unit:
            framework_material.append(
                _format_prompt_card(_material_card(unit, uid, "shared_framework"))
            )

    chapter_text_value = "；".join(
        f"{row.get('real_chapter') or row.get('chapter_id', '')} {row.get('chapter_title', '')}".strip()
        for row in result.get("chapter_mappings", []) or []
    ) or "未映射"
    validation_text = "；".join(
        str(issue) for issue in result.get("validation_checks", []) or []
    ) or "无"

    return f"""你是一位CAMS反洗钱讲师。你的学生零基础、没看过教材、非金融法律专业背景、英文非母语，他们通过做题来学习，想象你坐在学生旁边，拿笔在纸上画给他看。

你的目标是：简明扼要且直击重点地**讲透题目内部的逻辑关系**，让学生看完解析后，换一道类似的题也能自己判断。

---

## 你的学生

- 零基础，没看过教材，边做题边学
- 非金融/法律专业出身
- 英文非母语，需要中英对照来理解选项原意

## 你的任务

对这道题（固定答案为"{predicted}"，不得改判），你需要让学生：

1. **知道考什么** —— exam_point：一句话，30字内。只写核心概念或能力点，不写判断结论，不引入题干没有的信息，不复述题干具体情节。示例：
   - 好："按金额粒度及分散方式区分 structuring 与 microstructuring"
   - 好："识别房地产场景下的 placement 阶段特征"
   - 坏："判断时需关注存款是否被故意拆分为多笔略低于报告限额的小额交易"（复述题干）
   - 坏："区分结构化与微结构化、贸易洗钱等手法"（题干没有提及贸易洗钱）

2. **知道正确答案为什么对** —— core_analysis：一个自然段落。先给概念定义或判断规则，再结合题干关键事实，推出正确答案为什么成立。如果教材原文用的术语和选项中出现的术语不完全一致，先准确引用教材原文的术语和定义，再说明两者的关系。不要偷换主语——不要把"教材定义了X"写成"教材定义了Y"。引用定义时直接给出定义内容和页码，不要加"教材明确指出""教材将……定义为"等前缀。示例：
	   - 坏："教材明确定义逃税为使用非法手段逃避纳税义务（P28）。"
	   - 好："逃税指使用非法手段逃避纳税义务（P28）。"
	   写完自问："如果换一道题干相似但选项不同的题，学生看完还能自己判断吗？"如果核心解析本身已经说明了判断方法，不画蛇添足。

3. **知道为什么正确项更优** —— option_explanations：只写真正有迷惑性的错误项。正确项不写（core_analysis已覆盖）。每个错误项的写法不是"排除它"，而是"在教材框架下，正确项比它更直接匹配"：
   - 错误项本身可能也有一定关联（如信托确实能隐藏所有权），但它在教材中的定义位置和题干条件的匹配度不如正确项。你要解释的是**为什么匹配度不如**，而不是**绝对不可能**。
   - 避免非黑即白的排除语气（"不属于""不可能""因此错误"）。改用比较级（"不如X直接""题干更支持X而非Y""更吻合教材对X的定义"）。
   - 明显无关的选项（题干压根没涉及该选项所需的要素），直接指出缺失即可，不展开。
   - 明显无关的选项不凑数
   - 不加"故该项不选""因此该项正确"等套话结尾

4. **下次不踩同样的坑** —— easy_mistake：如果有真正容易混淆的概念对，给出教材中的核心区分标准。如果没有独立于一、二、三之外的增量信息，留空（"text": ""）。

## 你的材料

以下是你可以使用的全部信息：

**题目**
中文题干：{result.get('stem', '')}
英文题干：{standard_question.get('stem_en', '')}
中文选项：
{option_lines}
英文选项：
{option_en_lines or '未提供'}

注意：本题来自英文考试，中文选项为翻译版本。当中文翻译与英文原意有偏差时（如 bending the rules 被译为"违规操作"），以英文原意为准。

**盲判框架**
固定答案：{predicted}
框架类型：{framework.get('type', '')}
{chr(10).join(framework_material) if framework_material else '无'}
教材章节：{chapter_text_value}

**选项材料（已标注教材类型和教材页码）**
{chr(10).join(material_lines)}

引用教材内容时，必须标注页码。例如："放置阶段指非法资金进入金融系统（P53）。"

## 写作铁律

1. 不得在题干原文上添加任何程度副词、数量词、性质词。题干写"低于"就是低于，不是"略低于"也不是"远低于"。题干写"一个账户"就是一个，不是"跨账户"。题干写"支付发票"就是支付发票，不是"虚假发票"。同样，教材原文用"such as""include""for example"等举例措辞的，不得转写成"清单""界定""明确列举"等暗示穷举或硬性边界的词。原文是举例，解析就写举例。
2. 引用教材案例时，用描述性语言（"教材在Tamayo案例中展示了..."），不用规定性语言（"微结构化通常""教材规定必须"）。案例是教材展示概念的方式，不是教材制定的规则。案例中的具体数字不能写成硬性标准。
3. 当教材对某概念没有严格划分标准时，诚实说出，但给学生一条在当前条件下最合理的判断路径。
4. JSON 文本中的引号用中文引号「」，不要用 ASCII 双引号""（会破坏 JSON 结构）。
5. 易错提醒要么给出具体的区分标准，要么留空。不写"注意区分X和Y"这类空泛表达。
6. 区分两种信息来源：教材材料中有的（可引用、可定义），和材料中没有的（只能说"题干未提及X要素"）。当材料没有提供某个选项所涉及概念的定义时，不要用自己的知识去补那个定义。
7. 引用 unit 前检查其章节路径（材料卡片中的"章节："行）。若来自特例场景（如大使馆、外交使团、某类机构），该 unit 的陈述只在特例下有效，不能当普遍原则用。
8. 不以非黑即白的方式排除选项。即使错误项本身有一定关联，也不用"不属于""不可能""因此错误"等绝对否定语气。改用比较级——"在教材框架下，题干条件更直接匹配X而非Y""正确项比错误项更吻合教材定义"。
9. 区分事实和推理。教材原文（标注了页码的）是事实——确定的；你基于事实推导出的判断是推理——不确定的。推理部分用"由此可推断""在本题条件下""相比之下更可能"等表述，不要写成和教材事实一样的确定语气。归因于某个 unit 的断言词（如"所有权转移""转移资金"）必须真的出现在该 unit 的原文中，原文没有的词不能说成教材说的。同样，教材原文的涵盖范围不能缩窄——原文说"customers or sectors"，解析不能说"仅限于行业部门"。

10. 引用具体步骤或数据点的 unit 时，检查材料中是否存在描述该流程整体阶段框架的 unit（如一级/二级审查、三段式洗钱流程等层级结构）。若存在，一并引用：用框架 unit 建立程序先后，用步骤 unit 解释具体内容。

11. primary_unit_id：从 evidence_cards 中选出对本题答案判断最重要、最核心的那一个 unit_id。它应该是 core_analysis 引用的关键证据。如果有多条引用，选起决定性作用的那条。

## 输出 JSON

{{{{
  "answer": ["A"],
  "primary_unit_id": "v7u_N000001",
  "exam_point": {{{{
    "text": "一句话，30字内，只写考什么"
  }}}},
  "core_analysis": {{{{
    "text": "定义/规则 → 题干关键事实 → 为什么正确答案成立",
    "cited_unit_ids": ["v7u_N000001"],
    "source_quote": {{{{
      "unit_id": "v7u_N000001",
      "exact_excerpt": "可选，40-240字符英文原文片段"
    }}}}
  }}}},
  "option_explanations": [
    {{{{
      "option": "B",
      "analysis": "仅错误项，不超过两句。不仅说不选，还说何时选",
      "error_type": "概念混淆|主体或阶段错配|范围或程度偏差|题干要素不匹配|证据不足",
      "stem_quotes": ["题干逐字片段"],
      "option_quotes": ["选项逐字片段"]
    }}}}
  ],
  "easy_mistake": {{{{
    "text": "有增量信息时写，否则留空\"\"",
    "cited_unit_ids": ["v7u_N000001"]
  }}}}
}}}}"""


def _valid_citations(values: Any, allowed: set[str]) -> list[str]:
    if not isinstance(values, list):
        return []
    out: list[str] = []
    for value in values:
        uid = str(value or "").strip()
        if uid in allowed and uid not in out:
            out.append(uid)
    return out


def _grounded_block(value: Any, allowed: set[str]) -> dict[str, Any]:
    """规范化需要教材引用的文本块；无引用时降级为不足文本。"""
    value = value if isinstance(value, dict) else {}
    text = _clean_prose(value.get("text", ""))
    cited = _valid_citations(value.get("cited_unit_ids"), allowed)[:3]
    if not text or not cited:
        return {"text": INSUFFICIENT_TEXT, "cited_unit_ids": []}
    return {"text": text, "cited_unit_ids": cited}


def _clean_prose(value: Any) -> str:
    text = INTERNAL_UNIT_ID_RE.sub("", str(value or ""))
    text = re.sub(r"[ \t]+([，。；：、！？])", r"\1", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    return text.strip()


def _cited_source_text(
    cited_unit_ids: list[str], unit_map: dict[str, dict[str, Any]]
) -> str:
    parts: list[str] = []
    for uid in cited_unit_ids:
        unit = unit_map.get(uid, {})
        parts.extend(
            str(unit.get(field, "") or "")
            for field in ("knowledge_zh", "knowledge_en", "en_quote")
        )
    return " ".join(parts).casefold()


def _source_has_term(source: str, term: str) -> bool:
    folded = term.casefold()
    if folded.isascii():
        return bool(
            re.search(
                rf"(?<![a-z]){re.escape(folded)}(?![a-z])",
                source,
            )
        )
    return folded in source


def _unsupported_relation_issues(
    text: str,
    cited_unit_ids: list[str],
    unit_map: dict[str, dict[str, Any]],
) -> list[str]:
    """捕获引用文本中不存在的高风险关系或模态措辞。"""
    source = _cited_source_text(cited_unit_ids, unit_map)
    gates = (
        (
            "必要性",
            ("必须", "要求", "只能", "仅能", "仅需", "必要条件"),
            (
                "必须",
                "要求",
                "只能",
                "仅能",
                "仅需",
                "必要",
                "must",
                "require",
                "requires",
                "required",
                "requiring",
                "only",
                "necessary",
            ),
        ),
        (
            "定义性",
            ("特指",),
            ("特指", "定义", "是指", "意味着", "refers to", "means", "defined as"),
        ),
        (
            "分类关系",
            ("属于", "等同于", "相当于", "极端形式"),
            ("属于", "等同", "相当", "形式", "类型", "part of", "form of", "type of", "equivalent"),
        ),
        (
            "典型性",
            ("典型",),
            ("典型", "常见", "typical", "typically", "common", "commonly"),
        ),
        (
            "频率",
            ("通常", "往往", "一般会", "经常"),
            (
                "通常",
                "往往",
                "一般",
                "经常",
                "typical",
                "typically",
                "usual",
                "usually",
                "often",
                "general",
                "generally",
            ),
        ),
        (
            "关联关系",
            ("相关", "关联"),
            ("相关", "关联", "related", "associated", "connection"),
        ),
    )
    issues: list[str] = []
    for label, prose_terms, source_terms in gates:
        if any(_source_has_term(source, term) for term in source_terms):
            continue
        for term in prose_terms:
            start = text.find(term)
            if start < 0:
                continue
            prefix = text[max(0, start - 4) : start]
            if term == "要求" and any(
                marker in prefix for marker in ("题干", "本题", "该项", "选项")
            ):
                continue
            if term in ("属于", "等同于", "相当于") and prefix.endswith(
                ("不", "未", "非")
            ):
                continue
            issues.append(f"无原文支撑的{label}措辞“{term}”")
            break
    return issues


def _normalize_grounded_block(
    value: Any,
    allowed: set[str],
    unit_map: dict[str, dict[str, Any]],
    location: str,
) -> tuple[dict[str, Any], list[str]]:
    block = _grounded_block(value, allowed)
    if not block["cited_unit_ids"]:
        return block, [f"{location}缺少合法教材引用"]
    return block, []


def _core_context_issues(text: str) -> list[str]:
    phrases = ("这种模式符合", "这一模式符合", "整体模式符合")
    return [f"核心解析混合决定性信号与伴随事实“{phrase}”" for phrase in phrases if phrase in text]


def _fallback_core_analysis(
    option_explanations: list[dict[str, Any]],
) -> dict[str, Any] | None:
    rows = [
        row
        for row in option_explanations
        if row.get("judgement") == "correct"
        and row.get("basis_type") in {
            "textbook_direct",
            "textbook_definition_application",
        }
        and row.get("analysis") != INSUFFICIENT_TEXT
    ]
    if not rows:
        return None
    text = " ".join(str(row.get("analysis", "")).strip() for row in rows).strip()
    cited: list[str] = []
    for row in rows:
        for uid in row.get("cited_unit_ids", []) or []:
            if uid not in cited:
                cited.append(uid)
    if not text or not cited:
        return None
    return {"text": text, "cited_unit_ids": cited[:3]}


def _fallback_exam_point(
    option_explanations: list[dict[str, Any]],
    framework: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    """从错误项分析和框架信息推断考点。正确项不在 option_explanations 中（由核心解析覆盖）。"""
    framework = framework or {}
    fw_type = str(framework.get("type", "") or "")
    # is_scenario with no cited_unit_ids → 纯题干题，否则为教材依据题
    if fw_type in ("is_definition", "is_domain"):
        return {"text": "本题需依据教材规则/定义判断各选项与题干条件的对应关系。"}
    if fw_type == "is_scenario":
        if framework.get("cited_unit_ids"):
            return {"text": "本题需依据教材规则结合题干事实进行场景判断。"}
        return {"text": "本题需依据题干明确事实对各选项进行直接判断。"}
    # 回退：检查是否有错误项用了教材依据
    has_textbook = any(
        row.get("basis_type") in {"textbook_direct", "textbook_definition_application"}
        for row in option_explanations
    )
    if has_textbook:
        return {"text": "本题需依据教材规则/定义判断各选项与题干条件的对应关系。"}
    return {"text": "本题需依据题干明确事实对各选项进行直接判断。"}


def _fallback_easy_mistake(
    option_explanations: list[dict[str, Any]],
    unit_map: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    correct = next(
        (
            row
            for row in option_explanations
            if row.get("judgement") == "correct"
            and row.get("basis_type") in {
                "textbook_direct",
                "textbook_definition_application",
            }
            and row.get("cited_unit_ids")
        ),
        None,
    )
    distractor = next(
        (
            row
            for row in option_explanations
            if row.get("judgement") == "incorrect"
            and row.get("basis_type") in {
                "textbook_direct",
                "textbook_definition_application",
            }
            and row.get("cited_unit_ids")
        ),
        None,
    )
    if not correct or not distractor:
        return None

    correct_uid = correct["cited_unit_ids"][0]
    distractor_uid = distractor["cited_unit_ids"][0]
    # 防止正确项和干扰项引用同一个 unit 导致两边摘要一模一样
    if correct_uid == distractor_uid and len(distractor["cited_unit_ids"]) > 1:
        distractor_uid = distractor["cited_unit_ids"][1]
    elif correct_uid == distractor_uid:
        return None
    correct_fact = _clean_prose(unit_map.get(correct_uid, {}).get("knowledge_zh", ""))
    distractor_fact = _clean_prose(
        unit_map.get(distractor_uid, {}).get("knowledge_zh", "")
    )
    correct_fact = correct_fact.rstrip("。！？；; ")
    distractor_fact = distractor_fact.rstrip("。！？；; ")
    if not correct_fact or not distractor_fact:
        return None
    text = (
        f"易将选项{distractor['option']}与选项{correct['option']}混淆。"
        f"教材对选项{distractor['option']}的相关要点是“{distractor_fact}”；"
        f"对选项{correct['option']}的相关要点是“{correct_fact}”。"
        "判断时应逐项核对题干条件，不要扩大教材中的程度或范围。"
    )
    cited = list(dict.fromkeys([distractor_uid, correct_uid]))
    return {"text": text, "cited_unit_ids": cited}


def _exact_quote(value: Any, source: str) -> str:
    quote = str(value or "").strip()
    return quote if quote and quote in source else ""


def _exact_quotes(values: Any, source: str) -> list[str]:
    if not isinstance(values, list):
        return []
    out: list[str] = []
    for value in values:
        quote = _exact_quote(value, source)
        if quote and quote not in out:
            out.append(quote)
    return out[:3]


def _legacy_source_claims(
    cited_unit_ids: list[str], unit_map: dict[str, dict[str, Any]]
) -> list[dict[str, str]]:
    claims: list[dict[str, str]] = []
    for uid in cited_unit_ids[:3]:
        unit = unit_map.get(uid, {})
        excerpt = str(unit.get("knowledge_zh", "") or "").strip()
        if not excerpt:
            excerpt = str(unit.get("en_quote") or unit.get("knowledge_en", "") or "").strip()
        if excerpt:
            claims.append({"unit_id": uid, "exact_excerpt": excerpt})
    return claims


def _quoted_text(values: list[str]) -> str:
    return "、".join(f"“{value}”" for value in values if value)


def _render_structured_option_analysis(
    *,
    expected_judgement: str,
    basis_type: str,
    evidence_status: str,
    source_claims: list[dict[str, str]],
    stem_quotes: list[str],
    option_quotes: list[str],
) -> str:
    """fallback：当 LLM 未提供有效 analysis 文本时，用模板拼装兜底。"""
    stem_text = _quoted_text(stem_quotes)
    option_text = _quoted_text(option_quotes)

    if expected_judgement == "correct":
        return f"符合上述定义。"

    if basis_type in {"textbook_direct", "textbook_definition_application"}:
        source_text = _quoted_text(
            [claim["exact_excerpt"] for claim in source_claims]
        )
        if evidence_status == "negative":
            return (
                f"教材指出{source_text}，而{option_text or '该选项'}与此不符。"
            )
        return (
            f"教材指出{source_text}。"
            f"题干中{stem_text or '的条件'}与{option_text or '该选项'}不一致。"
        )

    if basis_type == "stem_contrast":
        return (
            f"{option_text or '该选项'}与题干给出的{stem_text or '事实'}不一致。"
        )

    return INSUFFICIENT_TEXT


def _stem_contrast_text(option_quote: str, stem_quotes: list[str]) -> str:
    stem_text = "”“".join(stem_quotes)
    return (
        f"选项涉及“{option_quote}”，而题干明确描述的是“{stem_text}”；"
        "两者的关键要素不一致，因此该项不符合题干场景。"
    )


def _candidate_mentions_quote(
    option_quote: str, unit_map: dict[str, dict[str, Any]]
) -> bool:
    """检测看似纯表面干扰的选项是否其实有教材覆盖。"""
    needle = re.sub(r"\W+", "", option_quote, flags=re.UNICODE).casefold()
    if len(needle) < 2:
        return False
    for unit in unit_map.values():
        haystack = " ".join(
            str(unit.get(field, "") or "")
            for field in ("knowledge_zh", "knowledge_en", "en_quote")
        )
        normalized = re.sub(r"\W+", "", haystack, flags=re.UNICODE).casefold()
        if needle in normalized:
            return True
    return False


def _build_source_evidence(
    unit_map: dict[str, dict[str, Any]],
    exam_point: dict[str, Any],
    core_analysis: dict[str, Any],
    option_explanations: list[dict[str, Any]],
    easy_mistake: dict[str, Any],
) -> list[dict[str, Any]]:
    """物化所有被引用单元的教材原文字段，不含模型撰写的释义。"""
    usage: dict[str, list[str]] = {}

    def add(values: list[str], location: str) -> None:
        for uid in values:
            bucket = usage.setdefault(uid, [])
            if location not in bucket:
                bucket.append(location)

    add(core_analysis["cited_unit_ids"], "核心解析")
    for row in option_explanations:
        add(row["cited_unit_ids"], f"选项{row['option']}")
    add(easy_mistake["cited_unit_ids"], "易错提醒")

    rows: list[dict[str, Any]] = []
    for uid, used_by in usage.items():
        unit = unit_map[uid]
        heading = unit.get("heading_context", []) or []
        if not isinstance(heading, list):
            heading = [str(heading)] if heading else []
        page_info = _get_unit_page_map().get(uid, {})
        rows.append(
            {
                "unit_id": uid,
                "used_by": used_by,
                "knowledge_zh": str(unit.get("knowledge_zh", "") or ""),
                "en_quote": str(
                    unit.get("en_quote") or unit.get("knowledge_en", "") or ""
                ),
                "heading_context": [str(x) for x in heading if str(x).strip()],
                "content_type": str(unit.get("type", "") or ""),
                "pdf_page": page_info.get("pdf_page"),
                "printed_page": page_info.get("printed_page", ""),
            }
        )
    return rows


def _normalize_source_quote(
    raw_core: Any,
    core_analysis: dict[str, Any],
    unit_map: dict[str, dict[str, Any]],
) -> tuple[dict[str, str], list[str]]:
    raw_core = raw_core if isinstance(raw_core, dict) else {}
    raw_quote = raw_core.get("source_quote", {})
    raw_quote = raw_quote if isinstance(raw_quote, dict) else {}
    uid = str(raw_quote.get("unit_id", "") or "").strip()
    excerpt = str(raw_quote.get("exact_excerpt", "") or "").strip()
    issues: list[str] = []

    if not uid and not excerpt:
        return {}, []  # source_quote 是可选的，不提供不报错
    if not uid or not excerpt:
        return {}, ["核心解析教材英文短引不完整（缺unit_id或缺excerpt）"]
    if uid not in core_analysis.get("cited_unit_ids", []):
        issues.append("教材英文短引unit未被核心解析引用")
    unit = unit_map.get(uid)
    if not unit:
        issues.append("教材英文短引unit不在本题证据池")
        source = ""
    else:
        source = str(unit.get("en_quote") or unit.get("knowledge_en", "") or "")
    if excerpt and excerpt not in source:
        issues.append("教材英文短引不是对应英文原文的连续子串")
    if not SOURCE_QUOTE_MIN_LENGTH <= len(excerpt) <= SOURCE_QUOTE_MAX_LENGTH:
        issues.append(
            f"教材英文短引长度不在{SOURCE_QUOTE_MIN_LENGTH}-{SOURCE_QUOTE_MAX_LENGTH}字符范围"
        )
    if issues:
        return {}, issues
    return {"unit_id": uid, "exact_excerpt": excerpt}, []


def _reference_answer_conflicts(
    predicted: list[str], reference: dict[str, Any]
) -> list[str]:
    predicted_set = set(predicted)
    conflicts: list[str] = []
    for field, label in (
        ("final_answer", "题库最终参考答案"),
        ("cn_answer", "中文参考答案"),
        ("en_answer", "英文参考答案"),
    ):
        values = [str(x).strip().upper() for x in reference.get(field, []) or []]
        if values and set(values) != predicted_set:
            conflicts.append(f"AI答案与{label}冲突")
    return conflicts


def _build_software_readiness(
    result: dict[str, Any],
    predicted: list[str],
    options: dict[str, Any],
    option_explanations: list[dict[str, Any]],
    reference: dict[str, Any],
    quote_issues: list[str],
    grounding_issues: list[str],
    normalization_warnings: list[str],
) -> dict[str, Any]:
    blockers: list[str] = []

    def add(message: str) -> None:
        if message and message not in blockers:
            blockers.append(message)

    if result.get("pipeline_status") != "ok":
        add(f"盲判状态不是ok: {result.get('pipeline_status', '')}")
    for issue in result.get("validation_checks", []) or []:
        add(f"盲判机械校验失败: {issue}")
    if result.get("citation_filter_drops"):
        add("盲判存在被过滤的非法引用")
    if not predicted:
        add("AI答案为空")
    for row in option_explanations:
        if row.get("basis_type") == "insufficient":
            add(f"选项{row.get('option', '')}证据不足")
    for issue in quote_issues:
        add(issue)
    for issue in grounding_issues:
        add(issue)
    for issue in _reference_answer_conflicts(predicted, reference):
        add(issue)

    risk_flags = [
        str(flag)
        for flag in reference.get("risk_flags", []) or []
        if str(flag).strip()
    ]
    if normalization_warnings:
        risk_flags.append("normalization_recovered")
    return {
        "ready": not blockers,
        "blocking_reasons": blockers,
        "risk_flags": list(dict.fromkeys(risk_flags)),
    }


def normalize_explanation(
    parsed: dict[str, Any] | None,
    result: dict[str, Any],
    reference: dict[str, Any],
    model: str,
) -> dict[str, Any]:
    parsed = parsed if isinstance(parsed, dict) else {}
    options = result.get("options", {}) or {}
    predicted = [
        str(x).strip().upper()
        for x in result.get("predicted_answer", []) or []
        if str(x).strip().upper() in options
    ]
    unit_map = candidate_by_unit(result)
    sources = {row["option"]: row for row in enriched_option_material(result)}
    framework = result.get("decision_framework", {}) or {}
    framework_ids = {
        str(uid)
        for uid in framework.get("cited_unit_ids", []) or []
        if str(uid) in unit_map
    }
    provided_evidence_ids = {
        card["unit_id"]
        for source in sources.values()
        for key in ("evidence_cards", "supplement_cards")
        for card in source.get(key, []) or []
    }
    provided_evidence_ids.update(framework_ids)
    raw_rows = parsed.get("option_explanations", [])
    raw_rows = raw_rows if isinstance(raw_rows, list) else []
    raw_by_label = {
        str(row.get("option", "")).strip().upper(): row
        for row in raw_rows
        if isinstance(row, dict)
    }

    option_explanations: list[dict[str, Any]] = []
    grounding_issues: list[str] = []
    normalization_warnings: list[str] = []
    for label in options:
        label = str(label).strip().upper()
        source = sources.get(label, {})
        raw = raw_by_label.get(label, {})
        expected_judgement = "correct" if label in predicted else "incorrect"
        if expected_judgement == "correct":
            continue  # 正确项由核心解析覆盖，不生成 option_explanations 条目
        blind_db = source.get("decision_basis", "")
        evidence_status = source.get("evidence_status", "")
        basis_type = _DECISION_TO_BASIS.get(blind_db)
        if basis_type is None:
            # 盲判 LLM 可能输出非法值（如把 framework.type 当 decision_basis），
            # 根据 evidence_status 做防御性回退
            if evidence_status in {"direct", "indirect", "negative"}:
                basis_type = "textbook_direct"
            else:
                basis_type = "insufficient"

        raw_option_quotes = raw.get("option_quotes")
        legacy_stem_schema = not isinstance(raw_option_quotes, list)
        if not isinstance(raw_option_quotes, list):
            legacy_option_quote = str(raw.get("option_quote", "") or "").strip()
            raw_option_quotes = [legacy_option_quote] if legacy_option_quote else []
        option_quotes = _exact_quotes(raw_option_quotes, str(options[label]))
        stem_quotes = _exact_quotes(
            raw.get("stem_quotes"), str(result.get("stem", ""))
        )

        evidence_status = source.get("evidence_status", "")

        source_claims: list[dict[str, str]] = []
        if basis_type in {"textbook_direct", "textbook_definition_application"}:
            # 直接从盲判 evidence_cards 反查 unit 原文构造 source_claims，
            # 不依赖 LLM 摘录，避免逐字校验失败
            claim_uids: list[str] = []
            for card in source.get("evidence_cards", []) or []:
                uid = str(card.get("unit_id", "") or "").strip()
                if uid and uid in unit_map and uid not in claim_uids:
                    claim_uids.append(uid)
            if basis_type == "textbook_definition_application":
                for uid in framework_ids:
                    if uid and uid in unit_map and uid not in claim_uids:
                        claim_uids.append(uid)
            source_claims = _legacy_source_claims(claim_uids, unit_map)
            if not option_quotes:
                option_quotes = [str(options[label])]

        # 兜底：LLM 未提供引文时自动用选项/题干全文
        if not option_quotes:
            option_quotes = [str(options[label])]
        if not stem_quotes and basis_type in (STEM_BASIS_TYPES | {"textbook_definition_application"}):
            stem_quotes = [str(result.get("stem", ""))]

        valid_basis = True
        if basis_type in {"textbook_direct", "textbook_definition_application"}:
            if not source_claims or not option_quotes:
                valid_basis = False
                grounding_issues.append(
                    f"选项{label}教材依据缺少逐字source_claims或option_quotes"
                )
            if basis_type == "textbook_definition_application" and not stem_quotes:
                valid_basis = False
                grounding_issues.append(
                    f"选项{label}定义应用缺少逐字stem_quotes"
                )
        elif basis_type in STEM_BASIS_TYPES:
            if not stem_quotes or not option_quotes:
                valid_basis = False
                grounding_issues.append(
                    f"选项{label}{basis_type}缺少逐字题干或选项片段"
                )
            if basis_type == "stem_contrast" and expected_judgement == "correct":
                valid_basis = False
                grounding_issues.append(f"选项{label}正确项不能使用stem_contrast")
            if (
                basis_type == "stem_contrast"
                and legacy_stem_schema
                and option_quotes
                and _candidate_mentions_quote(option_quotes[0], unit_map)
            ):
                valid_basis = False
                grounding_issues.append(
                    f"选项{label}旧版stem_contrast被候选教材覆盖"
                )
        else:
            valid_basis = False

        if valid_basis:
            cited = list(
                dict.fromkeys(claim["unit_id"] for claim in source_claims)
            )[:3]
            # 优先使用 LLM 输出的自然语言 analysis，模板拼装仅作 fallback
            raw_analysis = str(raw.get("analysis", "") or "").strip()
            if len(raw_analysis) >= 6:
                analysis = _clean_prose(raw_analysis)
            else:
                analysis = _render_structured_option_analysis(
                    expected_judgement=expected_judgement,
                    basis_type=basis_type,
                    evidence_status=evidence_status,
                    source_claims=source_claims,
                    stem_quotes=stem_quotes,
                    option_quotes=option_quotes,
                )
        else:
            basis_type = "insufficient"
            # 盲判证据不足时，仍优先使用 LLM 写的 analysis（LLM 可能基于材料推理）
            raw_analysis = str(raw.get("analysis", "") or "").strip()
            if len(raw_analysis) >= 6:
                analysis = _clean_prose(raw_analysis)
            else:
                # LLM 未提供可用分析且材料不足 → 不生成占位符
                continue
            cited = []
            source_claims = []
            option_quotes = []
            stem_quotes = []

        if expected_judgement == "correct":
            error_type = "正确"
        elif basis_type == "insufficient":
            error_type = "证据不足"
        else:
            llm_error = str(raw.get("error_type", "") or "").strip()
            valid_errors = {
                "概念混淆", "主体或阶段错配", "范围或程度偏差",
                "题干要素不匹配", "证据不足",
            }
            error_type = llm_error if llm_error in valid_errors else ""
        option_explanations.append(
            {
                "option": label,
                "judgement": expected_judgement,
                "error_type": error_type,
                "basis_type": basis_type,
                "evidence_status": evidence_status,
                "analysis": analysis,
                "cited_unit_ids": cited,
                "source_claims": source_claims,
                "option_quotes": option_quotes,
                "option_quote": option_quotes[0] if option_quotes else "",
                "stem_quotes": stem_quotes,
            }
        )

    correct_rows = [
        row for row in option_explanations if row.get("judgement") == "correct"
    ]
    correct_sources_list = [
        sources[row["option"]] for row in option_explanations
        if row.get("judgement") == "correct" and row["option"] in sources
    ]
    stem_only_answer = bool(correct_sources_list) and all(
        src.get("decision_basis") == "insufficient" for src in correct_sources_list
    )
    raw_core = parsed.get("core_analysis")

    if stem_only_answer:
        decisive_stem_quotes = list(
            dict.fromkeys(
                quote
                for row in correct_rows
                for quote in row.get("stem_quotes", []) or []
            )
        )[:3]
        core_text = " ".join(
            str(row.get("analysis", "")).strip() for row in correct_rows
        ).strip()
        exam_point = {"text": "本题考查依据题干明确事实进行直接判断。"}
        core_analysis = {
            "text": core_text or INSUFFICIENT_TEXT,
            "cited_unit_ids": [],
            "source_quote": {},
        }
        easy_mistake = {
            "text": (
                f"判断时只使用题干明确给出的{_quoted_text(decisive_stem_quotes)}，"
                "不要补入题干未提供的外部定义或通常做法。"
            ),
            "cited_unit_ids": [],
        }
        quote_issues: list[str] = []
    else:
        raw_ep = (
            parsed.get("exam_point") if isinstance(parsed.get("exam_point"), dict) else {}
        )
        raw_ep_text = str(raw_ep.get("text", "") or "").strip()
        if len(raw_ep_text) >= 6:
            exam_point = {"text": _clean_prose(raw_ep_text)}
        else:
            fallback_exam = _fallback_exam_point(option_explanations, framework)
            if fallback_exam and len(fallback_exam.get("text", "")) >= 6:
                exam_point = fallback_exam
            else:
                exam_point = {"text": INSUFFICIENT_TEXT}
        core_analysis, core_issues = _normalize_grounded_block(
            raw_core,
            provided_evidence_ids,
            unit_map,
            "核心解析",
        )
        if core_issues:
            fallback_core = _fallback_core_analysis(option_explanations)
            if fallback_core:
                core_analysis.update(fallback_core)
                normalization_warnings.extend(core_issues)
            else:
                grounding_issues.extend(core_issues)
        context_issues = _core_context_issues(core_analysis.get("text", ""))
        if context_issues:
            fallback_core = _fallback_core_analysis(option_explanations)
            if fallback_core:
                core_analysis.update(fallback_core)
                normalization_warnings.extend(context_issues)
            else:
                core_analysis = {"text": INSUFFICIENT_TEXT, "cited_unit_ids": []}
                grounding_issues.extend(context_issues)
        source_quote, quote_issues = _normalize_source_quote(
            raw_core, core_analysis, unit_map
        )
        core_analysis["source_quote"] = source_quote
        # easy_mistake：优先使用 LLM 原文，LLM 输出空则表示没有真正迷惑的干扰项
        raw_easy = (
            parsed.get("easy_mistake") if isinstance(parsed.get("easy_mistake"), dict) else {}
        )
        raw_easy_text = str(raw_easy.get("text", "") or "").strip()
        if len(raw_easy_text) >= 20:
            easy_mistake = {
                "text": _clean_prose(raw_easy_text),
                "cited_unit_ids": _valid_citations(
                    raw_easy.get("cited_unit_ids", []), provided_evidence_ids
                )[:3],
            }
        else:
            easy_mistake = {"text": "", "cited_unit_ids": []}
    source_evidence = _build_source_evidence(
        unit_map,
        exam_point,
        core_analysis,
        option_explanations,
        easy_mistake,
    )
    software_readiness = _build_software_readiness(
        result,
        predicted,
        options,
        option_explanations,
        reference,
        quote_issues,
        grounding_issues,
        normalization_warnings,
    )
    # 需人工复核标记
    review_flags: list[str] = []
    predicted_set = set(predicted)
    final_set = set(reference.get("final_answer", []) or [])
    if predicted_set and final_set and predicted_set != final_set:
        review_flags.append(f"答案冲突：解析{predicted_set} vs 题库{final_set}")
    if any(row.get("basis_type") == "insufficient" for row in option_explanations):
        insuff_labels = [row["option"] for row in option_explanations if row.get("basis_type") == "insufficient"]
        review_flags.append(f"部分选项证据不足：{', '.join(insuff_labels)}")
    if result.get("validation_checks"):
        review_flags.append("盲判校验未通过")
    if result.get("pipeline_status") != "ok":
        review_flags.append(f"盲判状态异常：{result.get('pipeline_status', '')}")

    return {
        "schema_version": SCHEMA_VERSION,
        "answer": predicted,
        "primary_unit_id": str(parsed.get("primary_unit_id", "") or "").strip(),
        "exam_point": exam_point,
        "core_analysis": core_analysis,
        "option_explanations": option_explanations,
        "easy_mistake": easy_mistake,
        "source_evidence": source_evidence,
        "software_readiness": software_readiness,
        "normalization_issues": grounding_issues,
        "normalization_warnings": normalization_warnings,
        "review_flags": review_flags,
        "chapter_mappings": result.get("chapter_mappings", []),
        "reference_appendix": reference,
        "generation_metadata": {
            "prompt_version": PROMPT_VERSION,
            "model": model,
            "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        },
    }


def chapter_text(result: dict[str, Any]) -> str:
    rows = result.get("chapter_mappings", []) or []
    if not rows:
        return "未映射"
    return "；".join(
        f"{row.get('real_chapter') or row.get('chapter_id', '')} {row.get('chapter_title', '')}".strip()
        for row in rows
    )


def render_markdown(
    result: dict[str, Any],
    explanation: dict[str, Any],
    standard_question: dict[str, Any] | None = None,
    export_mode: bool = False,
) -> str:
    lines: list[str] = [f"# {result.get('question_id', '')}\n\n"]
    lines.append(f"教材章节：{chapter_text(result)}\n\n")
    lines.append(f"题型：{result.get('question_type', '')}\n\n")
    lines.append(f"题干：{result.get('stem', '')}\n\n")
    standard_question = standard_question or {}
    stem_en = str(standard_question.get("stem_en", "") or "").strip()
    options_en = standard_question.get("options_en", {}) or {}
    if stem_en:
        lines.append(f"英文题干：{stem_en}\n\n")
    lines.append("选项：\n\n")
    for label, text in (result.get("options", {}) or {}).items():
        lines.append(f"- {label}. {text}\n")
        en_text = options_en.get(label, "")
        if en_text:
            lines.append(f"  English: {en_text}\n")

    answer = "、".join(explanation.get("answer", []) or []) or "未形成答案"

    if export_mode:
        # ── 导入格式 ──
        lines.append(f"\n答案：{answer}\n\n")
        lines.append("解析：\n\n")

        exam_point = explanation.get("exam_point", {}) or {}
        lines.append(f"考点：{exam_point.get('text') or INSUFFICIENT_TEXT}\n\n")

        core_analysis = explanation.get("core_analysis", {}) or {}
        lines.append(f"核心解析：{core_analysis.get('text') or INSUFFICIENT_TEXT}\n")
        source_quote = core_analysis.get("source_quote", {}) or {}
        if source_quote.get("exact_excerpt"):
            lines.append(f"教材原句：\"{source_quote['exact_excerpt']}\"\n")
        lines.append("\n")

        for row in explanation.get("option_explanations", []) or []:
            judgement = "正确" if row.get("judgement") == "correct" else "错误"
            lines.append(f"{row.get('option', '')}项{judgement}：{row.get('analysis', '')}\n")

        easy_mistake = explanation.get("easy_mistake", {}) or {}
        easy_text = (easy_mistake.get("text") or "").strip()
        if easy_text:
            lines.append(f"\n易错提醒：{easy_text}\n")
        return "".join(lines)

    # ── 教研格式（原样）──
    lines.append(f"\n## 【AI答案】\n\n{answer}\n\n")
    flags = explanation.get("review_flags", []) or []
    if flags:
        lines.append("> **需人工复核**\n>\n")
        for f in flags:
            lines.append(f"> - {f}\n")
        lines.append("\n")
    lines.append("## 【考点】\n\n")
    exam_point = explanation.get("exam_point", {}) or {}
    lines.append((exam_point.get("text") or INSUFFICIENT_TEXT) + "\n\n")

    lines.append("## 【核心解析】\n\n")
    core_analysis = explanation.get("core_analysis", {}) or {}
    lines.append((core_analysis.get("text") or INSUFFICIENT_TEXT) + "\n\n")
    source_quote = core_analysis.get("source_quote", {}) or {}
    if source_quote.get("exact_excerpt"):
        lines.append(f"教材原句：\"{source_quote['exact_excerpt']}\"\n\n")

    lines.append("## 【错误项分析】\n\n")
    judgement_labels = {
        "correct": "正确",
        "incorrect": "错误",
        "insufficient": "证据不足",
    }
    basis_labels = {
        "textbook_direct": "教材直接依据",
        "textbook_definition_application": "教材定义应用",
        "stem_contrast": "题干对照",
        "insufficient": "证据不足",
    }
    for row in explanation.get("option_explanations", []) or []:
        judgement = judgement_labels.get(
            str(row.get("judgement", "")), str(row.get("judgement", ""))
        )
        basis = basis_labels.get(
            str(row.get("basis_type", "")), str(row.get("basis_type", ""))
        )
        is_real_analysis = row.get("analysis", "") != INSUFFICIENT_TEXT
        if row.get("basis_type") == "insufficient" and is_real_analysis:
            basis = ""
        error_tag = ""
        if row.get("error_type") and not (row.get("basis_type") == "insufficient" and is_real_analysis):
            error_tag = f"｜{row['error_type']}"
        basis_tag = f"（{basis}）" if basis else ""
        lines.append(
            f"- **{row.get('option', '')} {judgement}{basis_tag}"
            f"{error_tag}**：{row.get('analysis', '')}\n"
        )

    lines.append("\n## 【易错提醒】\n\n")
    easy_mistake = explanation.get("easy_mistake", {}) or {}
    easy_text = (easy_mistake.get("text") or "").strip()
    lines.append((easy_text or "（无）") + "\n\n")

    lines.append("## 【教材原文依据】\n\n")
    primary_uid = str(explanation.get("primary_unit_id", "") or "").strip()
    if primary_uid:
        lines.append(f"> 核心引用单元：`{primary_uid}`\n\n")
    evidence_rows = explanation.get("source_evidence", []) or []
    if not evidence_rows:
        if not core_analysis.get("cited_unit_ids"):
            lines.append("本题依据题干明确事实直接推导，无教材引用。\n\n")
        else:
            lines.append(INSUFFICIENT_TEXT + "\n\n")
    for row in evidence_rows:
        used_by = "、".join(row.get("used_by", []) or []) or "未标注"
        heading = " > ".join(row.get("heading_context", []) or []) or "未标注"
        lines.append(f"### `{row.get('unit_id', '')}`\n\n")
        lines.append(f"- 用于：{used_by}\n")
        lines.append(f"- 章节：{heading}\n")
        pdf_page = row.get("pdf_page")
        printed_page = row.get("printed_page", "")
        if pdf_page is not None or printed_page:
            page_parts: list[str] = []
            if pdf_page is not None:
                page_parts.append(f"PDF第{pdf_page}页")
            if printed_page:
                page_parts.append(f"书内第{printed_page}页")
            lines.append(f"- 页码：{' / '.join(page_parts)}\n")
        lines.append(f"- 中文要点：{row.get('knowledge_zh', '') or '未提供'}\n")
        lines.append(f"- 英文原文：{row.get('en_quote', '') or '未提供'}\n\n")

    ref = explanation.get("reference_appendix", {}) or {}
    lines.append("## 【参考答案与参考解析】\n\n")
    lines.append(f"- 题库最终参考答案：{'、'.join(ref.get('final_answer', []) or []) or '未提供'}\n")
    lines.append(f"- 中文参考答案：{'、'.join(ref.get('cn_answer', []) or []) or '未提供'}\n\n")
    lines.append("### 中文参考解析\n\n")
    lines.append((ref.get("cn_explanation") or "未提供。") + "\n\n")
    lines.append(f"- 英文参考答案：{'、'.join(ref.get('en_answer', []) or []) or '未提供'}\n\n")
    lines.append("### 英文参考解析\n\n")
    lines.append((ref.get("en_explanation") or "未提供。") + "\n\n")
    lines.append("### 答案冲突提示\n\n")
    for message in ref.get("conflict_messages", []) or ["未发现答案冲突。"]:
        lines.append(f"- {message}\n")
    return "".join(lines)


def process_file(
    path: Path,
    output_dir: Path,
    api_key: str,
    base_url: str,
    model: str,
    write_back: bool,
    standard_question: dict[str, Any],
    workbook_row: dict[str, Any],
    reasoning_effort: str = "high",
    enable_thinking: bool = True,
) -> dict[str, Any]:
    result = load_question_result(path)
    qid = result.get("question_id", path.stem.removeprefix("q_"))
    prompt = build_prompt(result, standard_question)
    from openai import OpenAI

    client = OpenAI(api_key=api_key, base_url=base_url)
    raw = call_llm(client, prompt, model=model, reasoning_effort=reasoning_effort, enable_thinking=enable_thinking)
    reference = build_reference_context(
        qid, result.get("predicted_answer", []), standard_question, workbook_row
    )
    explanation = normalize_explanation(
        parse_json_object(raw), result, reference, model
    )

    explanations_dir = output_dir / "explanations"
    explanations_dir.mkdir(parents=True, exist_ok=True)
    md_path = explanations_dir / f"{qid}.md"
    md_path.write_text(render_markdown(result, explanation, standard_question), encoding="utf-8")

    # 导入格式输出到 explanations_export/
    export_dir = output_dir / "explanations_export"
    export_dir.mkdir(parents=True, exist_ok=True)
    export_path = export_dir / f"{qid}.md"
    export_path.write_text(
        render_markdown(result, explanation, standard_question, export_mode=True),
        encoding="utf-8",
    )

    if write_back:
        result["generated_explanation"] = explanation
        result["generated_explanation_prompt"] = prompt
        result["generated_explanation_raw_output"] = raw
        with open(path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
    return {
        "question_id": qid,
        "status": "ok",
        "answer": explanation["answer"],
        "chapter_mappings": result.get("chapter_mappings", []),
        "reference_conflict": (
            reference["cn_en_conflict"] or reference["blind_final_conflict"]
        ),
        "markdown_path": str(md_path),
    }


def select_question_files(
    output_dir: Path, question_ids: list[str], limit: int | None,
    resume: bool = False,
) -> list[Path]:
    questions_dir = output_dir / "questions"
    if not questions_dir.exists():
        raise RuntimeError(f"questions 目录不存在: {questions_dir}")
    if question_ids:
        files = [questions_dir / f"q_{qid}.json" for qid in question_ids]
        missing = [str(path) for path in files if not path.exists()]
        if missing:
            raise RuntimeError("指定题号输出不存在: " + ", ".join(missing))
        return files
    files = sorted(questions_dir.glob("q_*.json"))
    if resume:
        import json as _json
        remaining = []
        for f in files:
            try:
                d = _json.loads(f.read_text(encoding="utf-8"))
                if not d.get("generated_explanation"):
                    remaining.append(f)
            except Exception:
                remaining.append(f)
        print(f"[resume] 跳过 {len(files) - len(remaining)} 题已解析，剩余 {len(remaining)} 题")
        files = remaining
    return files[:limit] if limit is not None and limit > 0 else files


def collect_generated_rows(output_dir: Path) -> list[dict[str, Any]]:
    """从磁盘上所有已写入的 V3 母版构建累积索引。"""
    rows: list[dict[str, Any]] = []
    for path in sorted((output_dir / "questions").glob("q_*.json")):
        result = load_question_result(path)
        explanation = result.get("generated_explanation", {}) or {}
        if explanation.get("schema_version") != SCHEMA_VERSION:
            continue
        qid = str(result.get("question_id", path.stem.removeprefix("q_")))
        reference = explanation.get("reference_appendix", {}) or {}
        markdown_path = output_dir / "explanations" / f"{qid}.md"
        rows.append(
            {
                "question_id": qid,
                "status": "ok" if markdown_path.exists() else "markdown_missing",
                "pipeline_status": result.get("pipeline_status", ""),
                "software_ready": bool(
                    (explanation.get("software_readiness", {}) or {}).get("ready")
                ),
                "answer": explanation.get("answer", []) or [],
                "chapter_mappings": result.get("chapter_mappings", []) or [],
                "reference_conflict": bool(
                    reference.get("cn_en_conflict")
                    or reference.get("blind_final_conflict")
                ),
                "markdown_path": str(markdown_path) if markdown_path.exists() else "",
            }
        )
    return rows


def write_index(rows: list[dict[str, Any]], output_dir: Path) -> Path:
    path = output_dir / "explanations" / "index.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# V3.1 教研解析索引\n\n",
        "| 题号 | 章节 | 盲判状态 | 软件就绪 | 答案 | 参考冲突 | Markdown |\n",
        "|---|---|---|---|---|---|---|\n",
    ]
    for row in sorted(rows, key=lambda x: x.get("question_id", "")):
        chapters = ",".join(
            item.get("real_chapter") or item.get("chapter_id", "")
            for item in row.get("chapter_mappings", []) or []
        )
        name = Path(row.get("markdown_path", "")).name if row.get("markdown_path") else ""
        link = f"[打开]({name})" if name else ""
        lines.append(
            f"| {row.get('question_id', '')} | {chapters} | "
            f"{row.get('pipeline_status', '')} | {row.get('software_ready', False)} | "
            f"{','.join(row.get('answer', []) or [])} | "
            f"{row.get('reference_conflict', '')} | {link} |\n"
        )
    path.write_text("".join(lines), encoding="utf-8")
    return path


def write_chapter_drafts(
    rows: list[dict[str, Any]], output_dir: Path
) -> list[Path]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row.get("status") != "ok" or not row.get("markdown_path"):
            continue
        for mapping in row.get("chapter_mappings", []) or []:
            chapter_id = mapping.get("real_chapter") or mapping.get("chapter_id", "")
            if not chapter_id:
                continue
            group = grouped.setdefault(
                chapter_id,
                {
                    "chapter_title": mapping.get("chapter_title", ""),
                    "rows": [],
                },
            )
            if not any(x.get("question_id") == row.get("question_id") for x in group["rows"]):
                group["rows"].append(row)

    chapter_dir = output_dir / "explanations" / "chapters"
    chapter_dir.mkdir(parents=True, exist_ok=True)
    paths: list[Path] = []
    for chapter_id, group in sorted(grouped.items()):
        question_rows = sorted(group["rows"], key=lambda x: x.get("question_id", ""))
        lines = [
            f"# {chapter_id} {group['chapter_title']} 教研解析草稿\n\n",
            f"题目数：{len(question_rows)}\n\n",
        ]
        for row in question_rows:
            content = Path(row["markdown_path"]).read_text(encoding="utf-8")
            content = re.sub(
                r"^(#{1,5}) ", lambda match: "#" + match.group(1) + " ", content, flags=re.MULTILINE
            )
            lines.append(content.rstrip() + "\n\n---\n\n")
        path = chapter_dir / f"{chapter_id}.md"
        path.write_text("".join(lines), encoding="utf-8")
        paths.append(path)
    return paths


def main() -> None:
    parser = argparse.ArgumentParser(description="从盲判结果生成有据可查的 V3.1 教研解析。")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--question-id", action="append", default=[])
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--concurrency", type=int, default=5)
    parser.add_argument("--model", default="deepseek-v4-pro")
    parser.add_argument("--reasoning-effort", default="high", choices=["high", "max"])
    parser.add_argument("--no-thinking", action="store_true", help="关闭思考模式（temperature=0.4）")
    parser.add_argument("--write-back", action="store_true")
    parser.add_argument("--resume", action="store_true", help="跳过已有 generated_explanation 的题目")
    parser.add_argument("--questions-path", default=str(DEFAULT_QUESTIONS_PATH))
    parser.add_argument("--reference-workbook", default=str(DEFAULT_REFERENCE_WORKBOOK))
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    files = select_question_files(output_dir, args.question_id, args.limit or None, resume=args.resume)
    standard = load_standard_questions(args.questions_path)
    references = load_reference_workbook(args.reference_workbook)
    selected_qids = [path.stem.removeprefix("q_") for path in files]
    missing_standard = [qid for qid in selected_qids if qid not in standard]
    missing_reference = [qid for qid in selected_qids if qid not in references]
    if missing_standard or missing_reference:
        raise RuntimeError(
            f"生成前置校验失败: 标准题库缺失={missing_standard[:10]}，"
            f"参考工作簿缺失={missing_reference[:10]}"
        )

    api_key, base_url, env_name = get_llm_config()
    print(f"[input] output_dir={output_dir} | questions={len(files)}")
    print(f"[api] 使用 {env_name} | base_url={base_url} | model={args.model}")

    rows: list[dict[str, Any]] = []
    with ThreadPoolExecutor(max_workers=max(1, args.concurrency)) as executor:
        future_map = {}
        for path in files:
            qid = path.stem.removeprefix("q_")
            future = executor.submit(
                process_file,
                path,
                output_dir,
                api_key,
                base_url,
                args.model,
                args.write_back,
                standard[qid],
                references[qid],
                args.reasoning_effort,
                not args.no_thinking,
            )
            future_map[future] = path
        for i, future in enumerate(as_completed(future_map), start=1):
            path = future_map[future]
            try:
                row = future.result()
                rows.append(row)
                print(f"[{i}/{len(files)}] {row['question_id']} | ok")
            except Exception as exc:
                qid = path.stem.removeprefix("q_")
                rows.append(
                    {
                        "question_id": qid,
                        "status": "error",
                        "answer": [],
                        "error": str(exc),
                        "traceback": traceback.format_exc(),
                    }
                )
                print(f"[{i}/{len(files)}] {qid} | ERROR: {str(exc)[:160]}")

    cumulative_rows = collect_generated_rows(output_dir)
    index_path = write_index(cumulative_rows, output_dir)
    chapter_paths = write_chapter_drafts(cumulative_rows, output_dir)
    summary_path = output_dir / "explanations" / "generation_results.json"
    summary_path.write_text(
        json.dumps(cumulative_rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"[output] index={index_path}")
    print(f"[output] chapter_drafts={len(chapter_paths)}")
    print(f"[output] summary={summary_path}")


if __name__ == "__main__":
    main()
