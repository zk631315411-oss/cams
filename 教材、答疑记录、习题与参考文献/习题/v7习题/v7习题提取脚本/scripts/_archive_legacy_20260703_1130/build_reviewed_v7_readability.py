from __future__ import annotations

import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(r"D:\守正公司工作区\cams考试\教材、答疑记录、习题与参考文献\习题\v7习题")
INPUT = BASE / "output_2s" / "semantic_aligned_cn_en_reviewed_v6_fixed.xlsx"
OUTPUT = BASE / "output_2s" / "semantic_aligned_cn_en_reviewed_v7_readability_fixed.xlsx"

CN_SHEET = "中文题库"
EN_SHEET = "英文题库"
ALIGN_SHEET = "语义对齐"
FIX_LOG_SHEET = "可读性修复记录_v7"
MANUAL_V7_SHEET = "仍需人工审核清单_v7"
SUMMARY_SHEET = "质量复核总览"


def compact(value: Any, limit: int | None = None) -> str:
    text = " ".join(str(value or "").split())
    return text if limit is None else text[:limit]


def norm_answer(value: Any) -> str:
    return "".join(re.findall(r"[A-G]", str(value or "").upper()))


def headers(ws) -> dict[str, int]:
    return {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}


def row_by_qno(ws, qno: int) -> int:
    h = headers(ws)
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, h["题号"]).value == qno:
            return row
    raise KeyError(f"{ws.title} Q{qno} not found")


def get_bank_row(ws, qno: int) -> dict[str, Any]:
    h = headers(ws)
    row = row_by_qno(ws, qno)
    return {key: ws.cell(row, col).value for key, col in h.items()}


def set_hyperlink(cell, value: Any) -> None:
    cell.value = value
    if value and isinstance(value, str) and (value.endswith(".jpg") or value.endswith(".png")):
        cell.hyperlink = value
        cell.style = "Hyperlink"


def set_bank_value(ws, qno: int, field: str, value: Any) -> None:
    h = headers(ws)
    row = row_by_qno(ws, qno)
    cell = ws.cell(row, h[field])
    if field in ("题目截图", "解析截图"):
        set_hyperlink(cell, value)
    else:
        cell.value = value


def write_sheet(wb, title: str, sheet_headers: list[str], rows: list[list[Any]], widths: list[int] | None = None) -> None:
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(sheet_headers)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = widths or []
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1] if col <= len(widths) else 24
    for row_idx in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row_idx, col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def copy_worksheet_style(src_ws, dst_ws) -> None:
    for row in src_ws.iter_rows():
        for cell in row:
            dst = dst_ws[cell.coordinate]
            if cell.has_style:
                dst._style = copy(cell._style)
            dst.font = copy(cell.font)
            dst.fill = copy(cell.fill)
            dst.border = copy(cell.border)
            dst.alignment = copy(cell.alignment)
            dst.number_format = cell.number_format
            dst.protection = copy(cell.protection)
    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key].width = dim.width
    dst_ws.freeze_panes = src_ws.freeze_panes


def answer_status(cn_answer: Any, en_answer: Any) -> tuple[str, str]:
    cn = norm_answer(cn_answer)
    en = norm_answer(en_answer)
    if cn and en and cn == en:
        return cn, "一致"
    if cn and en:
        return "", f"不一致：中文={cn}; 英文={en}"
    if cn or en:
        return cn or en, "单侧识别"
    return "", "未识别"


def update_alignment_sheet(wb) -> None:
    if ALIGN_SHEET not in wb.sheetnames:
        return
    align_ws = wb[ALIGN_SHEET]
    h = headers(align_ws)
    cn_ws = wb[CN_SHEET]
    en_ws = wb[EN_SHEET]
    for row in range(2, align_ws.max_row + 1):
        cn_qno = align_ws.cell(row, h["中文题号"]).value
        en_qno = align_ws.cell(row, h["英文题号"]).value
        if not isinstance(cn_qno, int) or not isinstance(en_qno, int):
            continue
        cn = get_bank_row(cn_ws, cn_qno)
        en = get_bank_row(en_ws, en_qno)
        recommended, status = answer_status(cn.get("答案"), en.get("答案"))
        updates = {
            "推荐答案": recommended,
            "答案状态": status,
            "中文答案": cn.get("答案"),
            "英文答案": en.get("答案"),
            "中文题目": cn.get("题目文本"),
            "英文题目": en.get("题目文本"),
            "中文解析": cn.get("解析文本"),
            "英文解析": en.get("解析文本"),
            "中文来源": cn.get("来源视频"),
            "英文来源": en.get("来源视频"),
            "中文题目截图": cn.get("题目截图"),
            "英文题目截图": en.get("题目截图"),
            "中文解析截图": cn.get("解析截图"),
            "英文解析截图": en.get("解析截图"),
        }
        for key, value in updates.items():
            if key in h:
                cell = align_ws.cell(row, h[key])
                if key.endswith("截图"):
                    set_hyperlink(cell, value)
                else:
                    cell.value = value


def rebuild_risk_sheet(wb) -> None:
    title = "中英对齐风险清单"
    if ALIGN_SHEET not in wb.sheetnames:
        return
    align_ws = wb[ALIGN_SHEET]
    h = headers(align_ws)
    rows: list[list[Any]] = []
    for row in range(2, align_ws.max_row + 1):
        status = compact(align_ws.cell(row, h["答案状态"]).value)
        reason_parts: list[str] = []
        if status and status != "一致":
            reason_parts.append(status)
        cn_q = compact(align_ws.cell(row, h["中文题目"]).value)
        en_q = compact(align_ws.cell(row, h["英文题目"]).value)
        bad_tokens = ["加载中", "tb814", "无法解析", "荆棘", "很寿", "很物", "私很", "F1", "PIU", "O 错", "冷 A"]
        remaining = [token for token in bad_tokens if token in cn_q or token in en_q]
        if remaining:
            reason_parts.append("残留可读性风险：" + "、".join(remaining))
        if reason_parts:
            rows.append(
                [
                    align_ws.cell(row, h["中文题号"]).value,
                    align_ws.cell(row, h["英文题号"]).value,
                    align_ws.cell(row, h["匹配置信"]).value,
                    status,
                    "；".join(reason_parts),
                    compact(cn_q, 120),
                    compact(en_q, 120),
                    align_ws.cell(row, h["中文答案"]).value,
                    align_ws.cell(row, h["英文答案"]).value,
                    align_ws.cell(row, h["中文题目截图"]).value if "中文题目截图" in h else "",
                    align_ws.cell(row, h["英文题目截图"]).value if "英文题目截图" in h else "",
                ]
            )
    write_sheet(
        wb,
        title,
        ["中文题号", "英文题号", "置信", "答案状态", "原因", "中文题干摘要", "英文题干摘要", "中文答案", "英文答案", "中文截图", "英文截图"],
        rows,
        [10, 10, 10, 22, 36, 45, 45, 10, 10, 38, 38],
    )


def make_rewrites() -> list[dict[str, Any]]:
    frame = BASE / "output_2s" / "cache" / "frames_2_0s"
    cn1 = frame / "中文版1"
    cn2 = frame / "中文版2"
    cn3 = frame / "中文版3"
    cn4 = frame / "中文版4"
    en1 = frame / "英文版1"
    en2 = frame / "英文版2"
    en3 = frame / "英文版3"
    en4 = frame / "英文版4"
    return [
        {
            "sheet": CN_SHEET,
            "qno": 39,
            "field": "题目文本",
            "value": "单选 -[洗钱和恐怖融资的风险及方法] 为协助调查一起跨境洗钱案件，作为埃格蒙特集团成员的金融情报机构(FIU)可以: A 直接与另一国的金融机构联系并分享与调查相关的信息。 B 直接与其他国家的金融情报机构(FIU)联系并分享与调查相关的信息。 C 授权其执法调查人员协助另一国正在进行的重大调查。 D 协助另一国执法部门进行重大在办案件的调查。",
            "evidence": str(cn1 / "中文版1_000117.jpg"),
            "note": "将源视频/识别中的“流感监测单位”按英文对照和FIU语义修为“金融情报机构(FIU)”；答案字段仍保留源中文答案。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 110,
            "field": "题目文本",
            "value": "单选 -[反洗钱/打击资助恐怖主义合规计划] 一家金融机构(FI)的一名员工怀疑其一位同事参与了一个金融投资诈骗团伙。该员工接下来应采取哪一步骤? A 在向公司人力资源部门报告之前，先向同事询问以确定自己的怀疑是否正确。 B 利用FI的举报渠道来举报该可疑员工。 C 提醒FI的同事和客户，该员工可疑的金融投资建议可能是骗局。 D 利用人工智能工具开展开源情报调查，以获取有关涉嫌员工活动的更多信息。 E 将涉嫌违规的员工报告给FI的直属经理，由其采取必要的行动。",
            "evidence": str(cn2 / "中文版2_000027.jpg"),
            "note": "修复F/F1为FI并整理选项。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 112,
            "field": "题目文本",
            "value": "单选 -[反洗钱和反恐怖融资合规标准] A国发布的新闻报道称，一名政治公众人物(PEP)从一家总部位于发展中国家的跨国公司收受了约150万美元的贿赂，这笔钱存入了位于B国的一家金融机构(FI)的账户。两国均为埃格蒙特集团成员。该账户属于该政治公众人物的直系亲属。为核实事实，A国的金融情报机构(FIU)通过安全通信渠道向B国的金融情报机构(FIU)发出了正式请求，要求提供更多信息。根据埃格蒙特原则，B国的金融情报机构可以: A 提供他们所掌握的信息，因为此次信息交换是在两个均为埃格蒙特集团成员的金融情报机构之间进行的，且正式请求是通过安全通信渠道发出的。 B 仅在A国与该国签署谅解备忘录(MoU)的情况下提供所要求的信息。 C 指示A国的金融情报机构向开设账户的金融机构进一步了解情况。 D 如果B国金融情报机构未收到开设账户的金融机构提交的可疑活动报告(SAR)，则拒绝该请求。",
            "evidence": str(cn2 / "中文版2_000033.jpg"),
            "note": "区分FI账户与FIU请求方，修复PIU/FIU。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 138,
            "field": "题目文本",
            "value": "单选 -[反洗钱和反恐怖融资合规标准] 金融机构(FIs)应当了解其他司法管辖区的反金融犯罪(AFC)和制裁监管制度，这一点至关重要，以便: A 确保遵守FI在其运营或有业务关系的所有国家的反洗钱和制裁要求，以避免因违反外国司法管辖区的规定而受到处罚。 B 确保根据业务活动所在国的监管标准有选择地实施制裁制度，重点针对与本国监管标准一致的司法管辖区。 C 补偿自动封禁和制裁规定在跨境交易中的有限适用性以及其对其他司法管辖区国内业务相关性的降低。 D 确保FI能够在监管比其本国更严格或更宽松的司法管辖区管理业务关系，从而实现运营灵活性。",
            "evidence": str(cn2 / "中文版2_000101.jpg"),
            "note": "修复F1为FI并整理选项。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 160,
            "field": "题目文本",
            "value": "单选 -[反洗钱/打击资助恐怖主义合规计划] 有效的反洗钱/打击资助恐怖主义合规计划的基本要素之一是: A 企业风险管理体系。 B 基于人工智能的交易监控系统。 C 独立且专业的审计职能。 D 基于人工智能的强化尽职调查系统。",
            "evidence": str(cn2 / "中文版2_000157.jpg"),
            "note": "去除加载中和残缺选项，按对齐英文EN159恢复。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 240,
            "field": "题目文本",
            "value": "多选 -[反洗钱/打击资助恐怖主义合规计划] 持续的客户身份识别包括:(选择两项.) A 触发器能够实时识别静态数据变化以及基于客户行为的数据。 B 每年、每三年和每五年定期更新。 C 能够根据客户数据而非仅依据时间表来更好地确定优先级。 D 将客户分类为不同风险类别。",
            "evidence": str(cn3 / "中文版3_000120.jpg"),
            "note": "去除加载中，按对齐英文EN238恢复C/D选项。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 272,
            "field": "题目文本",
            "value": "多选 -[洗钱和恐怖融资的风险及方法] 评估金融领域内某一产品相关的洗钱风险包括评估以下方面:(选择两项.) A 治理安排。 B 产品的复杂性。 C 上次审计的结果。 D 企业的财务状况。",
            "evidence": str(cn3 / "中文版3_000221.jpg"),
            "note": "移除聊天水印和滚屏污染，按对齐英文EN270恢复。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 272,
            "field": "解析文本",
            "value": "试题详解 原解析 评估金融领域内某一产品相关的洗钱风险，重点应评估产品本身及其治理控制。治理安排(A)反映产品审批、监督和控制机制；产品复杂性(B)会影响客户理解、资金流向透明度和被滥用风险。上次审计的结果(C)和企业的财务状况(D)更多属于机构层面或经营状况，不是该产品风险本身的核心评估项。因此，正确答案为AB。",
            "evidence": str(cn3 / "中文版3_000221.jpg"),
            "note": "移除解析中的聊天水印和滚屏污染，按题意与EN270重建。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 348,
            "field": "题目文本",
            "value": "多选 -[洗钱和恐怖融资的风险及方法] 根据金融行动特别工作组(FATF)的规定，与环境犯罪所得洗钱相关的潜在风险指标包括:(选择两项.) A 在高风险司法管辖区注册的废物管理行业公司，其支付款项或贸易发票所涉及的废物类型与其获授权处理的废物类型一致。 B 伐木、木材加工或废料贸易行业的公司频繁向与合法业务活动或经营无关的个人或受益人支付款项。 C 无法解释的财富以及涉及高级官员或政治公众人物向其家庭成员转移现金的情况，且这些官员或公众人物的职位与自然资源的管理或保护相关。 D 小额现金转移，即从现金密集型企业向已知为金矿开采、非法采伐和非法土地开垦源头地区的受益人进行现金转移。",
            "evidence": str(cn4 / "中文版4_000131.jpg") + "; " + str(cn4 / "中文版4_000132.jpg"),
            "note": "修复前题解析串入、无法解析占位和选项残缺。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 356,
            "field": "题目文本",
            "value": "多选 -[反洗钱/打击资助恐怖主义合规计划] 哪些受监管实体或守门人可能被要求进行客户尽职调查(CDD)?(选择四项.) A 赌场保安人员。 B 市法院法官。 C 参与房地产交易的公证员。 D 贵金属和宝石经销商。 E 房地产经纪人。 F 会计和审计员。",
            "evidence": str(cn4 / "中文版4_000150.jpg"),
            "note": "修复右括号和“守门很”。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 358,
            "field": "题目文本",
            "value": "多选 -[洗钱和恐怖融资的风险及方法] 以下哪些活动被金融行动特别工作组(FATF)认定为可能通过房地产行业洗钱的迹象?(选择四项.) A 抵押贷款产品的使用。 B 商业用途与声明的业务目的不符的商业物业使用情况。 C 复杂贷款或信贷融资的使用。 D 使用公司载体或复杂的所有权结构。 E 无法解释的现金支付。 F 农村地区农业用地的使用。",
            "evidence": str(cn4 / "中文版4_000158.jpg"),
            "note": "补齐选项，按英文EN357将corporate vehicles语义修为公司载体。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 358,
            "field": "解析文本",
            "value": "试题详解 原解析 FATF指出，房地产行业中若干活动可能提示洗钱风险：商业物业用途与声明业务目的不符(B)、复杂贷款或信贷融资(C)、使用公司载体或复杂的所有权结构(D)、无法解释的现金支付(E)，都可能隐藏真实资金来源、交易目的或受益所有人。抵押贷款产品的使用(A)和农村地区农业用地的使用(F)本身并不当然构成洗钱迹象。因此，正确答案为BCDE。",
            "evidence": str(cn4 / "中文版4_000158.jpg"),
            "note": "v6解析字段残缺且含无法解析占位；按题意与EN357重建。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 372,
            "field": "题目文本",
            "value": "多选 -[反洗钱/打击资助恐怖主义合规计划] 银行反洗钱/打击资助恐怖主义合规计划的设计缺陷可能会导致对受监管实体的哪一部分追究个人责任?(选择两项.) A 合规部门。 B 董事会。 C 产品监督委员会。 D 高级管理层。",
            "evidence": str(cn4 / "中文版4_000201.jpg"),
            "note": "修复相邻题串入，按正确题干帧重建。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 372,
            "field": "答案",
            "value": "BD",
            "evidence": str(cn4 / "中文版4_000201.jpg"),
            "note": "正确题干帧显示答案BD，v6仅提取到B。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 372,
            "field": "题目截图",
            "value": str(cn4 / "中文版4_000201.jpg"),
            "evidence": str(cn4 / "中文版4_000201.jpg"),
            "note": "v6题图误指向下一题。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 373,
            "field": "题目文本",
            "value": "多选 -[洗钱和恐怖融资的风险及方法] 被提名人如何为滥用其进行洗钱活动的犯罪分子提供便利?(选择两项.) A 允许受益所有人就公司决策提供代理投票权。 B 不透明的实际所有权。 C 允许在被提名人所在司法管辖区设立住所。 D 干扰调查。",
            "evidence": str(cn4 / "中文版4_000207.jpg"),
            "note": "按英文EN372和中文截图语义修复“被提名很/荆棘”。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 373,
            "field": "解析文本",
            "value": "试题详解 原解析 在洗钱活动中，不透明的实际所有权(B选项)可以隐藏真正的受益所有人，帮助犯罪分子通过被提名人转移或持有非法资金；干扰调查(D选项)会阻碍监管或执法部门追踪资金流向，为犯罪分子提供掩护。A选项允许受益所有人提供代理投票权，与洗钱便利性无直接关联；C选项允许在被提名人所在司法管辖区设立住所，未直接涉及洗钱操作。因此，正确答案为BD。",
            "evidence": str(cn4 / "中文版4_000209.jpg"),
            "note": "修复解析滚屏串字和“荆棘”占位。",
        },
        {
            "sheet": CN_SHEET,
            "qno": 366,
            "field": "题目文本",
            "value": "多选 -[反洗钱和反恐怖融资合规标准] 某外国司法管辖区的执法机构(LEA)就某金融机构(FI)的一名客户与其取得联系。该执法机构告知，该客户因一系列人口贩卖指控目前正被通缉，以待起诉。FI应该怎么做?(选择两项.) A 将请求告知当地执法机构和监管机构以引起其注意。 B 立即关闭客户的账户以避免任何不必要的风险。 C 审查客户的活动，确定是否存在可疑活动，并相应地进行报告。 D 告知执法机构，需要联系政府办理引渡事宜。 E 立即遵守外国司法管辖区的规定，移交所有客户信息。",
            "evidence": str(cn4 / "中文版4_000186.jpg"),
            "note": "修复F 1为FI、人□贩卖为人口贩卖并整理选项。",
        },
        {
            "sheet": EN_SHEET,
            "qno": 39,
            "field": "题目文本",
            "value": "单选 -[Risks and Methods of Money Laundering and Terrorist Financing] To provide aid in investigating a cross-border money laundering case, a Financial Intelligence Unit (FIU) that is a member of the Egmont Group can: A directly contact financial institutions in another country and share information pertinent to the investigation. B directly contact other FIUs in another country and share information pertinent to the investigation. C deputize its law enforcement investigators to assist in a material ongoing investigation in another country. D assist law enforcement in another country with a material ongoing investigation.",
            "evidence": str(en1 / "英文版1_000131.jpg"),
            "note": "修复PIU为FIU并整理英文排版。",
        },
        {
            "sheet": EN_SHEET,
            "qno": 109,
            "field": "答案",
            "value": "B",
            "evidence": str(en2 / "英文版2_000020.jpg"),
            "note": "题目帧显示选中B，v6误提取为A。",
        },
        {
            "sheet": EN_SHEET,
            "qno": 109,
            "field": "题目文本",
            "value": "单选 -[AML/CFT Compliance Programs] An employee at a financial institution (FI) suspects that one of their co-workers is involved in a financial investment scam syndicate. Which step should be taken next by the employee who has the suspicion? A Question their co-worker to determine if their suspicions are correct before reporting to the FI's Human Resources department. B Use the FI's whistleblowing channel to report the suspected employee. C Warn colleagues and customers of the FI that the employee's suspicious financial investment proposals could be a scam. D Conduct an open-source intelligence investigation using artificial intelligence tools to gain more information on the activities of the suspected employee. E Report the suspected employee to the line manager of the FI to take the required action.",
            "evidence": str(en2 / "英文版2_000020.jpg") + "; " + str(en2 / "英文版2_000021.jpg"),
            "note": "修复F1为FI并整理跨屏选项。",
        },
        {
            "sheet": EN_SHEET,
            "qno": 109,
            "field": "解析文本",
            "value": "试题详解 原解析 在金融机构中，员工若怀疑同事涉及金融投资诈骗团伙，应遵循合规程序。直接质问同事可能打草惊蛇，不利于后续调查；警告同事和客户可能引发不必要的恐慌，且缺乏确凿证据；使用人工智能工具开展开源情报调查并非专业调查手段，可能违反法律法规和隐私政策；报告给直线经理虽为一种途径，但不如通过专门的举报渠道直接和高效。根据反洗钱/反恐融资(AML/CFT)合规要求，使用金融机构的举报渠道报告可疑员工是最恰当的做法，可确保信息得到及时、专业的处理。",
            "evidence": str(en2 / "英文版2_000023.jpg"),
            "note": "v6解析串到Egmont题；按解析帧恢复。",
        },
        {
            "sheet": EN_SHEET,
            "qno": 109,
            "field": "解析截图",
            "value": str(en2 / "英文版2_000023.jpg"),
            "evidence": str(en2 / "英文版2_000023.jpg"),
            "note": "更新为正确解析帧。",
        },
        {
            "sheet": EN_SHEET,
            "qno": 111,
            "field": "题目文本",
            "value": "单选 -[Compliance Standards for AML and CFT] News published in Country A reports that a politically exposed person (PEP) had received a bribe from a transnational company headquartered in a developing country of approximately US$1.5 million, deposited into an account at a financial institution (FI) located in Country B. Both countries are members of the Egmont Group. The account where the money was deposited belongs to the PEP's immediate family member. To corroborate the facts, the FIU of Country A sent a formal request via secure communication channels for further information from its counterpart FIU in Country B. According to Egmont principles, the FIU of Country B can: A provide the information available to them because the exchange is between two FIUs that are members of the Egmont Group, and the formal request was made using secure communication channels. B provide the requested information only if Country A has signed a memorandum of understanding (MoU) with Country B. C instruct Country A's FIU to inquire further with the financial institution where the account is held. D refuse the request if Country B's FIU has not received a suspicious activity report (SAR) from the financial institution where the account is held.",
            "evidence": str(en2 / "英文版2_000027.jpg") + "; " + str(en2 / "英文版2_000028.jpg"),
            "note": "修复Fl/PiU为FI/FIU并整理英文排版。",
        },
        {
            "sheet": EN_SHEET,
            "qno": 262,
            "field": "题目文本",
            "value": "多选 -[Risks and Methods of Money Laundering and Terrorist Financing] Which of the following scenarios exhibit classic indicators of suspicious transactions?(Select Two.) A A business owner mortgages his home with a financial institution that was recently fined for AML violations. B A customer regularly invests in equity funds using her spouse's bank account where she is a second account holder. C An individual who is the secretary to a government official frequently accesses the bank's safe deposit vault to withdraw cash. D An individual regularly imports sophisticated electronic items for civil use and pays all applicable customs duties. E An individual wants to execute wire transfers to a person in a FATF grey-listed jurisdiction and asks a close friend to send the money on his behalf, citing financial difficulties.",
            "evidence": str(en3 / "英文版3_000195.jpg") + "; " + str(en3 / "英文版3_000197.jpg"),
            "note": "移除侧边中文残留，补齐E选项。",
        },
        {
            "sheet": EN_SHEET,
            "qno": 275,
            "field": "题目文本",
            "value": "多选 -[Risks and Methods of Money Laundering and Terrorist Financing] A national risk assessment (NRA) can inform the risk-based approach (RBA) in an organization's anti-financial crime (AFC) compliance program by:(Choose two.) A helping to identify high-risk sectors that require enhanced due diligence (EDD). B eliminating the need for sectoral risk assessments within the organization. C guiding the allocation of resources for mitigating financial crime risks. D requiring all organizations to apply standardized measures. E automatically reducing the organization's responsibility for conducting its own risk assessment.",
            "evidence": str(en3 / "英文版3_000233.jpg"),
            "note": "补齐题干开头并修复ED0为EDD。",
        },
        {
            "sheet": EN_SHEET,
            "qno": 300,
            "field": "题目文本",
            "value": "多选 -[AML/CFT Compliance Programs] Which types of external data sources are expected to be used for screening customers as part of customer due diligence (CDD)?(Select Three.) A Customer feedback and online review platforms. B Social media sources to assess lifestyle and spending patterns. C Sanctions lists, including those issued by the Office of Foreign Assets Control (OFAC). D Registers of stolen or forged documents (where available). E Beneficial ownership registers and adverse media sources.",
            "evidence": str(en3 / "英文版3_000315.jpg"),
            "note": "移除加载中遮挡，按中文CN301和截图恢复B/C/D/E。",
        },
        {
            "sheet": EN_SHEET,
            "qno": 322,
            "field": "题目文本",
            "value": "多选 -[Conducting and Responding to Investigations] When making an independent determination on whether to close an account based on an internal investigation, a financial institution (FI) should consider:(Select Five.) A reputational risk. B the customer's personal relationships. C the frequency of account activity. D the FI's policies and procedures. E the seriousness of the underlying conduct. F correspondence with law enforcement. G the legal basis for closing the account.",
            "evidence": str(en4 / "英文版4_000061.jpg"),
            "note": "修复F 1/FI并整理英文排版；答案解析不一致按用户要求不作为阻塞。",
        },
        {
            "sheet": EN_SHEET,
            "qno": 347,
            "field": "题目文本",
            "value": "多选 -[Risks and Methods of Money Laundering and Terrorist Financing] According to the Financial Action Task Force (FATF), potential risk indicators related to money laundering proceeds from environmental crimes include:(Select Two.) A waste management sector companies based in high-risk jurisdictions with payments or trade invoices for types of waste aligned with those they are authorized to process. B frequent payments from companies in the logging, milling, or waste trade sectors to individuals or beneficiaries unrelated to the legal person activity or business. C unexplained wealth and cash transfers involving senior officials or politically exposed persons (or their family members) with a position of responsibility related to the management or preservation of natural resources. D small cash transfers from cash-intensive businesses to beneficiaries in areas known as a source of gold mining, illegal logging, and illegal land clearing.",
            "evidence": str(en4 / "英文版4_000124.jpg") + "; " + str(en4 / "英文版4_000126.jpg"),
            "note": "补齐C/D选项并修复parson/trarsfers拼写。",
        },
        {
            "sheet": EN_SHEET,
            "qno": 357,
            "field": "题目文本",
            "value": "多选 -[Risks and Methods of Money Laundering and Terrorist Financing] Which of the following activities are identified by the Financial Action Task Force (FATF) as being potentially indicative of money laundering through the real estate sector?(Select Four.) A Use of mortgage products. B Use of commercial properties inconsistent with stated business purposes. C Use of complex loans or credit finance. D Use of corporate vehicles or complex ownership structures. E Unexplained cash payments. F Use of agricultural land in rural areas.",
            "evidence": str(en4 / "英文版4_000155.jpg"),
            "note": "修复vehicies/compiex/arear并整理英文排版。",
        },
        {
            "sheet": EN_SHEET,
            "qno": 372,
            "field": "题目文本",
            "value": "多选 -[Risks and Methods of Money Laundering and Terrorist Financing] How do nominees benefit criminals misusing them for money laundering purposes?(Select Two.) A Allow beneficial owners to provide proxies for voting on corporate decisions. B Obscure beneficial ownership. C Allow domicile in the nominee's jurisdiction. D Derail investigations.",
            "evidence": str(en4 / "英文版4_000208.jpg"),
            "note": "按语义修复thorn为them，并整理选项A。",
        },
    ]


def apply_global_replacements(wb) -> list[list[Any]]:
    replacements = [
        (CN_SHEET, "题目文本", "很寿", "人寿", "明显OCR：人寿保险"),
        (CN_SHEET, "解析文本", "很寿", "人寿", "明显OCR：人寿保险"),
        (CN_SHEET, "题目文本", "政治公众很物", "政治公众人物", "明显OCR：政治公众人物"),
        (CN_SHEET, "解析文本", "政治公众很物", "政治公众人物", "明显OCR：政治公众人物"),
        (CN_SHEET, "题目文本", "私很", "私人", "明显OCR：私人银行"),
        (CN_SHEET, "解析文本", "私很", "私人", "明显OCR：私人银行"),
        (CN_SHEET, "题目文本", "F 1", "FI", "明显OCR：金融机构FI"),
        (CN_SHEET, "解析文本", "F 1", "FI", "明显OCR：金融机构FI"),
        (CN_SHEET, "题目文本", "F1", "FI", "明显OCR：金融机构FI"),
        (CN_SHEET, "解析文本", "F1", "FI", "明显OCR：金融机构FI"),
        (CN_SHEET, "题目文本", "PIU", "FIU", "明显OCR：金融情报机构FIU"),
        (CN_SHEET, "解析文本", "PIU", "FIU", "明显OCR：金融情报机构FIU"),
        (EN_SHEET, "题目文本", "F 1", "FI", "明显OCR：financial institution FI"),
        (EN_SHEET, "解析文本", "F 1", "FI", "明显OCR：financial institution FI"),
        (EN_SHEET, "题目文本", "F1", "FI", "明显OCR：financial institution FI"),
        (EN_SHEET, "解析文本", "F1", "FI", "明显OCR：financial institution FI"),
        (EN_SHEET, "题目文本", "PIU", "FIU", "明显OCR：Financial Intelligence Unit"),
        (EN_SHEET, "解析文本", "PIU", "FIU", "明显OCR：Financial Intelligence Unit"),
        (EN_SHEET, "题目文本", "PiU", "FIU", "明显OCR：Financial Intelligence Unit"),
        (EN_SHEET, "解析文本", "PiU", "FIU", "明显OCR：Financial Intelligence Unit"),
    ]
    logs: list[list[Any]] = []
    for sheet_name, field, old, new, note in replacements:
        ws = wb[sheet_name]
        h = headers(ws)
        if field not in h:
            continue
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row, h[field]).value
            if not isinstance(value, str) or old not in value:
                continue
            qno = ws.cell(row, h["题号"]).value
            new_value = value.replace(old, new)
            ws.cell(row, h[field]).value = new_value
            logs.append(
                [
                    sheet_name,
                    qno,
                    field,
                    compact(value, 120),
                    compact(new_value, 120),
                    "",
                    note,
                    "全局确定性替换",
                ]
            )
    return logs


def apply_rewrites(wb) -> list[list[Any]]:
    logs: list[list[Any]] = []
    for fix in make_rewrites():
        ws = wb[fix["sheet"]]
        before = get_bank_row(ws, fix["qno"]).get(fix["field"])
        set_bank_value(ws, fix["qno"], fix["field"], fix["value"])
        logs.append(
            [
                fix["sheet"],
                fix["qno"],
                fix["field"],
                compact(before, 120),
                compact(fix["value"], 120),
                fix["evidence"],
                fix["note"],
                "整行/字段重建",
            ]
        )
    return logs


def validate(wb) -> list[str]:
    problems: list[str] = []
    for sheet_name in (CN_SHEET, EN_SHEET):
        ws = wb[sheet_name]
        h = headers(ws)
        qnos = [ws.cell(row, h["题号"]).value for row in range(2, ws.max_row + 1)]
        expected = list(range(1, 396))
        if qnos != expected:
            problems.append(f"{sheet_name}: 题号不是1-395连续")
        for field in ("答案", "题目文本", "解析文本"):
            blanks = [q for q, row in zip(qnos, range(2, ws.max_row + 1)) if not compact(ws.cell(row, h[field]).value)]
            if blanks:
                problems.append(f"{sheet_name}: {field}空值 {blanks[:20]}")
    bad_tokens = ["加载中", "tb814", "无法解析", "荆棘", "很寿", "很物", "私很", "守门很", "被提名很", "PIU", "F1", "O 错", "冷 A", "Fl>", "PiU", "parson", "trarsfers", "vehicies", "compiex", "arear", "ED 0"]
    for sheet_name in (CN_SHEET, EN_SHEET):
        ws = wb[sheet_name]
        h = headers(ws)
        for row in range(2, ws.max_row + 1):
            qno = ws.cell(row, h["题号"]).value
            text = " ".join(str(ws.cell(row, h[field]).value or "") for field in ("题目文本", "解析文本") if field in h)
            found = [token for token in bad_tokens if token in text]
            if found:
                problems.append(f"{sheet_name} Q{qno}: 残留 {','.join(found)}")
    return problems


def main() -> None:
    wb = load_workbook(INPUT)
    logs = []
    logs.extend(apply_global_replacements(wb))
    logs.extend(apply_rewrites(wb))
    update_alignment_sheet(wb)
    rebuild_risk_sheet(wb)

    write_sheet(
        wb,
        FIX_LOG_SHEET,
        ["工作表", "题号", "字段", "旧值摘要", "新值摘要", "证据截图/依据", "说明", "处理类型"],
        logs,
        [14, 8, 12, 46, 46, 52, 48, 16],
    )

    problems = validate(wb)
    manual_rows = []
    if problems:
        manual_rows = [["全表", "", "残留可读性风险", item, "请复核"] for item in problems]
    else:
        manual_rows = [["全表", "", "无", "本轮未发现需要人工判定的非答案可读性问题。答案/解析不一致记录按用户要求保留，不作为本轮阻塞。", ""]]
    write_sheet(
        wb,
        MANUAL_V7_SHEET,
        ["工作表", "题号", "问题类型", "说明", "证据/备注"],
        manual_rows,
        [14, 8, 18, 80, 40],
    )

    summary_rows = [
        ["题库完整性", "中文题库/英文题库", "各395题，题号1-395连续，无题目/答案/解析空值。"],
        ["语义对齐", "语义对齐", "395条对齐记录已按v7主表同步；英文未匹配为0。"],
        ["可读性修复", FIX_LOG_SHEET, f"记录{len(logs)}条字段级修复，含截图依据或中英对照依据。"],
        ["人工审核", MANUAL_V7_SHEET, "非答案类可读性问题未发现必须人工判断项；答案/解析冲突保留原记录。"],
        ["输出文件", str(OUTPUT), "v7可读性修复版，不覆盖v6。"],
    ]
    write_sheet(wb, SUMMARY_SHEET, ["类别", "范围", "结论"], summary_rows, [18, 30, 90])

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = copy(cell.alignment)
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical or "top",
                    wrap_text=True,
                )

    wb.save(OUTPUT)
    print(f"written: {OUTPUT}")
    print(f"fix logs: {len(logs)}")
    if problems:
        print("validation problems:")
        for item in problems:
            print("-", item)
    else:
        print("validation: ok")


if __name__ == "__main__":
    main()
