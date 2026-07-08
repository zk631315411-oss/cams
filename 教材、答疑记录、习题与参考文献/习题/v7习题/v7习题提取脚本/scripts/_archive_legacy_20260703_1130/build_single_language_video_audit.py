from pathlib import Path
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def main() -> None:
    out = Path(sys.argv[1])
    review = out / "single_language_visual_review"
    book = out / "single_language_video_fidelity_audit.xlsx"

    rows = [
        [
            "中文",
            "CN101",
            "101",
            "建议你审核",
            "题目内容可从细帧恢复，但视频一开头就是本题，原2秒抽帧漏了题干上半段。",
            "中文版2.mp4 0.00-0.50s",
            str(review / "fine_cn" / "sheets" / "cn101_v2_01.jpg"),
            "请确认是否按这个视频开头细帧作为 CN101 的题干来源。",
        ],
        [
            "中文",
            "CN132",
            "131/132需确认",
            "建议你审核",
            "候选行编号与视频内题号有偏移：细帧显示对应内容是视频内 131/395，后面才进入 132。",
            "中文版2.mp4 169-171s",
            str(review / "contexts" / "cn_132_context.jpg"),
            "请确认最终题库编号按“视频内题号”还是按当前候选序号；内容本身可恢复。",
        ],
        [
            "英文",
            "EN284",
            "283/284需确认",
            "建议你审核",
            "上下文帧显示该段在 281、282、283、285 间切换；候选 EN284 的内容更像视频内 283/395。",
            "英文版3.mp4 512-516s",
            str(review / "contexts" / "en_284_context.jpg"),
            "请确认该题最终归属题号；题面内容可从上下文恢复。",
        ],
        [
            "英文",
            "EN308",
            "307/308需确认",
            "建议你审核",
            "该段跨了 306、307、308、309、310 多题；候选行混入了上一题解析和下一题选项，归属最容易错。",
            "英文版4.mp4 42-46s",
            str(review / "contexts" / "en_308_context.jpg"),
            "请人工确认这段到底应归到哪一道题，以及题干/答案边界。",
        ],
        [
            "中文",
            "CN168",
            "168",
            "可自动修",
            "2秒帧截到下半页/解析页，但细帧能看到完整题面。",
            "中文版2.mp4 353s附近",
            str(review / "fine_cn" / "sheets" / "cn168_01.jpg"),
            "我可按细帧重抽题干。",
        ],
        [
            "中文",
            "CN164",
            "164",
            "可自动修",
            "2秒帧只截到题干下半段，细帧能补全。",
            "中文版2.mp4 340s附近",
            str(review / "fine_cn" / "sheets" / "cn164_01.jpg"),
            "我可按细帧重抽题干。",
        ],
        [
            "中文",
            "CN378",
            "378",
            "可自动修",
            "2秒帧在下半页，细帧能补全。",
            "中文版4.mp4 433-435s",
            str(review / "fine_cn" / "sheets" / "cn378_01.jpg"),
            "我可按细帧重抽题干。",
        ],
        [
            "英文",
            "EN124",
            "124",
            "可自动修",
            "原帧跨页/混入中文解析，但邻近帧显示完整英文题面。",
            "英文版2.mp4 138-146s",
            str(review / "contexts" / "en_124_context.jpg"),
            "我可按邻近帧重抽题干。",
        ],
        [
            "英文",
            "EN221",
            "221",
            "可自动修",
            "原帧跨页/混入解析，但邻近帧显示完整英文题面。",
            "英文版3.mp4 108-112s",
            str(review / "contexts" / "en_221_context.jpg"),
            "我可按邻近帧重抽题干。",
        ],
        [
            "中文",
            "CN162",
            "162",
            "已通过",
            "题干、选项、答案在视频画面中可直接确认。",
            "中文版2.mp4 328-334s",
            str(review / "cn_review_01.jpg"),
            "无需人工审核。",
        ],
        [
            "中文",
            "CN187",
            "187",
            "已通过",
            "题干、选项、答案在视频画面中可直接确认。",
            "中文版2.mp4 446-448s",
            str(review / "cn_review_02.jpg"),
            "无需人工审核。",
        ],
        [
            "中文",
            "CN191",
            "191",
            "已通过",
            "题干、选项、答案在视频画面中可直接确认。",
            "中文版2.mp4 466s",
            str(review / "cn_review_02.jpg"),
            "无需人工审核。",
        ],
        [
            "中文",
            "CN212",
            "212",
            "已通过",
            "题干、选项、答案在视频画面中可直接确认。",
            "中文版3.mp4 60-62s",
            str(review / "cn_review_02.jpg"),
            "无需人工审核。",
        ],
        [
            "中文",
            "CN352",
            "352",
            "已通过",
            "题干、选项、答案在视频画面中可直接确认。",
            "中文版4.mp4 284-286s",
            str(review / "cn_review_02.jpg"),
            "无需人工审核。",
        ],
        [
            "中文",
            "CN389",
            "389",
            "已通过",
            "题干、选项、答案在视频画面中可直接确认。",
            "中文版4.mp4 494-496s",
            str(review / "cn_review_03.jpg"),
            "无需人工审核。",
        ],
        [
            "中文",
            "CN194",
            "194",
            "已通过",
            "题干、选项、答案在视频画面中可直接确认。",
            "中文版2.mp4 478-480s",
            str(review / "cn_review_03.jpg"),
            "无需人工审核。",
        ],
        [
            "英文",
            "EN21",
            "21",
            "已通过",
            "题干、选项、答案在视频画面中可直接确认。",
            "英文版1.mp4 140-142s",
            str(review / "en_review_01.jpg"),
            "无需人工审核。",
        ],
        [
            "英文",
            "EN189",
            "189",
            "已通过",
            "题干、选项、答案在视频画面中可直接确认。",
            "英文版2.mp4 548-550s",
            str(review / "en_review_02.jpg"),
            "无需人工审核。",
        ],
        [
            "英文",
            "EN351",
            "351",
            "已通过",
            "题干、选项、答案在视频画面中可直接确认。",
            "英文版4.mp4 280s",
            str(review / "en_review_02.jpg"),
            "无需人工审核。",
        ],
    ]

    headers = [
        "语言",
        "候选编号",
        "视频内题号判断",
        "审核等级",
        "复核结论",
        "视频时间",
        "证据截图",
        "处理建议",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "单语言复核总览"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    need = wb.create_sheet("需要你审核")
    need.append(headers)
    for row in rows:
        if row[3] == "建议你审核":
            need.append(row)

    summary = wb.create_sheet("汇总")
    summary.append(["项目", "数量"])
    summary.append(["候选总数", len(rows)])
    summary.append(["建议你审核", sum(1 for r in rows if r[3] == "建议你审核")])
    summary.append(["可自动修", sum(1 for r in rows if r[3] == "可自动修")])
    summary.append(["已通过", sum(1 for r in rows if r[3] == "已通过")])
    summary.append(
        [
            "口径",
            "只看中文题目是否符合中文视频、英文题目是否符合英文视频；未做中英语义对齐判断。",
        ]
    )

    fills = {
        "建议你审核": PatternFill("solid", fgColor="F4CCCC"),
        "可自动修": PatternFill("solid", fgColor="FFF2CC"),
        "已通过": PatternFill("solid", fgColor="D9EAD3"),
    }

    for sheet in [ws, need, summary]:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows(min_row=2):
            if sheet.title != "汇总" and row[3].value in fills:
                row[3].fill = fills[row[3].value]
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        if sheet.title != "汇总":
            for r in range(2, sheet.max_row + 1):
                cell = sheet.cell(r, 7)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
        widths = [12, 14, 18, 14, 54, 24, 72, 42] if sheet.title != "汇总" else [20, 80]
        for i, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width
        sheet.freeze_panes = "A2"

    wb.save(book)
    print(book)


if __name__ == "__main__":
    main()
