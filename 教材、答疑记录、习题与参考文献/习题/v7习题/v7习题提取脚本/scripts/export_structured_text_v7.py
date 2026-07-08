from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parents[1]
SOURCE = BASE / "CAMS_v7\u9898\u5e93_\u4e2d\u82f1\u5bf9\u7167_v8\u7cbe\u4fee\u7248.xlsx"
OUT_DIR = BASE / "v7\u7ed3\u6784\u5316\u6587\u672c"

SHEET_ALIGN = "\u4e2d\u82f1\u5bf9\u7167\u8868"
SHEET_RISK = "v8\u9700\u4eba\u5de5\u5ba1\u6838"
COL_CN_QNO = "\u4e2d\u6587\u9898\u53f7"
COL_EN_QNO = "\u82f1\u6587\u9898\u53f7"
COL_CN_QUESTION = "\u4e2d\u6587\u9898\u76ee"
COL_CN_ANSWER = "\u4e2d\u6587\u7b54\u6848"
COL_CN_EXPLANATION = "\u4e2d\u6587\u89e3\u6790"
COL_EN_QUESTION = "\u82f1\u6587\u9898\u76ee"
COL_EN_ANSWER = "\u82f1\u6587\u7b54\u6848"
COL_EN_EXPLANATION = "\u82f1\u6587\u89e3\u6790"


def compact(value: Any) -> str:
    return " ".join(str(value or "").replace("\r", "\n").split())


def headers(ws) -> dict[str, int]:
    return {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}


def answer_array(value: Any) -> list[str]:
    return re.findall(r"[A-G]", str(value or "").upper())


def detect_type(*texts: str) -> str:
    joined = " ".join(texts).lower()
    if "\u591a\u9009" in joined or "select" in joined or "choose two" in joined or "choose three" in joined:
        return "multiple"
    if "\u5355\u9009" in joined:
        return "single"
    return "unknown"


def extract_topic(text: str) -> str:
    match = re.search(r"-\[(.*?)\]", text)
    return compact(match.group(1)) if match else ""


def strip_prefix(text: str) -> str:
    text = compact(text)
    text = re.sub(r"^(?:\u7ffb\u8bd1|\u7ffb\u6cfd|\u7ffb\u8b6f)\s*", "", text)
    text = re.sub(r"^(?:\u5355\u9009|\u591a\u9009)\s*", "", text)
    text = re.sub(r"-\[[^\]]+\]\s*", "", text, count=1)
    text = re.sub(r"^(?:\u5355\u9009|\u591a\u9009)\s*", "", text)
    return compact(text)


def split_stem_options(text: str) -> tuple[str, list[dict[str, str]], str]:
    body = strip_prefix(text)
    marker = r"(?:(?<=^)|(?<=[\s\u3002\uff01\uff1f?;:\uff1b\uff1a]))([A-G])(?:[\.、\)]|\s+|(?=[\u4e00-\u9fff]))"
    matches = list(re.finditer(marker, body))
    if len(matches) < 2:
        return body, [], "option_parse_low_confidence"

    stem = compact(body[: matches[0].start()])
    options: list[dict[str, str]] = []
    for idx, match in enumerate(matches):
        key = match.group(1)
        start = match.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(body)
        text_part = compact(body[start:end].strip(" .\u3002\uff1b;:\uff1a\u3001)"))
        if text_part:
            options.append({"key": key, "text": text_part})

    by_key: dict[str, str] = {}
    for opt in options:
        if len(opt["text"]) > len(by_key.get(opt["key"], "")):
            by_key[opt["key"]] = opt["text"]
    options = [{"key": key, "text": by_key[key]} for key in sorted(by_key)]
    status = "ok" if len(options) >= 2 else "option_parse_low_confidence"
    return stem, options, status


def load_risks(wb) -> dict[int, list[dict[str, str]]]:
    risks: dict[int, list[dict[str, str]]] = defaultdict(list)
    if SHEET_RISK not in wb.sheetnames:
        return risks
    ws = wb[SHEET_RISK]
    h = headers(ws)
    for row in range(2, ws.max_row + 1):
        qno = ws.cell(row, h["\u9898\u53f7"]).value
        if not isinstance(qno, int):
            continue
        risks[qno].append(
            {
                "language": compact(ws.cell(row, h["\u8bed\u8a00"]).value),
                "issue": compact(ws.cell(row, h["\u95ee\u9898"]).value),
                "note": compact(ws.cell(row, h["\u8bf4\u660e"]).value),
            }
        )
    return risks


def build_records() -> list[dict[str, Any]]:
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb[SHEET_ALIGN]
    h = headers(ws)
    risks = load_risks(wb)
    records: list[dict[str, Any]] = []

    for row in range(2, ws.max_row + 1):
        cn_qno = ws.cell(row, h[COL_CN_QNO]).value
        en_qno = ws.cell(row, h[COL_EN_QNO]).value
        raw_cn = compact(ws.cell(row, h[COL_CN_QUESTION]).value)
        raw_en = compact(ws.cell(row, h[COL_EN_QUESTION]).value)
        ans_cn = answer_array(ws.cell(row, h[COL_CN_ANSWER]).value)
        ans_en = answer_array(ws.cell(row, h[COL_EN_ANSWER]).value)
        stem_cn, options_cn, parse_cn = split_stem_options(raw_cn)
        stem_en, options_en, parse_en = split_stem_options(raw_en)
        if ans_cn and parse_cn == "ok" and not set(ans_cn).issubset({opt["key"] for opt in options_cn}):
            parse_cn = "option_parse_low_confidence"
        if ans_en and parse_en == "ok" and not set(ans_en).issubset({opt["key"] for opt in options_en}):
            parse_en = "option_parse_low_confidence"
        qtype = detect_type(raw_cn, raw_en)
        answer_status = "\u4e00\u81f4" if ans_cn and ans_cn == ans_en else "\u4e0d\u4e00\u81f4" if ans_cn and ans_en else "\u5355\u4fa7\u8bc6\u522b"

        records.append(
            {
                "id": f"CAMS-V7-{int(cn_qno):04d}",
                "cn_qno": cn_qno,
                "en_qno": en_qno,
                "type": qtype,
                "topic_cn": extract_topic(raw_cn),
                "topic_en": extract_topic(raw_en),
                "stem_cn": stem_cn,
                "stem_en": stem_en,
                "options_cn": options_cn,
                "options_en": options_en,
                "answer_cn": ans_cn,
                "answer_en": ans_en,
                "answer_final": ans_cn if ans_cn else ans_en,
                "answer_status": answer_status,
                "explanation_cn": compact(ws.cell(row, h[COL_CN_EXPLANATION]).value),
                "explanation_en": compact(ws.cell(row, h[COL_EN_EXPLANATION]).value),
                "raw_question_cn": raw_cn,
                "raw_question_en": raw_en,
                "parse_status": {"cn": parse_cn, "en": parse_en},
                "risk_flags": risks.get(cn_qno, []),
                "source": {"workbook": SOURCE.name, "sheet": SHEET_ALIGN, "row": row},
            }
        )
    return records


def write_jsonl(records: list[dict[str, Any]]) -> None:
    with (OUT_DIR / "CAMS_v7_questions.jsonl").open("w", encoding="utf-8", newline="\n") as f:
        for item in records:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")


def write_json(records: list[dict[str, Any]]) -> None:
    (OUT_DIR / "CAMS_v7_questions.json").write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


def write_schema() -> None:
    schema = {
        "$schema": "https://json-schema.org/draft/2020-12/schema",
        "title": "CAMS v7 structured question",
        "type": "object",
        "required": ["id", "cn_qno", "en_qno", "type", "raw_question_cn", "raw_question_en", "answer_final"],
        "properties": {
            "id": {"type": "string"},
            "cn_qno": {"type": "integer"},
            "en_qno": {"type": "integer"},
            "type": {"enum": ["single", "multiple", "unknown"]},
            "topic_cn": {"type": "string"},
            "topic_en": {"type": "string"},
            "stem_cn": {"type": "string"},
            "stem_en": {"type": "string"},
            "options_cn": {"type": "array", "items": {"$ref": "#/$defs/option"}},
            "options_en": {"type": "array", "items": {"$ref": "#/$defs/option"}},
            "answer_cn": {"type": "array", "items": {"type": "string"}},
            "answer_en": {"type": "array", "items": {"type": "string"}},
            "answer_final": {"type": "array", "items": {"type": "string"}},
            "answer_status": {"type": "string"},
            "explanation_cn": {"type": "string"},
            "explanation_en": {"type": "string"},
            "raw_question_cn": {"type": "string"},
            "raw_question_en": {"type": "string"},
            "parse_status": {"type": "object"},
            "risk_flags": {"type": "array"},
            "source": {"type": "object"},
        },
        "$defs": {
            "option": {
                "type": "object",
                "required": ["key", "text"],
                "properties": {"key": {"type": "string"}, "text": {"type": "string"}},
            }
        },
    }
    (OUT_DIR / "CAMS_v7_questions.schema.json").write_text(json.dumps(schema, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(records: list[dict[str, Any]]) -> None:
    fields = [
        "id",
        "cn_qno",
        "en_qno",
        "type",
        "topic_cn",
        "topic_en",
        "stem_cn",
        "stem_en",
        "options_cn_json",
        "options_en_json",
        "answer_cn",
        "answer_en",
        "answer_final",
        "answer_status",
        "risk_flags_json",
        "raw_question_cn",
        "raw_question_en",
        "explanation_cn",
        "explanation_en",
    ]
    with (OUT_DIR / "CAMS_v7_questions.csv").open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for item in records:
            writer.writerow(
                {
                    "id": item["id"],
                    "cn_qno": item["cn_qno"],
                    "en_qno": item["en_qno"],
                    "type": item["type"],
                    "topic_cn": item["topic_cn"],
                    "topic_en": item["topic_en"],
                    "stem_cn": item["stem_cn"],
                    "stem_en": item["stem_en"],
                    "options_cn_json": json.dumps(item["options_cn"], ensure_ascii=False),
                    "options_en_json": json.dumps(item["options_en"], ensure_ascii=False),
                    "answer_cn": "".join(item["answer_cn"]),
                    "answer_en": "".join(item["answer_en"]),
                    "answer_final": "".join(item["answer_final"]),
                    "answer_status": item["answer_status"],
                    "risk_flags_json": json.dumps(item["risk_flags"], ensure_ascii=False),
                    "raw_question_cn": item["raw_question_cn"],
                    "raw_question_en": item["raw_question_en"],
                    "explanation_cn": item["explanation_cn"],
                    "explanation_en": item["explanation_en"],
                }
            )


def write_markdown(records: list[dict[str, Any]]) -> None:
    lines = ["# CAMS v7 structured question bank", ""]
    for item in records:
        lines.extend(
            [
                f"## {item['id']} / CN {item['cn_qno']} / EN {item['en_qno']}",
                "",
                f"- Type: `{item['type']}`",
                f"- Answer: `{''.join(item['answer_final'])}`",
                f"- Answer status: `{item['answer_status']}`",
                "",
                "### 中文",
                "",
                item["stem_cn"] or item["raw_question_cn"],
                "",
            ]
        )
        for opt in item["options_cn"]:
            lines.append(f"- {opt['key']}. {opt['text']}")
        lines.extend(["", item["explanation_cn"], "", "### English", "", item["stem_en"] or item["raw_question_en"], ""])
        for opt in item["options_en"]:
            lines.append(f"- {opt['key']}. {opt['text']}")
        lines.extend(["", item["explanation_en"], ""])
        if item["risk_flags"]:
            lines.extend(["### Risk Flags", ""])
            for flag in item["risk_flags"]:
                lines.append(f"- {flag['language']} / {flag['issue']}: {flag['note']}")
            lines.append("")
    (OUT_DIR / "CAMS_v7_questions.md").write_text("\n".join(lines), encoding="utf-8")


def validate(records: list[dict[str, Any]]) -> str:
    lines: list[str] = []
    lines.append(f"source: {SOURCE.name}")
    lines.append(f"records: {len(records)}")
    lines.append(f"qno_continuous: {[r['cn_qno'] for r in records] == list(range(1, 396))}")
    lines.append("")
    lines.append("type_counts:")
    for key, count in Counter(r["type"] for r in records).most_common():
        lines.append(f"  {key}: {count}")
    lines.append("")
    lines.append("answer_status_counts:")
    for key, count in Counter(r["answer_status"] for r in records).most_common():
        lines.append(f"  {key}: {count}")
    lines.append("")
    lines.append("parse_status_counts:")
    parse_counts = Counter()
    for item in records:
        parse_counts[f"cn:{item['parse_status']['cn']}"] += 1
        parse_counts[f"en:{item['parse_status']['en']}"] += 1
    for key, count in parse_counts.most_common():
        lines.append(f"  {key}: {count}")
    lines.append("")

    issues: list[str] = []
    low_confidence: list[str] = []
    for item in records:
        if not item["raw_question_cn"] or not item["raw_question_en"]:
            issues.append(f"{item['id']}: blank raw question")
        if not item["answer_final"]:
            issues.append(f"{item['id']}: blank answer_final")
        if item["type"] == "multiple" and len(item["answer_final"]) < 2:
            issues.append(f"{item['id']}: multiple choice with fewer than 2 answers")
        for lang in ("cn", "en"):
            if item["parse_status"][lang] != "ok":
                low_confidence.append(f"{item['id']}: {lang} option parse low confidence")
                continue
            opt_keys = {opt["key"] for opt in item[f"options_{lang}"]}
            missing = [ans for ans in item[f"answer_{lang}"] if ans not in opt_keys]
            if item[f"answer_{lang}"] and missing:
                issues.append(f"{item['id']}: answer_{lang} not found in parsed options: {''.join(missing)}")

    lines.append(f"validation_issues: {len(issues)}")
    lines.extend(issues[:300])
    if len(issues) > 300:
        lines.append(f"... truncated {len(issues) - 300} more issues")
    lines.append("")
    lines.append(f"parse_low_confidence: {len(low_confidence)}")
    lines.extend(low_confidence[:200])
    if len(low_confidence) > 200:
        lines.append(f"... truncated {len(low_confidence) - 200} more low-confidence parse notices")
    return "\n".join(lines) + "\n"


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    records = build_records()
    write_jsonl(records)
    write_json(records)
    write_schema()
    write_csv(records)
    write_markdown(records)
    report = validate(records)
    (OUT_DIR / "validation_report.txt").write_text(report, encoding="utf-8")
    print(f"written: {OUT_DIR}")
    print(f"records: {len(records)}")


if __name__ == "__main__":
    main()
