from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from extract_video_ocr_sample import (
    FrameOcr,
    OcrLine,
    cluster_answer,
    cluster_frames,
    cluster_page_no,
    clean_lines,
    detect_answer,
    detect_page_no,
    normalize_text,
    question_signature,
    run_ocr,
    similarity,
    write_excel,
)


VIDEO_RE = re.compile(r"(?P<lang>中文|英文)版(?P<part>\d+)")


def resolve_ffmpeg() -> Path:
    for name in ("ffmpeg", "ffmpeg.exe"):
        found = shutil.which(name)
        if found:
            return Path(found)

    local_appdata = os.environ.get("LOCALAPPDATA")
    if local_appdata:
        root = Path(local_appdata) / "Microsoft" / "WinGet" / "Packages"
        matches = sorted(root.rglob("ffmpeg.exe"))
        if matches:
            return matches[0]
    raise FileNotFoundError("ffmpeg.exe not found")


def parse_video_meta(video_path: Path) -> tuple[str, int]:
    match = VIDEO_RE.search(video_path.stem)
    if not match:
        raise ValueError(f"unrecognized video name: {video_path.name}")
    lang = "cn" if match.group("lang") == "中文" else "en"
    part = int(match.group("part"))
    return lang, part


def extract_frames(video_path: Path, frame_dir: Path, interval_sec: float) -> list[Path]:
    frame_dir.mkdir(parents=True, exist_ok=True)
    existing = sorted(frame_dir.glob("*.jpg"))
    if existing:
        return existing

    ffmpeg = resolve_ffmpeg()
    pattern = frame_dir / f"{video_path.stem}_%06d.jpg"
    fps = 1 / interval_sec
    subprocess.run(
        [
            str(ffmpeg),
            "-hide_banner",
            "-loglevel",
            "error",
            "-y",
            "-i",
            str(video_path),
            "-vf",
            f"fps={fps}",
            str(pattern),
        ],
        check=True,
    )
    return sorted(frame_dir.glob("*.jpg"))


def frame_time(index: int, interval_sec: float) -> int:
    return int(round((index - 1) * interval_sec))


def page_range_for_part(part: int) -> tuple[int, int]:
    if part <= 1:
        return 1, 100
    if part == 2:
        return 101, 200
    if part == 3:
        return 201, 300
    return 301, 395


def filter_frame_page(frame: FrameOcr, part: int) -> None:
    if not frame.page_no or not str(frame.page_no).isdigit():
        return
    start, end = page_range_for_part(part)
    page_no = int(frame.page_no)
    if start <= page_no <= end:
        return

    # The flag icon before the page label is sometimes OCR'd as a leading "1",
    # e.g. 41/395 -> 141/395 or 80/395 -> 180/395. Prefer correcting that
    # common error before dropping the page number.
    for offset in (100, 200, 300):
        corrected = page_no - offset
        if start <= corrected <= end:
            frame.page_no = str(corrected)
            return

    frame.page_no = ""


def frame_to_dict(frame: FrameOcr) -> dict:
    return {
        "index": frame.index,
        "time_sec": frame.time_sec,
        "image": frame.image,
        "page_no": frame.page_no,
        "answer": frame.answer,
        "has_explanation": frame.has_explanation,
        "clean_text": frame.clean_text,
        "raw_text": frame.text,
        "lines": [
            {
                "y": line.y,
                "x": line.x,
                "text": line.text,
                "score": line.score,
            }
            for line in frame.lines
        ],
    }


def frame_from_dict(data: dict) -> FrameOcr:
    image_path = Path(data["image"])
    if not image_path.is_absolute():
        image_path = Path.cwd() / image_path
    return FrameOcr(
        index=int(data["index"]),
        time_sec=int(data["time_sec"]),
        image=str(image_path.resolve()),
        page_no=data.get("page_no", ""),
        answer=data.get("answer", ""),
        has_explanation=bool(data.get("has_explanation", False)),
        text=data.get("raw_text", ""),
        clean_text=data.get("clean_text", ""),
        lines=[
            OcrLine(
                y=float(line["y"]),
                x=float(line["x"]),
                text=line["text"],
                score=float(line["score"]),
            )
            for line in data.get("lines", [])
        ],
    )


def refresh_frame_derived_fields(frame: FrameOcr) -> None:
    raw_text = "\n".join(line.text for line in frame.lines if line.score >= 0.58)
    cleaned = "\n".join(clean_lines(frame.lines))
    flat = normalize_text(raw_text)
    frame.text = raw_text
    frame.clean_text = normalize_text(cleaned)
    frame.page_no = detect_page_no(frame.lines)
    frame.answer = detect_answer(flat)
    frame.has_explanation = ("试题详解" in flat or "原解析" in flat)


def is_question_like_frame(frame: FrameOcr) -> bool:
    text = frame.clean_text
    if not text:
        return False
    compact = re.sub(r"\s+", "", text)
    if compact.startswith(("正确答案", "试题详解", "原解析")):
        return False
    if not any(
        token in text
        for token in (
            "单选",
            "多选",
            "Select",
            "Choose",
            "Which",
            "What",
            "When",
            "According",
            "-[",
        )
    ):
        return False
    return len(question_signature(text)) >= 24


def is_explanation_only_frame(frame: FrameOcr) -> bool:
    text = frame.clean_text
    compact = re.sub(r"\s+", "", text)
    if not compact:
        return True
    if is_question_like_frame(frame):
        return False
    for token in ("正确答案", "试题详解", "原解析"):
        pos = compact.find(token)
        if 0 <= pos <= 35:
            return True
    return False


def is_potential_question_start(frame: FrameOcr) -> bool:
    if is_question_like_frame(frame):
        return True
    if is_explanation_only_frame(frame):
        return False
    sig = question_signature(frame.clean_text)
    if len(sig) < 24:
        return False
    if frame.answer:
        return True
    return not frame.has_explanation


def infer_gap_pages(
    frames: list[FrameOcr],
    run_indices: list[int],
    missing_pages: list[int],
    left_idx: int | None,
    right_idx: int | None,
) -> None:
    if not run_indices or not missing_pages:
        return

    left_sig = question_signature(frames[left_idx].clean_text) if left_idx is not None else ""
    right_sig = question_signature(frames[right_idx].clean_text) if right_idx is not None else ""

    candidates: list[int] = []
    for idx in run_indices:
        frame = frames[idx]
        if not is_potential_question_start(frame):
            continue
        sig = question_signature(frame.clean_text)
        if right_sig and similarity(sig, right_sig) >= 0.45:
            continue
        if left_sig and similarity(sig, left_sig) >= 0.45:
            continue
        if candidates and similarity(sig, question_signature(frames[candidates[-1]].clean_text)) >= 0.45:
            continue
        candidates.append(idx)

    if len(candidates) < len(missing_pages):
        fallback = next(
            (idx for idx in run_indices if not is_explanation_only_frame(frames[idx])),
            run_indices[0],
        )
        if fallback not in candidates:
            candidates.insert(0, fallback)

    if len(candidates) < len(missing_pages):
        candidates = sorted(set(candidates))
        while len(candidates) < len(missing_pages):
            pos = len(candidates)
            approx = run_indices[min(len(run_indices) - 1, round(pos * len(run_indices) / len(missing_pages)))]
            if approx not in candidates:
                candidates.append(approx)
            else:
                break

    candidates = sorted(candidates)
    if len(candidates) < len(missing_pages):
        return

    selected = candidates[: len(missing_pages)]
    next_page_like = right_idx if right_idx is not None else run_indices[-1] + 1
    if right_sig and selected:
        for idx in run_indices:
            if idx <= selected[-1]:
                continue
            if is_potential_question_start(frames[idx]) and similarity(
                question_signature(frames[idx].clean_text),
                right_sig,
            ) >= 0.45:
                next_page_like = idx
                break

    for pos, page_no in enumerate(missing_pages):
        group_start = selected[pos]
        group_end = selected[pos + 1] if pos + 1 < len(selected) else next_page_like
        for idx in range(group_start, group_end):
            if not frames[idx].page_no:
                frames[idx].page_no = str(page_no)


def infer_missing_frame_pages(frames: list[FrameOcr], part: int) -> None:
    start_page, end_page = page_range_for_part(part)

    known_indices = [
        idx
        for idx, frame in enumerate(frames)
        if frame.page_no and str(frame.page_no).isdigit()
    ]
    if len(known_indices) < 2:
        return

    first_known_idx = known_indices[0]
    first_known_page = int(frames[first_known_idx].page_no)
    if first_known_page > start_page and first_known_idx > 0:
        infer_gap_pages(
            frames,
            [idx for idx in range(0, first_known_idx) if not frames[idx].page_no],
            list(range(start_page, first_known_page)),
            None,
            first_known_idx,
        )

    for left_idx, right_idx in zip(known_indices, known_indices[1:]):
        if right_idx <= left_idx + 1:
            continue
        left_page = int(frames[left_idx].page_no)
        right_page = int(frames[right_idx].page_no)
        if not (start_page <= left_page <= end_page and start_page <= right_page <= end_page):
            continue
        if right_page <= left_page + 1:
            continue

        missing_pages = list(range(left_page + 1, right_page))
        if not missing_pages:
            continue

        run_indices = [
            idx
            for idx in range(left_idx + 1, right_idx)
            if not frames[idx].page_no
        ]
        if not run_indices:
            continue

        infer_gap_pages(frames, run_indices, missing_pages, left_idx, right_idx)


def process_one_video(video_path: str, output_root: str, interval_sec: float) -> dict:
    video = Path(video_path)
    output = Path(output_root)
    lang, part = parse_video_meta(video)

    frame_dir = output / "cache" / f"frames_{str(interval_sec).replace('.', '_')}s" / video.stem
    frame_paths = extract_frames(video, frame_dir, interval_sec)
    ocr_cache_dir = output / "cache" / f"ocr_{str(interval_sec).replace('.', '_')}s" / video.stem
    ocr_cache_dir.mkdir(parents=True, exist_ok=True)

    from rapidocr_onnxruntime import RapidOCR

    ocr = None
    frames: list[FrameOcr] = []
    for idx, path in enumerate(frame_paths, start=1):
        cache_path = ocr_cache_dir / f"{idx:06d}.json"
        if cache_path.exists():
            frame = frame_from_dict(json.loads(cache_path.read_text(encoding="utf-8")))
            refresh_frame_derived_fields(frame)
        else:
            if ocr is None:
                ocr = RapidOCR()
            frame = run_ocr(ocr, path, idx)
            frame.time_sec = frame_time(idx, interval_sec)
            cache_path.write_text(json.dumps(frame_to_dict(frame), ensure_ascii=False), encoding="utf-8")
        filter_frame_page(frame, part)
        frames.append(frame)
        if idx == 1 or idx % 10 == 0 or idx == len(frame_paths):
            print(f"{video.name}: OCR {idx}/{len(frame_paths)}", flush=True)

    infer_missing_frame_pages(frames, part)
    clusters = cluster_frames(frames)

    video_out = output / "videos"
    json_out = output / "json"
    video_out.mkdir(parents=True, exist_ok=True)
    json_out.mkdir(parents=True, exist_ok=True)

    xlsx_path = video_out / f"{video.stem}.xlsx"
    write_excel(xlsx_path, frames, clusters)

    raw_json_path = json_out / f"{video.stem}_frames.json"
    raw_json_path.write_text(
        json.dumps(
            [frame_to_dict(f) for f in frames],
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    cluster_json_path = json_out / f"{video.stem}_clusters.json"
    cluster_json_path.write_text(
        json.dumps(
            {
                "video": video.name,
                "lang": lang,
                "part": part,
                "interval_sec": interval_sec,
                "clusters": [
                    {
                        "cluster_id": cluster.cluster_id,
                        "page_no": cluster_page_no(cluster),
                        "answer": cluster_answer(cluster),
                        "first_time": cluster.first_time,
                        "last_time": cluster.last_time,
                        "frame_count": len(cluster.frames),
                        "question_text": cluster.best_question_frame.clean_text,
                        "explanation_text": (
                            cluster.best_explanation_frame.clean_text
                            if any(frame.has_explanation for frame in cluster.frames)
                            else ""
                        ),
                        "question_image": cluster.best_question_frame.image,
                        "explanation_image": cluster.best_explanation_frame.image,
                    }
                    for cluster in clusters
                ],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    return {
        "video": video.name,
        "lang": lang,
        "part": part,
        "frame_count": len(frames),
        "cluster_count": len(clusters),
        "xlsx": str(xlsx_path),
        "raw_json": str(raw_json_path),
        "cluster_json": str(cluster_json_path),
    }


def pick_best_cluster(clusters: list[dict]) -> dict:
    def quality(item: dict) -> tuple[int, int, int]:
        text = str(item.get("question_text", "") or "")
        compact = re.sub(r"\s+", "", text)
        score = min(len(text), 1200) // 3
        if any(
            token in text
            for token in (
                "单选",
                "多选",
                "Which",
                "What",
                "When",
                "According",
                "Select",
                "Choose",
                "-[",
            )
        ):
            score += 600
        if any(token in text for token in ("正确答案", "试题详解", "原解析")):
            score -= 250
        if compact.startswith(("正确答案", "试题详解", "原解析")):
            score -= 1200
        option_hits = len(re.findall(r"(^|[\s,，。:：])([A-F])([\s\w\u4e00-\u9fff]|$)", text))
        if option_hits >= 2:
            score += option_hits * 60
        return (score, int(item.get("frame_count", 0)), len(text))

    return max(clusters, key=quality)


def write_merged_workbook(summary_paths: list[Path], output_path: Path) -> None:
    cn_rows_by_page: dict[str, dict] = {}
    en_rows_by_page: dict[str, dict] = {}

    for path in summary_paths:
        data = json.loads(path.read_text(encoding="utf-8"))
        lang = data["lang"]
        part = int(data["part"])
        clusters = data["clusters"]

        by_page: dict[str, list[dict]] = defaultdict(list)
        for cluster in clusters:
            page_no = str(cluster.get("page_no", "")).strip()
            if not page_no:
                continue
            by_page[page_no].append(cluster)

        for page_no, entries in by_page.items():
            chosen = pick_best_cluster(entries)
            row = {
                "page_no": int(page_no) if page_no.isdigit() else page_no,
                "part": part,
                "answer": chosen.get("answer", ""),
                "first_time": chosen.get("first_time", ""),
                "last_time": chosen.get("last_time", ""),
                "frame_count": chosen.get("frame_count", 0),
                "question_text": chosen.get("question_text", ""),
                "explanation_text": chosen.get("explanation_text", ""),
                "question_image": chosen.get("question_image", ""),
                "explanation_image": chosen.get("explanation_image", ""),
                "video": data["video"],
            }
            target = cn_rows_by_page if lang == "cn" else en_rows_by_page
            existing = target.get(page_no)
            if existing is None or (
                row["frame_count"],
                len(row["question_text"]),
            ) > (
                existing["frame_count"],
                len(existing["question_text"]),
            ):
                target[page_no] = row

    cn_rows = sorted(cn_rows_by_page.values(), key=lambda row: row["page_no"] if isinstance(row["page_no"], int) else 9999)
    en_rows = sorted(en_rows_by_page.values(), key=lambda row: row["page_no"] if isinstance(row["page_no"], int) else 9999)

    cn_by_page = {str(row["page_no"]): row for row in cn_rows}
    en_by_page = {str(row["page_no"]): row for row in en_rows}
    all_pages = sorted(
        set(cn_by_page) | set(en_by_page),
        key=lambda value: int(value) if value.isdigit() else 9999,
    )

    wb = Workbook()
    ws_cn = wb.active
    ws_cn.title = "中文合并"
    ws_en = wb.create_sheet("英文合并")
    ws_merged = wb.create_sheet("中英对照")

    def style_header(ws):
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def set_widths(ws, widths: dict[str, int]):
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    ws_cn.append(
        [
            "题号",
            "分卷",
            "答案",
            "起止秒",
            "帧数",
            "题目候选文本",
            "解析候选文本",
            "来源视频",
            "题目截图",
            "解析截图",
        ]
    )
    for row in cn_rows:
        ws_cn.append(
            [
                row["page_no"],
                row["part"],
                row["answer"],
                f'{row["first_time"]}-{row["last_time"]}',
                row["frame_count"],
                row["question_text"],
                row["explanation_text"],
                row["video"],
                row["question_image"],
                row["explanation_image"],
            ]
        )

    ws_en.append(
        [
            "Question",
            "Part",
            "Answer",
            "Time",
            "Frames",
            "Question Text",
            "Explanation Text",
            "Source",
            "Question Image",
            "Explanation Image",
        ]
    )
    for row in en_rows:
        ws_en.append(
            [
                row["page_no"],
                row["part"],
                row["answer"],
                f'{row["first_time"]}-{row["last_time"]}',
                row["frame_count"],
                row["question_text"],
                row["explanation_text"],
                row["video"],
                row["question_image"],
                row["explanation_image"],
            ]
        )

    ws_merged.append(
        [
            "题号",
            "中文分卷",
            "英文分卷",
            "推荐答案",
            "答案核对状态",
            "中文答案",
            "英文答案",
            "中文题目候选文本",
            "英文题目候选文本",
            "中文解析候选文本",
            "英文解析候选文本",
            "中文时间",
            "英文时间",
            "中文来源",
            "英文来源",
            "中文题目截图",
            "英文题目截图",
            "中文解析截图",
            "英文解析截图",
        ]
    )
    for page in all_pages:
        cn = cn_by_page.get(page, {})
        en = en_by_page.get(page, {})
        cn_answer = cn.get("answer", "")
        en_answer = en.get("answer", "")
        if cn_answer and en_answer:
            if cn_answer == en_answer:
                suggested_answer = cn_answer
                answer_status = "一致"
            else:
                suggested_answer = ""
                answer_status = f"不一致: 中文={cn_answer}; 英文={en_answer}"
        elif cn_answer or en_answer:
            suggested_answer = cn_answer or en_answer
            answer_status = "单侧识别"
        else:
            suggested_answer = ""
            answer_status = "未识别"
        ws_merged.append(
            [
                int(page) if page.isdigit() else page,
                cn.get("part", ""),
                en.get("part", ""),
                suggested_answer,
                answer_status,
                cn_answer,
                en_answer,
                cn.get("question_text", ""),
                en.get("question_text", ""),
                cn.get("explanation_text", ""),
                en.get("explanation_text", ""),
                f'{cn.get("first_time", "")}-{cn.get("last_time", "")}' if cn else "",
                f'{en.get("first_time", "")}-{en.get("last_time", "")}' if en else "",
                cn.get("video", ""),
                en.get("video", ""),
                cn.get("question_image", ""),
                en.get("question_image", ""),
                cn.get("explanation_image", ""),
                en.get("explanation_image", ""),
            ]
        )

    for ws in (ws_cn, ws_en, ws_merged):
        style_header(ws)

    set_widths(
        ws_cn,
        {
            "A": 8,
            "B": 8,
            "C": 8,
            "D": 14,
            "E": 8,
            "F": 72,
            "G": 72,
            "H": 18,
            "I": 55,
            "J": 55,
        },
    )
    set_widths(
        ws_en,
        {
            "A": 10,
            "B": 8,
            "C": 8,
            "D": 14,
            "E": 8,
            "F": 72,
            "G": 72,
            "H": 18,
            "I": 55,
            "J": 55,
        },
    )
    set_widths(
        ws_merged,
        {
            "A": 8,
            "B": 10,
            "C": 10,
            "D": 10,
            "E": 24,
            "F": 8,
            "G": 8,
            "H": 55,
            "I": 55,
            "J": 55,
            "K": 55,
            "L": 14,
            "M": 14,
            "N": 16,
            "O": 16,
            "P": 55,
            "Q": 55,
            "R": 55,
            "S": 55,
        },
    )

    for ws in (ws_cn, ws_en, ws_merged):
        for idx in range(1, ws.max_row + 1):
            ws.row_dimensions[idx].height = 60 if idx > 1 else 24

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--videos-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--interval-sec", type=float, default=5.0)
    parser.add_argument("--workers", type=int, default=1)
    parser.add_argument("--only", nargs="*", default=[])
    args = parser.parse_args()

    videos = sorted(args.videos_dir.glob("*.mp4"))
    if args.only:
        wanted = set(args.only)
        videos = [video for video in videos if video.name in wanted or video.stem in wanted]

    if not videos:
        raise SystemExit("no videos found")

    args.output_dir.mkdir(parents=True, exist_ok=True)

    summaries: list[Path] = []
    if args.workers <= 1:
        for video in videos:
            print(f"start {video.name}")
            result = process_one_video(str(video), str(args.output_dir), args.interval_sec)
            summaries.append(Path(result["cluster_json"]))
            print(
                f'{result["video"]}: frames={result["frame_count"]} clusters={result["cluster_count"]} '
                f'xlsx={result["xlsx"]}'
            )
    else:
        tasks = []
        with ProcessPoolExecutor(max_workers=max(1, args.workers)) as pool:
            for video in videos:
                tasks.append(pool.submit(process_one_video, str(video), str(args.output_dir), args.interval_sec))

            for future in as_completed(tasks):
                result = future.result()
                summaries.append(Path(result["cluster_json"]))
                print(
                    f'{result["video"]}: frames={result["frame_count"]} clusters={result["cluster_count"]} '
                    f'xlsx={result["xlsx"]}'
                )

    merged_path = args.output_dir / "merged_cn_en.xlsx"
    write_merged_workbook(sorted(summaries), merged_path)
    print(f"merged={merged_path}")


if __name__ == "__main__":
    main()
