from __future__ import annotations

import csv
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent
WORK_DIR = ROOT / "英文题干选项逐题审核_工作区"
INPUT_DIR = WORK_DIR / "inputs"
AGENT_DIR = WORK_DIR / "agent_outputs"
OUTPUT_PATH = ROOT / "英文题干选项逐题审核表.xlsx"
VALIDATION_PATH = WORK_DIR / "validation_report.json"

CHAPTERS = ["第二章", "第三章", "第四章", "第五章", "第六章"]
EXPECTED_COUNTS = {"第二章": 177, "第三章": 201, "第四章": 210, "第五章": 111, "第六章": 16}
OPTION_LABELS = list("ABCDEF")
ALLOWED_CONFIDENCE = {"高", "中", "低"}
ALLOWED_STATUS = {"已确认", "仅语言润色", "待人工确认"}
EXPECTED_CROSS_CHAPTER_MAPPINGS = {
    ("第三章", 15, "第五章.docx", 96),
    ("第三章", 19, "第五章.docx", 107),
    ("第三章", 20, "第五章.docx", 94),
    ("第二章", 151, "第三章.docx", 66),
    ("第五章", 105, "第三章.docx", 185),
    ("第六章", 1, "第二章.docx", 100),
    ("第四章", 98, "第三章.docx", 29),
    ("第四章", 101, "第三章.docx", 126),
}
EXPECTED_UNMATCHED = {("第三章", 113)}


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def read_agent_csv(path: Path, columns: list[str]) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != columns:
            raise ValueError(f"{path.name}: unexpected columns {reader.fieldnames}")
        return [{key: (value or "").strip() for key, value in row.items()} for row in reader]


def parse_option_mapping(text: str) -> dict[str, str]:
    return {
        english.upper(): chinese.upper()
        for english, chinese in re.findall(r"([A-F])\s*(?:→|->|=)\s*([A-F])", text.upper())
    }


def normalized_answer(text: str) -> str:
    return "".join(re.findall(r"[A-F]", text.upper()))


def requested_answer_count(question: str) -> int | None:
    match = re.search(
        r"\b(?:choose|select)\s+(one|two|three|four|five|six|\d+)\b",
        question.lower(),
    )
    if not match:
        return None
    words = {"one": 1, "two": 2, "three": 3, "four": 4, "five": 5, "six": 6}
    token = match.group(1)
    return words.get(token, int(token) if token.isdigit() else None)


def validate_chapter(
    chapter: str,
    inputs: list[dict],
    outputs: list[dict[str, str]],
    chinese_index: dict[tuple[str, int], dict],
) -> list[str]:
    errors: list[str] = []
    expected = EXPECTED_COUNTS[chapter]
    if len(inputs) != expected or len(outputs) != expected:
        errors.append(f"{chapter}: expected {expected}, input={len(inputs)}, output={len(outputs)}")
        return errors

    seen = set()
    input_by_number = {record["english"]["number"]: record for record in inputs}
    for row in outputs:
        try:
            number = int(row["current_question_number"])
        except ValueError:
            errors.append(f"{chapter}: invalid question number {row['current_question_number']!r}")
            continue
        if number in seen:
            errors.append(f"{chapter} Q{number}: duplicate output row")
            continue
        seen.add(number)
        source = input_by_number.get(number)
        if source is None:
            errors.append(f"{chapter} Q{number}: not found in input")
            continue
        english = source["english"]
        english_labels = set(english["options"])

        if row["mapping_confidence"] not in ALLOWED_CONFIDENCE:
            errors.append(f"{chapter} Q{number}: invalid confidence {row['mapping_confidence']!r}")
        if row["mapping_status"] not in ALLOWED_STATUS:
            errors.append(f"{chapter} Q{number}: invalid status {row['mapping_status']!r}")
        if normalized_answer(row["current_answer"]) != english["answer"]:
            errors.append(f"{chapter} Q{number}: current answer changed")
        if not row["proposed_question"]:
            errors.append(f"{chapter} Q{number}: missing proposed question")
        for label in OPTION_LABELS:
            proposed = row[f"proposed_{label}"]
            if label in english_labels and not proposed:
                errors.append(f"{chapter} Q{number}: missing proposed option {label}")
            if label not in english_labels and proposed:
                errors.append(f"{chapter} Q{number}: unexpected proposed option {label}")
        proposed_answer = normalized_answer(row["proposed_answer"])
        if not proposed_answer or any(label not in english_labels for label in proposed_answer):
            errors.append(f"{chapter} Q{number}: invalid proposed answer {row['proposed_answer']!r}")
        if len(set(proposed_answer)) != len(proposed_answer):
            errors.append(f"{chapter} Q{number}: proposed answer contains duplicate labels")
        requested_count = requested_answer_count(row["proposed_question"])
        if requested_count is not None and len(proposed_answer) != requested_count:
            errors.append(
                f"{chapter} Q{number}: proposed answer count {len(proposed_answer)} != requested {requested_count}"
            )

        mapping_file = row["mapping_source_file"]
        try:
            mapping_number = int(row["mapping_source_number"]) if row["mapping_source_number"] else None
        except ValueError:
            mapping_number = None
            errors.append(f"{chapter} Q{number}: invalid mapping source number")
        chinese = chinese_index.get((mapping_file, mapping_number)) if mapping_number else None
        if chinese:
            if normalized_answer(row["chinese_answer"]) != chinese["answer"]:
                errors.append(f"{chapter} Q{number}: Chinese answer does not match source")
        elif row["chinese_answer"]:
            errors.append(f"{chapter} Q{number}: Chinese answer supplied without valid source")

        if row["mapping_confidence"] == "高":
            if chinese is None:
                errors.append(f"{chapter} Q{number}: high confidence without valid Chinese source")
                continue
            option_map = parse_option_mapping(row["option_mapping"])
            chinese_labels = set(chinese["options"])
            if set(option_map) != english_labels:
                errors.append(f"{chapter} Q{number}: incomplete English option mapping")
            if set(option_map.values()) != chinese_labels or len(option_map.values()) != len(set(option_map.values())):
                errors.append(f"{chapter} Q{number}: mapping is not a Chinese-option bijection")
            derived = "".join(
                label
                for label in OPTION_LABELS
                if label in option_map and option_map[label] in set(chinese["answer"])
            )
            if proposed_answer != derived:
                errors.append(
                    f"{chapter} Q{number}: proposed answer {proposed_answer} != derived {derived}"
                )
        else:
            if proposed_answer != english["answer"]:
                errors.append(f"{chapter} Q{number}: non-high-confidence answer must remain current")

    if seen != set(range(1, expected + 1)):
        missing = sorted(set(range(1, expected + 1)) - seen)
        errors.append(f"{chapter}: missing question numbers {missing[:20]}")
    return errors


def style_sheet(sheet, widths: list[int], freeze: str = "A2") -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions


def main() -> None:
    agent_columns = read_json(INPUT_DIR / "agent_columns.json")
    terminology = read_json(INPUT_DIR / "terminology.json")
    chinese_rows = read_json(INPUT_DIR / "中文题库_全部.json")
    chinese_index = {(row["file_name"], int(row["number"])): row for row in chinese_rows}

    chapter_inputs = {chapter: read_json(INPUT_DIR / f"{chapter}.json") for chapter in CHAPTERS}
    chapter_outputs = {
        chapter: read_agent_csv(AGENT_DIR / f"{chapter}.csv", agent_columns) for chapter in CHAPTERS
    }

    errors = []
    cross_mappings = set()
    unmatched = set()
    for chapter in CHAPTERS:
        errors.extend(
            validate_chapter(chapter, chapter_inputs[chapter], chapter_outputs[chapter], chinese_index)
        )
        input_by_number = {
            int(record["english"]["number"]): record for record in chapter_inputs[chapter]
        }
        for row in chapter_outputs[chapter]:
            number = int(row["current_question_number"])
            english_file = input_by_number[number]["english"]["file_name"]
            source_file = row["mapping_source_file"]
            source_number = int(row["mapping_source_number"]) if row["mapping_source_number"] else None
            if not source_file:
                unmatched.add((chapter, number))
            elif source_file != english_file.removeprefix("英文版"):
                cross_mappings.add((chapter, number, source_file, source_number))
    if cross_mappings != EXPECTED_CROSS_CHAPTER_MAPPINGS:
        errors.append(
            "cross-chapter mapping set mismatch: "
            f"expected={sorted(EXPECTED_CROSS_CHAPTER_MAPPINGS)}, actual={sorted(cross_mappings)}"
        )
    if unmatched != EXPECTED_UNMATCHED:
        errors.append(
            f"unmatched-question set mismatch: expected={sorted(EXPECTED_UNMATCHED)}, actual={sorted(unmatched)}"
        )
    validation = {
        "valid": not errors,
        "errors": errors,
        "counts": {chapter: len(chapter_outputs[chapter]) for chapter in CHAPTERS},
        "total": sum(len(rows) for rows in chapter_outputs.values()),
        "cross_chapter_mappings": sorted(cross_mappings),
        "unmatched_questions": sorted(unmatched),
    }
    VALIDATION_PATH.write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")
    if errors:
        raise ValueError("Validation failed:\n" + "\n".join(errors[:100]))

    workbook = Workbook()
    summary = workbook.active
    summary.title = "总览"
    summary.append(
        [
            "章节",
            "题数",
            "高置信",
            "中置信",
            "低置信",
            "建议答案变化",
            "有风险标记",
            "待人工确认",
        ]
    )

    audit = workbook.create_sheet("映射审计")
    audit_headers = [
        "英文章节",
        "英文文档",
        "当前题号",
        "原英文题干",
        "当前答案",
        "中文来源文档",
        "中文题号",
        "中文题干",
        "中文答案",
        "映射置信度",
        "映射状态",
        "选项对应关系",
        "映射说明",
    ]
    audit.append(audit_headers)

    total_counters = Counter()
    chapter_headers = (
        ["当前题号", "原英文题干"]
        + [f"原英文{label}" for label in OPTION_LABELS]
        + ["中文来源文档", "中文题号", "中文题干"]
        + [f"中文{label}" for label in OPTION_LABELS]
        + ["建议英文题干"]
        + [f"建议英文{label}" for label in OPTION_LABELS]
        + [
            "当前答案",
            "中文答案",
            "建议答案",
            "映射置信度",
            "映射状态",
            "选项对应关系",
            "修改类型",
            "答案/解析风险",
            "代理说明",
            "人工审核状态",
            "审核备注",
        ]
    )

    for chapter in CHAPTERS:
        inputs_by_number = {
            int(record["english"]["number"]): record for record in chapter_inputs[chapter]
        }
        outputs = sorted(chapter_outputs[chapter], key=lambda row: int(row["current_question_number"]))
        sheet = workbook.create_sheet(chapter)
        sheet.append(chapter_headers)
        counters = Counter()

        for row in outputs:
            number = int(row["current_question_number"])
            source = inputs_by_number[number]
            english = source["english"]
            source_number = int(row["mapping_source_number"]) if row["mapping_source_number"] else None
            chinese = chinese_index.get((row["mapping_source_file"], source_number)) if source_number else None
            chinese_options = chinese["options"] if chinese else {}

            audit.append(
                [
                    chapter,
                    english["file_name"],
                    number,
                    english["question"],
                    english["answer"],
                    row["mapping_source_file"],
                    source_number or "",
                    chinese["question"] if chinese else "",
                    chinese["answer"] if chinese else "",
                    row["mapping_confidence"],
                    row["mapping_status"],
                    row["option_mapping"],
                    row["mapping_reason"],
                ]
            )

            sheet.append(
                [number, english["question"]]
                + [english["options"].get(label, "") for label in OPTION_LABELS]
                + [
                    row["mapping_source_file"],
                    source_number or "",
                    chinese["question"] if chinese else "",
                ]
                + [chinese_options.get(label, "") for label in OPTION_LABELS]
                + [row["proposed_question"]]
                + [row[f"proposed_{label}"] for label in OPTION_LABELS]
                + [
                    english["answer"],
                    chinese["answer"] if chinese else "",
                    normalized_answer(row["proposed_answer"]),
                    row["mapping_confidence"],
                    row["mapping_status"],
                    row["option_mapping"],
                    row["change_type"],
                    row["answer_analysis_risk"],
                    row["agent_notes"],
                    "待审核",
                    "",
                ]
            )

            counters[row["mapping_confidence"]] += 1
            if normalized_answer(row["proposed_answer"]) != english["answer"]:
                counters["答案变化"] += 1
            if row["answer_analysis_risk"] and row["answer_analysis_risk"] not in {"无", "无风险"}:
                counters["风险"] += 1
            if row["mapping_status"] == "待人工确认":
                counters["待确认"] += 1

        summary.append(
            [
                chapter,
                len(outputs),
                counters["高"],
                counters["中"],
                counters["低"],
                counters["答案变化"],
                counters["风险"],
                counters["待确认"],
            ]
        )
        total_counters.update(counters)

        status_column = len(chapter_headers) - 1
        validation = DataValidation(
            type="list", formula1='"待审核,已通过,需修改,暂缓"', allow_blank=False
        )
        sheet.add_data_validation(validation)
        validation.add(f"{sheet.cell(2, status_column).coordinate}:{sheet.cell(sheet.max_row, status_column).coordinate}")
        style_sheet(
            sheet,
            [9, 55] + [40] * 6 + [20, 10, 55] + [40] * 6 + [55] + [40] * 6
            + [10, 10, 10, 10, 14, 22, 16, 28, 48, 14, 30],
        )
        risk_column = len(chapter_headers) - 3
        risk_letter = sheet.cell(1, risk_column).column_letter
        sheet.conditional_formatting.add(
            f"{risk_letter}2:{risk_letter}{sheet.max_row}",
            FormulaRule(
                formula=[f'AND(${risk_letter}2<>"",${risk_letter}2<>"无",${risk_letter}2<>"无风险")'],
                fill=PatternFill("solid", fgColor="FCE4D6"),
            ),
        )

    summary.append(
        [
            "合计",
            715,
            total_counters["高"],
            total_counters["中"],
            total_counters["低"],
            total_counters["答案变化"],
            total_counters["风险"],
            total_counters["待确认"],
        ]
    )
    terminology_sheet = workbook.create_sheet("术语表")
    terminology_sheet.append(["中文术语", "统一英文"])
    for chinese, english in terminology.items():
        terminology_sheet.append([chinese, english])

    style_sheet(summary, [14, 10, 10, 10, 10, 14, 14, 14])
    style_sheet(audit, [12, 22, 10, 55, 10, 22, 10, 55, 10, 10, 14, 28, 50])
    style_sheet(terminology_sheet, [28, 45])
    workbook.save(OUTPUT_PATH)
    print(json.dumps(validation_report(chapter_outputs), ensure_ascii=False))
    print(OUTPUT_PATH)


def validation_report(chapter_outputs: dict[str, list[dict[str, str]]]) -> dict:
    result = {"total": 0, "chapters": {}}
    for chapter, rows in chapter_outputs.items():
        confidence = Counter(row["mapping_confidence"] for row in rows)
        status = Counter(row["mapping_status"] for row in rows)
        result["chapters"][chapter] = {
            "count": len(rows),
            "confidence": dict(confidence),
            "status": dict(status),
        }
        result["total"] += len(rows)
    return result


if __name__ == "__main__":
    main()
