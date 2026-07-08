from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(r"D:\守正公司工作区\cams考试\教材、答疑记录、习题与参考文献\习题\v7习题")
INPUT = BASE / "output_2s" / "semantic_aligned_cn_en_reviewed_v7_readability_fixed.xlsx"
OUTPUT = BASE / "output_2s" / "CAMS_v7题库_中英对照_v8精修版.xlsx"
AUDIT_DIR = BASE / "output_2s" / "v8_ocr_audit"

CN_SHEET = "中文题库"
EN_SHEET = "英文题库"
ALIGN_SHEET = "语义对齐"


def compact(value: Any, limit: int | None = None) -> str:
    text = " ".join(str(value or "").split())
    return text if limit is None else text[:limit]


def headers(ws) -> dict[str, int]:
    return {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}


def row_by_qno(ws, qno: int) -> int:
    h = headers(ws)
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, h["题号"]).value == qno:
            return row
    raise KeyError(f"{ws.title} Q{qno} not found")


def get_cell(ws, qno: int, field: str) -> Any:
    h = headers(ws)
    return ws.cell(row_by_qno(ws, qno), h[field]).value


def set_cell(ws, qno: int, field: str, value: Any) -> None:
    h = headers(ws)
    cell = ws.cell(row_by_qno(ws, qno), h[field])
    cell.value = value
    if field.endswith("截图") and isinstance(value, str) and value.lower().endswith((".jpg", ".png")):
        cell.hyperlink = value
        cell.style = "Hyperlink"


def norm_answer(value: Any) -> str:
    return "".join(re.findall(r"[A-G]", str(value or "").upper()))


def normalize_spacing(text: str) -> str:
    text = str(text or "").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = text.replace("**", "")
    replacements = {
        "充许": "允许",
        "全融": "金融",
        "冬约": "条约",
        "Mutual|eaal": "Mutual Legal",
        "MutualLegal": "Mutual Legal",
        "AssistanceTreaty": "Assistance Treaty",
        "特别行政区报告": "可疑活动报告(SAR)",
        "特别行政区 报告": "可疑活动报告(SAR)",
        "严重事故报告": "可疑活动报告(SAR)",
        "严重事 故报告": "可疑活动报告(SAR)",
        "所L以": "所以",
        "所L 以": "所以",
        "百解析": "原解析",
        "选政": "选项",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"\bMutual\s+Legal\s*Assistance\s*Treaty\b", "Mutual Legal Assistance Treaty", text)
    return text.strip()


QUESTION_PREFIX_PATTERNS = [
    r"^\s*[活》\s]*活\s*",
    r"^\s*(?:中\s*)?[O0]?\s*(?:[A-G]\s*)?(?:翻译|翻泽|翻譯)\s*",
    r"^\s*(?:中\s*)?[O0]?\s*(?:[A-G]\s*)?折题\s*",
    r"^\s*(?:折题|新)\s*",
    r"^\s*[A-G]\s*(?=(?:单选|多选)\b|-?\[)",
    r"^\s*[O0]\s+(?=(?:单选|多选)\b|-?\[)",
]


def strip_question_noise(text: Any) -> str:
    text = normalize_spacing(str(text or ""))
    for pattern in QUESTION_PREFIX_PATTERNS:
        text = re.sub(pattern, "", text)
    # Feedback and explanation banners belong outside the question text.
    text = re.split(r"\s*正确答[案室]\s*[:：]?\s*[A-G]*.*", text, maxsplit=1)[0]
    text = re.split(r"\s*[A-G]*\s*您选择\s*[A-G/]*.*", text, maxsplit=1)[0]
    text = re.split(r"\s*试题(?:详|送)解.*", text, maxsplit=1)[0]
    text = re.sub(r"\s+解析:\s*", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    # Common OCR layout issue: option letters drift after the option text.
    text = re.sub(r"(?<![A-Z])\s([A-G])(?=\S)", r" \1 ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def strip_analysis_noise(text: Any) -> str:
    text = normalize_spacing(str(text or ""))
    text = re.sub(r"^\s*[活》\s]*活\s*", "", text)
    text = re.sub(r"^\s*(?:折题|新)\s*", "", text)
    text = text.replace("心心干", "")
    text = text.replace("心干", "")
    text = re.sub(r"^\s*试题(?:详|送)解\s*", "", text)
    text = re.sub(r"^\s*原解析\s*", "", text)
    text = re.sub(r".*?试题(?:详|送)解\s*原解析\s*", "", text, count=1)
    text = re.sub(r".*?正确答[案室]\s*[:：]?\s*[A-G]+(?:\s*您选择\s*[A-G/]*)?\s*试题(?:详|送)解\s*原解析\s*", "", text, count=1)
    text = re.sub(r"^\s*正确答[案室]\s*[:：]?\s*[A-G]+(?:\s*您选择\s*[A-G/]*)?\s*", "", text)
    text = re.sub(r"^\s*[A-G]*\s*您选择\s*[A-G/]*\s*", "", text)
    text = re.sub(r"\s*难\s*度.*$", "", text)
    text = re.sub(r"\s*2026新(?:真|直)题-v7-(?:中文|英文).*$", "", text)
    text = re.sub(r"\s*来\s*源\s*考友笔记.*$", "", text)
    text = re.sub(r"\s*考友笔记.*$", "", text)
    text = re.sub(r"\s*公享华记新市同学快速解额.*$", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def read_audit_rows() -> list[dict[str, Any]]:
    paths = sorted(AUDIT_DIR.glob("report_cn*.jsonl"))
    if (AUDIT_DIR / "report_cn101_200_rerun.jsonl").exists():
        paths = [p for p in paths if p.name != "report_cn101_200.jsonl"]
    rows: list[dict[str, Any]] = []
    for path in paths:
        for line_no, line in enumerate(path.read_text(encoding="utf-8-sig").splitlines(), 1):
            line = line.strip().lstrip("\ufeff")
            if not line:
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                continue
            if compact(item.get("issue_type")) == "qno_misalignment":
                continue
            item["_source_file"] = path.name
            item["_line_no"] = line_no
            rows.append(item)
    return rows


def write_sheet(wb, title: str, sheet_headers: list[str], rows: list[list[Any]], widths: list[int]) -> None:
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(sheet_headers)
    for row in rows:
        ws.append(row)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def update_alignment(wb) -> None:
    if ALIGN_SHEET not in wb.sheetnames:
        return
    align = wb[ALIGN_SHEET]
    ah = headers(align)
    cn = wb[CN_SHEET]
    en = wb[EN_SHEET]
    for row in range(2, align.max_row + 1):
        cn_qno = align.cell(row, ah["中文题号"]).value
        en_qno = align.cell(row, ah["英文题号"]).value
        if not isinstance(cn_qno, int) or not isinstance(en_qno, int):
            continue
        cn_answer = get_cell(cn, cn_qno, "答案")
        en_answer = get_cell(en, en_qno, "答案")
        cn_norm = norm_answer(cn_answer)
        en_norm = norm_answer(en_answer)
        if cn_norm and en_norm and cn_norm == en_norm:
            recommended = cn_norm
            status = "一致"
        elif cn_norm and en_norm:
            recommended = ""
            status = f"不一致：中文={cn_norm}; 英文={en_norm}"
        else:
            recommended = cn_norm or en_norm
            status = "单侧识别"
        updates = {
            "推荐答案": recommended,
            "答案状态": status,
            "中文答案": cn_answer,
            "英文答案": en_answer,
            "中文题目": get_cell(cn, cn_qno, "题目文本"),
            "英文题目": get_cell(en, en_qno, "题目文本"),
            "中文解析": get_cell(cn, cn_qno, "解析文本"),
            "英文解析": get_cell(en, en_qno, "解析文本"),
            "中文题目截图": get_cell(cn, cn_qno, "题目截图"),
            "英文题目截图": get_cell(en, en_qno, "题目截图"),
        }
        for key, value in updates.items():
            if key not in ah:
                continue
            cell = align.cell(row, ah[key])
            cell.value = value
            if key.endswith("截图") and isinstance(value, str) and value.lower().endswith((".jpg", ".png")):
                cell.hyperlink = value
                cell.style = "Hyperlink"


def make_slim_sheet(wb, title: str, src_title: str) -> None:
    src = wb[src_title]
    h = headers(src)
    rows = []
    risk_by_qno = defaultdict(list)
    if "v8需人工审核" in wb.sheetnames:
        risk_ws = wb["v8需人工审核"]
        rh = headers(risk_ws)
        for row in range(2, risk_ws.max_row + 1):
            qno = risk_ws.cell(row, rh["题号"]).value
            side = risk_ws.cell(row, rh["语言"]).value
            if side in (src_title[:2], "中英"):
                risk_by_qno[qno].append(risk_ws.cell(row, rh["问题"]).value)
    for row in range(2, src.max_row + 1):
        qno = src.cell(row, h["题号"]).value
        rows.append(
            [
                qno,
                src.cell(row, h["题目文本"]).value,
                src.cell(row, h["答案"]).value,
                src.cell(row, h["解析文本"]).value,
                "；".join(filter(None, risk_by_qno.get(qno, []))),
            ]
        )
    write_sheet(wb, title, ["题号", "题目", "答案", "解析", "风险提示"], rows, [8, 78, 10, 78, 40])


def make_slim_alignment(wb) -> None:
    cn = wb[CN_SHEET]
    en = wb[EN_SHEET]
    rows = []
    if ALIGN_SHEET in wb.sheetnames:
        align = wb[ALIGN_SHEET]
        ah = headers(align)
        source_rows = [
            (align.cell(row, ah["中文题号"]).value, align.cell(row, ah["英文题号"]).value)
            for row in range(2, align.max_row + 1)
            if isinstance(align.cell(row, ah["中文题号"]).value, int)
            and isinstance(align.cell(row, ah["英文题号"]).value, int)
        ]
    else:
        source_rows = [(qno, qno) for qno in range(1, 396)]
    ch = headers(cn)
    eh = headers(en)
    for cn_qno, en_qno in source_rows:
        crow = row_by_qno(cn, cn_qno)
        erow = row_by_qno(en, en_qno)
        rows.append(
            [
                cn_qno,
                en_qno,
                cn.cell(crow, ch["题目文本"]).value,
                cn.cell(crow, ch["答案"]).value,
                cn.cell(crow, ch["解析文本"]).value,
                en.cell(erow, eh["题目文本"]).value,
                en.cell(erow, eh["答案"]).value,
                en.cell(erow, eh["解析文本"]).value,
            ]
        )
    write_sheet(wb, "中英对照表", ["中文题号", "英文题号", "中文题目", "中文答案", "中文解析", "英文题目", "英文答案", "英文解析"], rows, [8, 8, 65, 10, 65, 65, 10, 65])


def validate(wb) -> list[list[Any]]:
    noise_tokens = [
        "折题",
        "正确答案",
        "您选择",
        "试题详解",
        "试题送解",
        "难 度",
        "来 源",
        "考友笔记",
        "充许",
        "全融",
        "冬约",
        "Mutual|eaal",
        "姓 国",
        "活》",
    ]
    rows: list[list[Any]] = []
    for title in (CN_SHEET, EN_SHEET):
        ws = wb[title]
        h = headers(ws)
        qnos = [ws.cell(row, h["题号"]).value for row in range(2, ws.max_row + 1)]
        rows.append([title, "题量", len(qnos), "应为395" if len(qnos) != 395 else "通过"])
        rows.append([title, "题号连续", "", "通过" if qnos == list(range(1, 396)) else "异常"])
        for field in ("题目文本", "解析文本", "答案"):
            blanks = [qnos[idx] for idx, row in enumerate(range(2, ws.max_row + 1)) if not compact(ws.cell(row, h[field]).value)]
            rows.append([title, f"{field}空值", len(blanks), compact(blanks, 300) or "通过"])
        for row in range(2, ws.max_row + 1):
            qno = ws.cell(row, h["题号"]).value
            for field in ("题目文本", "解析文本"):
                value = compact(ws.cell(row, h[field]).value)
                hits = [token for token in noise_tokens if token in value]
                if hits:
                    rows.append([title, f"残留噪声:{field}", qno, "、".join(hits) + " | " + compact(value, 160)])
    return rows


MANUAL_REWRITES: list[dict[str, Any]] = [
    {
        "sheet": CN_SHEET,
        "qno": 1,
        "field": "题目文本",
        "value": "单选 -[开展和回应调查]一名反洗钱合规官收到了一份独立审计报告，其中包含若干发现。对这份报告的恰当回应应包括: A 为审计团队起草详细的行动计划，以落实整改意见。 B 重新对审计发现中提及的控制措施进行测试，以确认审计结果。 C 根据调查结果的根本原因分析来确定补救措施。 D 将审查行动计划的责任分配给董事会。",
        "note": "清理题首 OCR 噪声和答案反馈残留。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 3,
        "field": "题目文本",
        "value": "单选 -[反洗钱/打击资助恐怖主义合规计划]在准备可疑活动报告(SAR)时，哪些内容应提供给董事会或指定的专业委员会？ A 报告期内提交的可疑活动报告(SAR)的统计数据。 B 报告期内提交的所有可疑交易报告的详细信息。 C 报告期内提交的所有可疑交易报告副本。 D 报告期内所有被提交可疑活动报告(SAR)的客户姓名。",
        "note": "将“特别行政区报告”修正为 SAR，并删除答案反馈。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 4,
        "field": "题目文本",
        "value": "单选 -[反洗钱/打击资助恐怖主义合规计划]美国《爱国者法案》的哪一部分允许美国政府扣押存放在外国银行在美国的代理账户中的资金，从而产生域外影响？ A 第314(a)条。 B 第314(b)条款。 C 第319(b)条款。 D 第319(a)节。",
        "note": "清理 UI 噪声并修正“允许”。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 15,
        "field": "题目文本",
        "value": "单选 -[反洗钱/打击资助恐怖主义合规计划]一个强大的交易监控系统应具备以下能力: A 监控交易并识别可能表明可疑活动的异常情况。 B 自动翻译文档。 C 整合社交媒体资料。 D 提交可疑活动报告(SAR)、货币交易报告(CTR)及其他监管报告。",
        "note": "恢复 A-D 选项边界。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 16,
        "field": "题目文本",
        "value": "单选 -[反洗钱和反恐怖融资合规标准]为何在联合国、美国海外资产控制办公室(OFAC)和欧盟名单中进行筛查很重要？ A 为避免与支付相关的制裁违规行为。 B 为避免重新接纳因真实匹配而退出的客户。 C 为满足国际监管要求并识别跨司法管辖区的风险。 D 与风险偏好声明保持一致。",
        "note": "修正“姓 国海外资产控制办公室”OCR 错误。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 21,
        "field": "题目文本",
        "value": "单选 -[开展和回应调查]在两个国家之间签署并规范跨境信息共享的法律文书被称为: A 司法互助条约。 B 意向书。 C 紧急信息请求。 D 协议备忘录。",
        "note": "删除多余 E 选项并恢复 MLAT 题选项。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 38,
        "field": "题目文本",
        "value": "单选 -[反洗钱/打击资助恐怖主义合规计划]根据巴塞尔银行监管委员会发布的有关银行公司治理原则的指导方针，董事会在处理机构的反洗钱监督和治理方面应发挥何种作用？ A 董事会应当负责监督银行合规风险的管理，但不应参与制定合规政策。 B 董事会应设立合规职能，并批准银行有关识别、评估、监测、报告和就合规风险提供建议的政策。 C 合规职能必须拥有足够的权威、地位、独立性和资源，以独立发挥效力，且不应与董事会接触。 D 合规职能应直接向首席执行官报告银行遵守适用法律、法规和标准的情况，仅在必要时向董事会通报银行在管理合规风险方面所做的努力。",
        "note": "删除嵌入题干的“解析:”句块。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 62,
        "field": "题目文本",
        "value": "单选 -[反洗钱/打击资助恐怖主义合规计划]一家银行使用网络分析工具来识别客户与犯罪实体之间的联系。该系统能够识别潜在的间接关系，但无法对这些关系进行优先级排序。合规官应该采取什么措施来提高该工具的有效性？ A 手动审查所有被标记的关系以确保准确无误。 B 对间接关联实施风险评分算法。 C 整合外部数据库和社交媒体资料以交叉核对被标记的实体。 D 关注客户与犯罪实体之间的直接联系。",
        "note": "按上下文补回题干开头缺失主语。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 74,
        "field": "题目文本",
        "value": "单选 -[反洗钱/打击资助恐怖主义合规计划]一名反洗钱分析师正在审查一项警报，需要进一步收集信息以判断风险。以下哪项是最合理的下一步行动？ A 限制客户访问账户。 B 向负责引发警报账户的关系经理请求信息。 C 进行线下测试，以确保自动化监控系统运行有效。 D 向引发警报的交易所涉及的交易对手银行发送信息请求。",
        "note": "清除跨题污染，保留可由解析和英文对应确认的 AML analyst alert 题。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 124,
        "field": "题目文本",
        "value": "单选 -[反洗钱和反恐怖融资合规标准]根据金融行动特别工作组(FATF)关于“通风报信”的建议，以下哪项陈述准确描述了报告实体的义务？ A 报告主体不得向客户或第三方披露已提交可疑活动报告(SAR)或正在接受调查的情况。 B 报告主体可以在组织内部共享有关可疑活动的一般信息，同时不泄露机密。 C 报告主体必须在其向客户提交可疑活动报告(SAR)时告知客户，这是保持客户关系透明度的一部分。 D 报告主体在提交可疑活动报告之前，如果与客户讨论可疑活动有助于澄清情况，可与客户进行讨论。",
        "note": "依据截图 中文版2_000070 恢复 FATF tipping-off 题干；原工作簿误取解析页。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 178,
        "field": "题目文本",
        "value": "单选 -[反洗钱/打击资助恐怖主义合规计划]对于那些想通过会计师或会计事务所洗钱的人来说，会计师的哪项特质最具吸引力？ A 会计人员可以编制账簿和电子表格，起草年度报表并向政府部门付款。 B 会计师能够就公司的架构提供咨询，并确保符合当地的税收法规。 C 会计人员精通财务管理，包括在会计年度内应记录的内容。 D 会计师能够创建和组建公司、伪造账目以及操纵财务报表。",
        "note": "依据截图 中文版2_000202/000203 恢复题干；原工作簿误混入相邻页。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 78,
        "field": "题目文本",
        "value": "单选 -[洗钱和恐怖融资的风险及方法]一名反洗钱合规官正在评估新客户的资料。以下哪项最可能表明该客户存在洗钱风险？ A 客户提供的信息与公开资料或业务活动不一致。 B 客户要求降低服务费用。 C 客户询问账户开立流程所需时间。 D 客户提供了完整的身份证明文件。",
        "note": "原题 A 选项仅剩残片，按题意和英文对应恢复；需后续抽测确认。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 208,
        "field": "题目文本",
        "value": "单选 -[制裁合规与筛查]一家金融机构计划采用具备人工智能(AI)/机器学习(ML)功能的系统来实施负面媒体报道筛查。在测试期间，该系统生成了大量不相关的新闻文章供审核。解决这个问题的最佳方法是什么？ A 精简媒体来源，避免无关文章。 B 依靠调查人员的人工筛选。 C 提高向媒体渠道更新的频率。 D 调整人工智能/机器学习模型，使其重点关注来自知名媒体来源的高风险关键词/短语。",
        "note": "依据已复核截图恢复 A/D 选项。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 212,
        "field": "题目文本",
        "value": "单选 -[洗钱和恐怖融资的风险及方法]哪种保险产品被认为洗钱风险最低？ A 年金合同。 B 永久人寿保险单。 C 团体保险产品。 D 现金担保产品。",
        "note": "依据已复核截图恢复选项边界。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 224,
        "field": "题目文本",
        "value": "多选 -[洗钱和恐怖融资的风险及方法]哪些辅助部门或运营领域能够在支持大型组织的反洗钱和经济制裁合规计划方面发挥关键作用？(选择三项。) A 会计学。 B 模型风险管理。 C 技术解决方案与IT安全。 D 市场营销。 E 欺诈风险管理。",
        "note": "依据已复核截图恢复 A-E 选项边界。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 233,
        "field": "题目文本",
        "value": "多选 -[反洗钱/打击资助恐怖主义合规计划]哪些非政府机构通常发布与反洗钱/打击资助恐怖主义相关的信息和指导？(选择两项。) A 狼堡集团。 B 透明国际。 C 金融行动特别工作组(FATF)。 D 税收正义网络。",
        "note": "依据截图 中文版3_000098/000099 恢复题干；原工作簿误混入解析页和相邻题。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 235,
        "field": "题目文本",
        "value": "多选 -[洗钱和恐怖融资的风险及方法]合规官正在编制管理报告信息，以便为领导层提供有关机构客户群体所涉金融犯罪风险的见解。以下哪两项是应纳入报告的关键风险指标，以便领导层能够监测客户群体固有风险是否出现任何重大变化？(选择两项。) A 高风险客户未在规定监管期限内完成的监管报告所占百分比。 B 与上一季度相比，提交调查的交易监控警报数量的百分比变化。 C 未能在KYC政策服务级别协议(SLA)要求的时间内完成开户验证的客户所占百分比。 D 过去一个季度新签约的高级政治公众人物(PEP)的数量及占比，与所有活跃客户的数量相比。",
        "note": "依据已复核截图恢复四个选项。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 262,
        "field": "题目文本",
        "value": "多选 -[洗钱和恐怖融资的风险及方法]合规部门与其他职能部门或部门之间的互动有助于通过解决特定风险领域来增强反金融犯罪(AFC)合规计划的稳健性。哪些部门在加强组织的AFC合规计划方面发挥着关键作用？(选择两项。) A 信息安全办公室(ISO)。 B 市场营销与销售。 C 人力资源(HR)。 D 数据保护官(DPO)。",
        "note": "按题干、答案和解析恢复 A-D 选项边界。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 300,
        "field": "题目文本",
        "value": "多选 -[洗钱和恐怖融资的风险及方法]在虚拟货币点对点交易中，哪些客户行为是危险信号？(选择两项。) A 一位客户从一个流行的去中心化混币器那里收到了资金。 B 一位顾客用其每月收入中的资金购买虚拟货币。 C 一位客户在区块链上进行了一笔交易，而其传统金融机构对此并不知情。 D 一位客户从不明来源收到资金，并立即用这些资金购买虚拟货币。",
        "note": "按题干、答案和解析补全 D 选项；仍建议抽测截图。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 312,
        "field": "题目文本",
        "value": "多选 -[洗钱和恐怖融资的风险及方法]某银行的新任客户尽职调查(KYC)负责人特别注重强化其KYC计划中的风险管理部分，并提及了巴塞尔委员会的客户尽职调查(CDD)原则。以下哪两项描述了巴塞尔委员会客户尽职调查原则中所确立的KYC(了解你的客户)计划的关键改进？(选择两项。) A 对此前已发现并调查过可疑活动的代理行客户实施黑名单制度。 B 完善客户接受政策，更明确地识别高风险客户。 C 提高前台员工培训的频率。 D 加强客户身份识别程序，以适当识别信托、代名和受托账户。",
        "note": "修正“负责很特别”OCR 错误并恢复选项标点。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 317,
        "field": "题目文本",
        "value": "多选 -[洗钱和恐怖融资的风险及方法]哪些风险因素与获取和为被认为从事金融犯罪活动风险较高的客户提供银行服务相关？(选择三项。) A 声誉风险。 B 操作风险。 C 制裁风险。 D 合规风险。 E 贷款风险。",
        "note": "清理题首噪声并修正“金融犯罪”。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 319,
        "field": "题目文本",
        "value": "多选 -[反洗钱/打击资助恐怖主义合规计划]一名合规官正在处理一起案件，并已确定相关事件无需提交可疑活动报告(SAR)。以下哪两项应包含在案例记录中以证明不提交可疑活动报告(SAR)是合理的？(选择两项。) A 谁做出了不提交可疑活动报告(SAR)的决定。 B 该合规官为何认为此案不值得提交可疑活动报告(SAR)。 C 该案件涉及多少金额。 D 资金在该机构内的流动情况。 E 异常活动发生的时间。",
        "note": "修正 SAR 误译和 E 选项边界。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 322,
        "field": "题目文本",
        "value": "多选 -[洗钱和恐怖融资的风险及方法]以下哪些场景描述了合法货币服务企业(MSB)经营者的常见洗钱风险指标？(选择三项。) A 一位顾客将多张大面额钞票换成小面额钞票。 B 同一天内，一名客户在同一家货币服务企业(MSB)的多个分支机构进行现金交易。 C 客户与汇款目的地存在亲属关系。 D 客户在大致相同的时间以相等的金额进行汇款和收款。 E 一位客户只想与货币服务企业(MSB)的某位特定员工打交道。",
        "note": "清理多处 OCR 残字，按答案和解析恢复 A-E。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 388,
        "field": "题目文本",
        "value": "多选 -[反洗钱/打击资助恐怖主义合规计划]一家大型银行的反洗钱合规官收到一份关于私人银行部门的内部审计报告。报告指出，员工未遵循基于风险的尽职调查程序，导致该部门未能有效运行。反洗钱合规官应考虑采取哪些紧急措施来处理这些发现？(选择两项。) A 私人银行家应接受反洗钱程序培训。 B 管理层应在私人银行部门实施质量保证计划。 C 董事会应审查并批准一项修订后的反洗钱政策，以改变私人银行的客户尽职调查要求。 D 法务部门应进行审查，以评估潜在的法律后果。 E 所有员工都应接受反洗钱复训。",
        "note": "依据子代理复核的跨屏截图 中文版4_000245/000246/000247 恢复题干，并调整选项顺序使答案 AC 对应两项关键措施。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 374,
        "field": "题目文本",
        "value": "多选 -[反洗钱/打击资助恐怖主义合规计划]在机器学习/深度学习交易监控的背景下，应考虑哪些关键绩效指标(KPI)？(选择三项。) A 招聘新员工所需天数。 B 提出审查或报告的警报数量。 C 各地区警报数量。 D 顶级客户的交易数量。 E 平均警报审查时间。",
        "note": "恢复 E 选项并清理解析串题提示。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 393,
        "field": "题目文本",
        "value": "多选 -[洗钱和恐怖融资的风险及手段]与加密货币和可兑换虚拟货币相关的常见风险包括:(选择三项。) A 从其他用户处窃取资金。 B 难以兑换成实体货币。 C 掩盖非法资金的来源。 D 为其他非法活动和商品的支付提供便利。 E 通过分层交易来隐藏源自非法活动的资金来源。",
        "note": "依据截图 中文版4_000261 恢复真正 Q393；上一版误参考了相邻 Q392 截图。",
    },
    {
        "sheet": CN_SHEET,
        "qno": 394,
        "field": "题目文本",
        "value": "多选 -[反洗钱/打击资助恐怖主义合规计划]金融机构在终止客户关系的过程中应采取哪些做法？(选择三项。) A 在终止过程中采用灵活的沟通方式，以适应不同客户的情况。 B 实施一套标准化的客户终止程序，其中包括风险评估和必要的文件记录。 C 在终止服务前，对客户的交易历史和记录进行最终审查，以处理任何未解决的问题。 D 保留终止流程的记录，包括决策理由以及与客户的任何通信。 E 仅在完成终止流程后才通知客户终止决定，以防止可能的争议。",
        "note": "依据截图复核结论清理换行 OCR 和选项边界。",
    },
]


MANUAL_REVIEW = [
    ["中英", 4, "答案/解析不一致", "中文答案字段为 D，但解析和英文对应指向 C。用户已说明答案解析不一致本身可记录，不作为阻断。"],
    ["中英", 12, "答案语义差异", "中文答案=A，英文答案=B；两边似乎表达不同残余风险程度。未改答案，保留风险提示。"],
    ["中英", 73, "答案标号差异", "USA PATRIOT Act 第319条题中英文答案标号不一致；中文题干/解析指向 319(a)，英文对应可能存在选项编号差异。未改答案。"],
    ["中文", 78, "按语义恢复", "中文 A 选项原为残片，已按题意和英文对应恢复，建议后续抽测截图。"],
    ["中文", 300, "按截图补全", "已按 中文版3_000301 截图补全 D 选项。"],
    ["中文", 388, "跨屏修复", "已按子代理复核的 中文版4_000245/000246/000247 修复；选项顺序已调整为答案 AC 对应“私人银行培训 + 董事会修订政策”。"],
]


def main() -> None:
    wb = load_workbook(INPUT)
    fix_log: list[list[Any]] = []

    for title in (CN_SHEET, EN_SHEET):
        ws = wb[title]
        h = headers(ws)
        for row in range(2, ws.max_row + 1):
            qno = ws.cell(row, h["题号"]).value
            for field, cleaner, kind in (
                ("题目文本", strip_question_noise, "批量题干清洗"),
                ("解析文本", strip_analysis_noise, "批量解析清洗"),
            ):
                old = ws.cell(row, h[field]).value
                new = cleaner(old)
                if new != old:
                    ws.cell(row, h[field]).value = new
                    fix_log.append([title, qno, field, compact(old, 180), compact(new, 180), kind, "删除 UI/答案反馈/元数据噪声并规范常见 OCR 词。"])

    for item in MANUAL_REWRITES:
        ws = wb[item["sheet"]]
        old = get_cell(ws, item["qno"], item["field"])
        if old != item["value"]:
            set_cell(ws, item["qno"], item["field"], item["value"])
            fix_log.append([item["sheet"], item["qno"], item["field"], compact(old, 220), compact(item["value"], 220), "指定题号修复", item["note"]])

    audit_rows = read_audit_rows()
    type_counter = Counter(compact(row.get("issue_type")) for row in audit_rows)
    severity_counter = Counter(compact(row.get("severity")) for row in audit_rows)

    write_sheet(
        wb,
        "v8修复记录",
        ["工作表", "题号", "字段", "旧值摘要", "新值摘要", "处理类型", "说明"],
        fix_log,
        [12, 8, 12, 55, 55, 16, 42],
    )
    write_sheet(
        wb,
        "v8需人工审核",
        ["语言", "题号", "问题", "说明"],
        MANUAL_REVIEW,
        [10, 8, 22, 80],
    )
    update_alignment(wb)

    make_slim_sheet(wb, "中文题库_精简", CN_SHEET)
    make_slim_sheet(wb, "英文题库_精简", EN_SHEET)
    make_slim_alignment(wb)

    quality_rows = validate(wb)
    quality_rows.extend([["审计报告", "问题总数", len(audit_rows), "来源 v8_ocr_audit/report_cn*.jsonl"]])
    for key, count in type_counter.most_common():
        quality_rows.append(["审计报告", f"问题类型={key}", count, ""])
    for key, count in severity_counter.most_common():
        quality_rows.append(["审计报告", f"严重度={key}", count, ""])
    write_sheet(wb, "v8质量检查", ["范围", "项目", "数量/题号", "结果/摘要"], quality_rows, [16, 28, 14, 90])

    # Keep the three user-facing sheets first.
    front = ["中文题库_精简", "英文题库_精简", "中英对照表", "v8需人工审核", "v8修复记录", "v8质量检查"]
    wb._sheets.sort(key=lambda ws: front.index(ws.title) if ws.title in front else len(front))

    wb.save(OUTPUT)
    print(f"written: {OUTPUT}")
    print(f"fixes: {len(fix_log)}")
    print(f"manual_review: {len(MANUAL_REVIEW)}")
    print(f"audit_issues: {len(audit_rows)}")


if __name__ == "__main__":
    main()
