from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

out_dir = Path(r'D:\守正公司工作区\cams考试\教材、答疑记录、习题与参考文献\习题\v7习题\output_2s')
source = out_dir / 'semantic_aligned_cn_en.xlsx'
target = out_dir / 'semantic_aligned_cn_en_single_language_fixed.xlsx'
review = out_dir / 'single_language_visual_review'
cn_sheet = review / 'fine_cn' / 'sheets'
wb = load_workbook(source)
cn_ws, en_ws, align_ws = wb.worksheets[0], wb.worksheets[1], wb.worksheets[2]

def compact(v):
    return ' '.join(str(v or '').split())

def head(v, n=120):
    s = compact(v)
    return s if len(s) <= n else s[:n] + '...'

def row_by_qno(ws, qno):
    for r in range(2, ws.max_row + 1):
        try:
            if int(ws.cell(r, 1).value) == qno:
                return r
        except (TypeError, ValueError):
            pass
    raise KeyError((ws.title, qno))

def analysis_from_old(ws, qno):
    v = str(ws.cell(row_by_qno(ws, qno), 7).value or '')
    i = v.find('试题详解')
    return v[i:] if i >= 0 else v

CN101 = '''-[洗钱和恐怖融资的风险及方法] 欧洲一家银行的合规官正在调查该行过去两年中为一家客户开设的账户。根据行记录显示，该公司的主要经济活动是进出口石化产品。在一年的时间里，该账户的交易额已超过5亿美元，先是收到来自也门供应商的多笔大额汇入款项，随后又向阿塞拜疆的交易对手汇出款项。如果在调查中发现以下哪个因素，合规官应当最为关注？
A 媒体搜索显示，三年前该客户曾被指控行为不当。
B 欧盟银行认为交易涉及的起始国和目的国属于高风险国家。
C 该客户的活动包括涉及多个交易对手的多笔跨境交易。
D 欧盟银行未收到有关电汇的发起人或受益人信息。
E 该客户专门从事石化产品的进出口业务。
正确答案：D'''
CN132 = '''-[反洗钱/打击资助恐怖主义合规计划] 谁应对批准金融机构与政治公众人物的关系承担最终责任？
A 知识产权分析师
B 高级管理层
C 客户经理
D 增强型尽职调查合规官
正确答案：B'''
CN132_ANALYSIS = '''试题详解 原解析
在金融机构的反洗钱和打击资助恐怖主义合规计划中，与政治公众人物建立关系需高度谨慎，因其可能涉及更高洗钱风险。根据多个技术资料显示，高级管理层对金融机构的重大决策和风险管理承担最终责任，包括批准与政治公众人物的关系。知识产权分析师、客户经理和增强型尽职调查合规官虽在不同环节发挥作用，但均不承担最终批准责任。因此，正确答案为B。'''
CN164 = '''-[洗钱和恐怖融资的风险及方法] 以下哪一项是非政府组织（NGO）打击洗钱活动的常见策略？
A 协助金融情报机构（FIU）分析可疑活动报告（SAR）
B 直接在法庭上起诉洗钱者
C 向各国政府提供财政援助，以加强其反洗钱工作
D 提高对洗钱问题及其后果的认识
正确答案：D'''
CN168 = '''-[反洗钱和反恐怖融资合规标准] 以下哪一项是公私合作（PPP）成功的范例？
A 金融行动特别工作组（FATF）
B 澳大利亚交易报告和分析中心金融情报联盟
C 埃格蒙特集团
D 狼堡集团
正确答案：B'''
CN378 = '''-[洗钱和恐怖融资的风险及方法] 与使用“哈瓦拉”相关的金融犯罪风险可能源于：（选择两项。）
A 难以追踪交易的发起方、接收方和资金来源。
B 由第三方程序管理员进行远程身份验证。
C 高级政治人物的频繁使用。
D 用于跨境交易的非正式网络，这些交易游离于正规银行体系之外。
E 交易被退回的风险增加。
正确答案：AD'''
EN124 = '''-[Compliance Standards for AML and CFT] Under the Egmont Group Principles, information exchange among financial intelligence units (FIUs) should be conducted:
A Without the expectation of reciprocity on how the information will be used.
B Only if the status of the foreign FIU is related to law enforcement.
C Freely, spontaneously, and upon request, on the basis of reciprocity.
D With set limits on the amount of financial and administrative information provided.
Correct answer: C'''
EN221 = '''-[Risks and Methods of Money Laundering and Terrorist Financing] A periodic review has been completed for an existing virtual asset service provider (VASP) customer. Which of the following are indicators of potential money laundering? (Select Three.)
A Use of shell companies for deposits and withdrawals into the VASP
B Using a peer-to-peer network to reduce costs associated with server maintenance and data storage
C Rapid market fluctuations resulting in quick changes in the value of underlying assets
D Frequent use of mixers and tumblers for holdings and transactions
E Receiving funds from countries known for weak money laundering regulations and frameworks
Correct answer: ADE'''
EN284 = '''-[Risks and Methods of Money Laundering and Terrorist Financing] Which criminal activities could possibly serve as a predicate offense to financial crimes or money laundering activity? (Select Three.)
A Assault
B Arson
C Bribery or corruption
D Fraud
E Organized crime or racketeering
Correct answer: CDE'''
EN308 = '''-[AML/CFT Compliance Programs] Which of the following AML/CFT arrangements should a bank have in place at minimum before opening for business? (Choose three.)
A Independent compliance testing
B Qualified and knowledgeable staff
C Third-party screening software
D An automated monitoring system
E Written policy and procedures
Correct answer: ABE'''

fixes = [
    dict(lang='CN', qno=101, answer='D', question=CN101, analysis=analysis_from_old(cn_ws,101), time='0-2', frames=2, evidence=str(cn_sheet/'cn101_v2_01.jpg'), note='细帧补全题干上半部分。'),
    dict(lang='CN', qno=132, answer='B', question=CN132, analysis=CN132_ANALYSIS, time='173-176', frames=4, evidence=str(cn_sheet/'cn132_03.jpg'), note='原行混入上一题解释，按132/395视频页重建。'),
    dict(lang='CN', qno=164, answer='D', question=CN164, analysis=analysis_from_old(cn_ws,164), time='339-341', frames=3, evidence=str(cn_sheet/'cn164_02.jpg'), note='细帧补全题干上半部分。'),
    dict(lang='CN', qno=168, answer='B', question=CN168, analysis=analysis_from_old(cn_ws,168), time='353-355', frames=3, evidence=str(cn_sheet/'cn168_03.jpg'), note='细帧补全PPP题干。'),
    dict(lang='CN', qno=378, answer='AD', question=CN378, analysis=analysis_from_old(cn_ws,378), time='433-435', frames=3, evidence=str(cn_sheet/'cn378_02.jpg'), note='细帧补全哈瓦拉题干。'),
    dict(lang='EN', qno=124, answer='C', question=EN124, analysis='', time='144-146', frames=2, evidence=str(review/'contexts'/'en_124_context.jpg'), note='原行混入123题解释；解析帧未完整抽取，先清空错误解析。'),
    dict(lang='EN', qno=221, answer='ADE', question=EN221, analysis='', time='112-114', frames=2, evidence=str(review/'fine_en221_b'/'en221b_001.jpg'), note='原行混入220题解释；解析帧未完整抽取，先清空错误解析。'),
    dict(lang='EN', qno=284, answer='CDE', question=EN284, analysis='', time='516-518', frames=2, evidence=str(review/'fine_en284'/'en284_012.jpg'), note='原行内容属于相邻283题，按284/395视频页重建；解析待后续复核。'),
    dict(lang='EN', qno=308, answer='ABE', question=EN308, analysis=analysis_from_old(en_ws,308), time='45-46', frames=2, evidence=str(review/'fine_en308'/'en308_020.jpg'), note='原行混入307题解释，按308/395视频页重建。'),
    dict(lang='EN', qno=123, answer='A', question=None, analysis=None, time=None, frames=None, evidence=str(review/'contexts'/'en_124_context.jpg'), note='相邻边界题答案补齐：视频123/395 tipping-off题答案A。'),
    dict(lang='EN', qno=220, answer='CD', question=None, analysis=None, time=None, frames=None, evidence=str(review/'contexts'/'en_221_context.jpg'), note='相邻边界题答案补齐：视频220/395 charity/NPO题答案CD。'),
    dict(lang='EN', qno=283, answer='ABE', question=None, analysis=None, time=None, frames=None, evidence=str(review/'contexts'/'en_284_context.jpg'), note='相邻边界题答案补齐：视频283/395 laws/regulations题答案ABE。'),
]

def apply_to_sheet(ws, f):
    r = row_by_qno(ws, f['qno'])
    old = dict(row=r, old_answer=ws.cell(r,3).value, old_question=head(ws.cell(r,6).value), old_analysis=head(ws.cell(r,7).value))
    ws.cell(r,3).value = f['answer']
    if f['time'] is not None: ws.cell(r,4).value = f['time']
    if f['frames'] is not None: ws.cell(r,5).value = f['frames']
    if f['question'] is not None: ws.cell(r,6).value = f['question']
    if f['analysis'] is not None: ws.cell(r,7).value = f['analysis']
    if f['evidence']:
        for c in (9,10):
            ws.cell(r,c).value = f['evidence']
            ws.cell(r,c).hyperlink = f['evidence']
            ws.cell(r,c).style = 'Hyperlink'
    for c in range(1, ws.max_column+1): ws.cell(r,c).alignment = Alignment(wrap_text=True, vertical='top')
    return old

def sync_align(f):
    cn = f['lang'] == 'CN'
    qcol, acol, qtext, ancol, tcol, scol = (9,11,13,15,19,21) if cn else (10,12,14,16,20,22)
    n = 0
    for r in range(2, align_ws.max_row+1):
        try:
            if int(align_ws.cell(r,qcol).value) != f['qno']:
                continue
        except (TypeError, ValueError):
            continue
        align_ws.cell(r,acol).value = f['answer']
        if f['question'] is not None: align_ws.cell(r,qtext).value = f['question']
        if f['analysis'] is not None: align_ws.cell(r,ancol).value = f['analysis']
        if f['time'] is not None: align_ws.cell(r,tcol).value = f['time']
        if f['evidence']:
            align_ws.cell(r,scol).value = f['evidence']
            align_ws.cell(r,scol).hyperlink = f['evidence']
            align_ws.cell(r,scol).style = 'Hyperlink'
        for c in range(1, align_ws.max_column+1): align_ws.cell(r,c).alignment = Alignment(wrap_text=True, vertical='top')
        n += 1
    return n

records=[]
for f in fixes:
    ws = cn_ws if f['lang']=='CN' else en_ws
    old = apply_to_sheet(ws, f)
    n = sync_align(f)
    if f['analysis'] == '': ast = '清空错误/相邻题解析'
    elif f['analysis'] is None: ast = '未改'
    else: ast = '同步修复'
    records.append([f['lang'], f['qno'], old['row'], old['old_answer'], f['answer'], old['old_question'], head(f['question']), ast, f['evidence'], f['note'], n])

log_title='单语视频修复记录'
if log_title in wb.sheetnames: del wb[log_title]
log = wb.create_sheet(log_title)
headers=['语言','题号','工作表行','旧答案','新答案','旧题干摘要','新题干摘要','解析处理','证据截图','备注','语义对齐同步行数']
log.append(headers)
for rec in records: log.append(rec)
for cell in log[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill('solid', fgColor='D9EAF7')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
widths=[10,10,10,10,10,45,45,22,72,50,16]
for i,w in enumerate(widths,1): log.column_dimensions[get_column_letter(i)].width = w
for r in range(2, log.max_row+1):
    for c in range(1, log.max_column+1): log.cell(r,c).alignment = Alignment(wrap_text=True, vertical='top')
    if log.cell(r,9).value:
        log.cell(r,9).hyperlink = log.cell(r,9).value
        log.cell(r,9).style='Hyperlink'
log.freeze_panes='A2'
wb.save(target)
print(target)
print(len(records))
